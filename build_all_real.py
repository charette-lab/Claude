"""Rebuild the framework + both analyses using the REAL data from Compensation.xlsx 'Static' sheet."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
import matplotlib; matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np

WB="/home/user/Claude/Performance_Fee_Framework.xlsx"
PNG_ROLE="/home/user/Claude/Take_share_by_AUM.png"
PNG_PP="/home/user/Claude/Take_share_per_person.png"

# ---- REAL data from Compensation.xlsx > Static ----
RET=0.12; FEE=0.20; TAKE=0.60; MGMT=0.015; THR=300; LOAD=1.62
scn_aum=[20,300,500,1000,2000,5000]
# (role, salary<300, salary>300, weight, headcount per scenario)
roles=[
 ("CIO",       0.20,0.30,16, [1,1,1,1,1,1]),
 ("PM",        0.13,0.20, 8, [2,2,3,3,4,4]),
 ("Analyst",   0.00,0.20,2.5,[0,2,3,4,8,8]),
 ("CEO",       0.15,0.30, 8, [1,1,1,1,1,1]),
 ("CFO",       0.10,0.15, 3, [0.2,0.2,0.2,1,1,1]),
 ("COO",       0.12,0.20, 3, [0.5,0.5,0.5,1,1,1]),
 ("IR",        0.12,0.12, 1, [1,2,2,2,2,2]),
 ("Support",   0.12,0.12, 1, [0.5,0.5,0.5,1,1,1]),
 ("Execution", 0.00,0.20, 2, [0,0,0,1,1,1]),
 ("Risk",      0.00,0.20, 2, [0,0,0,1,1,1]),
 ("Compliance",0.00,0.20, 1, [0,0,0,1,1,1]),
]
nR=len(roles)

# ---------- styles ----------
NAVY="1F3864";BLUE="2E5496";LIGHT="D6E0F0";LIGHTER="EAF0FA";GREY="F2F2F2";GOLD="FFF2CC";GREEN="E2EFDA"
title_f=Font(size=16,bold=True,color="FFFFFF");sub_f=Font(size=10,italic=True,color="FFFFFF")
hdr_f=Font(size=10,bold=True,color="FFFFFF");hdr_s=Font(size=9,bold=True,color="FFFFFF")
sec_f=Font(size=12,bold=True,color=NAVY);bold=Font(size=10,bold=True);norm=Font(size=10);small=Font(size=9)
ital=Font(size=9,italic=True,color="595959");inp_f=Font(size=10,bold=True,color="C00000")
wf=lambda c:Font(size=10,bold=True,color=c)
fnavy=PatternFill("solid",fgColor=NAVY);fblue=PatternFill("solid",fgColor=BLUE);fhdr=PatternFill("solid",fgColor=BLUE)
flight=PatternFill("solid",fgColor=LIGHT);flighter=PatternFill("solid",fgColor=LIGHTER)
fgrey=PatternFill("solid",fgColor=GREY);fgold=PatternFill("solid",fgColor=GOLD);fgreen=PatternFill("solid",fgColor=GREEN)
ctr=Alignment("center","center",wrap_text=True);lft=Alignment("left","center",wrap_text=True);rgt=Alignment("right","center")
thin=Side("thin",color="BFBFBF");box=Border(thin,thin,thin,thin)
L=openpyxl.utils.get_column_letter
SC=["H","I","J","K","L","M"]
MUSD='#,##0.000';MM='#,##0.0';PCT='0.0%';PCT2='0.00%';NUM='0.0';W='0.0'

wb=openpyxl.Workbook(); wb.remove(wb.active)

def st(ws,coord,v,f=norm,fill=None,al=None,fmt=None,bd=False):
    c=ws[coord];c.value=v;c.font=f
    if fill:c.fill=fill
    if al:c.alignment=al
    if fmt:c.number_format=fmt
    if bd:c.border=box
    return c

# ================================================================= STATIC
ws=wb.create_sheet("Static")
ws.merge_cells("A1:M1");st(ws,"A1","Performance Fee Distribution Framework",title_f,fnavy,ctr)
ws.merge_cells("A2:M2");st(ws,"A2","Athanase Industrial Partner — real roles, salaries & performance weights (source: Compensation.xlsx › Static). Amounts in millions.",sub_f,fnavy,ctr)
ws.row_dimensions[1].height=26;ws.row_dimensions[2].height=16

st(ws,"A4","1.  Global assumptions  (edit the gold cells)",sec_f)
gi=[(5,"Performance-fee rate (of profits)",FEE,PCT),
    (6,"Firm take rate — share of perf-fee to the comp pool",TAKE,PCT),
    (7,"Assumed gross return on AUM",RET,PCT),
    (8,"Management fee (of AUM)",MGMT,PCT),
    (9,"AUM tier threshold (millions)",THR,'#,##0'),
    (10,"Employer cost loading on salary",LOAD,'0.00')]
for r,lab,val,fmt in gi:
    ws.merge_cells(f"A{r}:E{r}");st(ws,f"A{r}",lab,norm,fgrey,lft,bd=True)
    st(ws,f"F{r}",val,inp_f,fgold,ctr,fmt,bd=True)
st(ws,"A11","Comp pool = AUM × return × fee rate × take rate.  (The firm retains the other 40% of the performance fee.)",ital)

st(ws,"A13","2.  AUM scenarios & performance-fee economics  (millions)",sec_f)
st(ws,"A14","Metric",hdr_f,fhdr,lft,bd=True)
for i,col in enumerate(SC):st(ws,f"{col}14",f"Scenario {i+1}",hdr_f,fhdr,ctr,bd=True)
st(ws,"A15","AUM (millions)",bold,flight,lft,bd=True)
for col,a in zip(SC,scn_aum):st(ws,f"{col}15",a,inp_f,fgold,ctr,'#,##0',bd=True)
st(ws,"A16","Gross profit  (= AUM × return)",norm,flighter,lft,bd=True)
for col in SC:st(ws,f"{col}16",f"={col}15*$F$7",norm,None,ctr,MM,bd=True)
st(ws,"A17","Performance fee  (= profit × fee rate)",norm,flighter,lft,bd=True)
for col in SC:st(ws,f"{col}17",f"={col}16*$F$5",norm,None,ctr,MM,bd=True)
st(ws,"A18","Comp pool  (= perf fee × take rate)",bold,fgreen,lft,bd=True)
for col in SC:st(ws,f"{col}18",f"={col}17*$F$6",bold,fgreen,ctr,MM,bd=True)

st(ws,"A20","3.  Organisation, weights & headcount per scenario",sec_f)
st(ws,"A20","3.  Organisation, weights & headcount",sec_f)
st(ws,"B20","Salary\n< $300m",hdr_s,fhdr,ctr,bd=True)
st(ws,"C20","Salary\n≥ $300m",hdr_s,fhdr,ctr,bd=True)
st(ws,"D20","Performance\nweight",hdr_s,fhdr,ctr,bd=True)
st(ws,"E20","",hdr_s,fhdr,ctr,bd=True);st(ws,"F20","",hdr_s,fhdr,ctr,bd=True)
st(ws,"G20","Headcount →",hdr_s,fhdr,ctr,bd=True)
for i,col in enumerate(SC):st(ws,f"{col}20",f"@ ${scn_aum[i]:,}m",hdr_s,fhdr,ctr,bd=True)
ws.row_dimensions[20].height=30
first=21
for idx,(name,lo,hi,wt,hc) in enumerate(roles):
    r=first+idx;fr=flighter if idx%2 else None
    st(ws,f"A{r}",name,norm,fr,lft,bd=True)
    st(ws,f"B{r}",lo,inp_f,fgold,ctr,MUSD,bd=True)
    st(ws,f"C{r}",hi,inp_f,fgold,ctr,MUSD,bd=True)
    st(ws,f"D{r}",wt,inp_f,fgold,ctr,W,bd=True)
    st(ws,f"E{r}","",norm,fr,bd=True);st(ws,f"F{r}","",norm,fr,bd=True);st(ws,f"G{r}","",norm,fr,bd=True)
    for col,n in zip(SC,hc):st(ws,f"{col}{r}",n,inp_f,fgold,ctr,'0.0',bd=True)
tot=first+nR  # row for totals
st(ws,f"A{tot}","Total performance units",bold,fblue,lft,bd=True)
for c in "BCDEFG":st(ws,f"{c}{tot}","",bold,fblue,bd=True)
st(ws,f"G{tot}","Σ weight×HC →",bold,fblue,rgt,bd=True)
for col in SC:st(ws,f"{col}{tot}",f"=SUMPRODUCT($D${first}:$D${tot-1},{col}{first}:{col}{tot-1})",wf("FFFFFF"),fblue,ctr,'#,##0.0',bd=True)
uv=tot+1
st(ws,f"A{uv}","Value per performance unit (millions)",bold,fgreen,lft,bd=True)
for c in "BCDEFG":st(ws,f"{c}{uv}","",bold,fgreen,bd=True)
st(ws,f"G{uv}","pool ÷ units →",bold,fgreen,rgt,bd=True)
for col in SC:st(ws,f"{col}{uv}",f"=IF({col}{tot}=0,0,{col}18/{col}{tot})",bold,fgreen,ctr,MUSD,bd=True)
pr=uv+1
st(ws,f"A{pr}","Total fixed payroll, loaded (millions)",norm,fgrey,lft,bd=True)
for c in "BCDEFG":st(ws,f"{c}{pr}","",norm,fgrey,bd=True)
for col in SC:st(ws,f"{col}{pr}",f"=IF({col}15<$F$9,SUMPRODUCT($B${first}:$B${tot-1},{col}{first}:{col}{tot-1}),SUMPRODUCT($C${first}:$C${tot-1},{col}{first}:{col}{tot-1}))*$F$10",norm,fgrey,ctr,MM,bd=True)
hcrow=pr+1
st(ws,f"A{hcrow}","Total headcount",norm,fgrey,lft,bd=True)
for c in "BCDEFG":st(ws,f"{c}{hcrow}","",norm,fgrey,bd=True)
for col in SC:st(ws,f"{col}{hcrow}",f"=SUM({col}{first}:{col}{tot-1})",norm,fgrey,ctr,'0.0',bd=True)

# downstream output tables (per person $ ; role share % ; per-person share %)
def block(sr,title,fn,fmt):
    st(ws,f"A{sr}",title,sec_f);hr=sr+1
    st(ws,f"A{hr}","Role",hdr_s,fhdr,lft,bd=True);st(ws,f"D{hr}","Weight",hdr_s,fhdr,ctr,bd=True)
    for i,col in enumerate(SC):st(ws,f"{col}{hr}",f"@ ${scn_aum[i]:,}m",hdr_s,fhdr,ctr,bd=True)
    for idx in range(nR):
        r=hr+1+idx;src=first+idx;fr=flighter if idx%2 else None
        st(ws,f"A{r}",f"=A{src}",norm,fr,lft,bd=True);st(ws,f"D{r}",f"=D{src}",norm,fr,ctr,W,bd=True)
        for col in SC:st(ws,f"{col}{r}",fn(col,src,r),norm,fr,ctr,fmt,bd=True)
    return hr+1+nR
eA=block(hcrow+2,"4.  Performance payout PER PERSON, by scenario  (millions — absolute, grows with AUM)",
         lambda col,src,r:f"=$D{src}*{col}${uv}",MUSD)
eB=block(eA+2,"5.  Each role's SHARE of the take  (% — falls as the firm grows)",
         lambda col,src,r:f"=IF({col}${tot}=0,0,{col}{src}*$D{src}/{col}${tot})",PCT)
eC=block(eB+2,"6.  Share of the take PER PERSON  (% — one individual's slice)",
         lambda col,src,r:f"=$D{src}/{col}${tot}",PCT2)

ws.column_dimensions["A"].width=30
for c in "BCD":ws.column_dimensions[c].width=12
ws.column_dimensions["E"].width=3;ws.column_dimensions["F"].width=11;ws.column_dimensions["G"].width=13
for col in SC:ws.column_dimensions[col].width=12
ws.sheet_view.showGridLines=False;ws.freeze_panes="B21"

# ================================================================= SWEEP analyses
aum_grid=list(range(100,2001,100))
def interp(hc,A):
    A=max(scn_aum[0],min(scn_aum[-1],A))
    for k in range(len(scn_aum)-1):
        if scn_aum[k]<=A<=scn_aum[k+1]:
            t=(A-scn_aum[k])/(scn_aum[k+1]-scn_aum[k]);return hc[k]+t*(hc[k+1]-hc[k])
    return hc[-1]
units=[sum(wt*interp(hc,A) for(_,_,_,wt,hc) in roles) for A in aum_grid]
role_share=np.array([[ (wt*interp(hc,A))/units[j] for j,A in enumerate(aum_grid)] for(_,_,_,wt,hc) in roles])
pp_share=np.array([[ wt/units[j] for j in range(len(aum_grid))] for(_,_,_,wt,_) in roles])

def sweep_sheet(name,title,sub,mat,fmt,denom_label):
    ws=wb.create_sheet(name)
    n=len(aum_grid)
    ws.merge_cells(f"A1:{L(1+n)}1");st(ws,"A1",title,Font(size=14,bold=True,color="FFFFFF"),fnavy,ctr)
    ws.merge_cells(f"A2:{L(1+n)}2");st(ws,"A2",sub,sub_f,fnavy,ctr);ws.row_dimensions[1].height=22
    hr=4;st(ws,f"A{hr}","AUM (millions) →",hdr_s,fhdr,lft,bd=True)
    for j,A in enumerate(aum_grid):st(ws,f"{L(2+j)}{hr}",A,hdr_s,fhdr,ctr,'#,##0',bd=True)
    f0=hr+1
    for i,(nm,_,_,wt,_) in enumerate(roles):
        r=f0+i;fr=flighter if i%2 else None
        st(ws,f"A{r}",f"{nm}  (wt {wt:g})",norm,fr,lft,bd=True)
        for j in range(n):st(ws,f"{L(2+j)}{r}",float(mat[i,j]),norm,fr,ctr,fmt,bd=True)
    tr=f0+nR;st(ws,f"A{tr}",denom_label,hdr_s,fhdr,lft,bd=True)
    for j in range(n):
        if denom_label.startswith("Total units"):
            st(ws,f"{L(2+j)}{tr}",round(units[j],1),hdr_s,fhdr,ctr,'#,##0.0',bd=True)
        else:
            col=L(2+j);st(ws,f"{col}{tr}",f"=SUM({col}{f0}:{col}{tr-1})",hdr_s,fhdr,ctr,'0%',bd=True)
    ch=LineChart();ch.title=title;ch.style=2;ch.height=11;ch.width=26
    ch.y_axis.numFmt=fmt;ch.x_axis.title="AUM (millions)";ch.y_axis.title="% of take"
    d=Reference(ws,min_col=1,min_row=f0,max_row=tr-1,max_col=1+n)
    cats=Reference(ws,min_col=2,min_row=hr,max_col=1+n,max_row=hr)
    ch.add_data(d,titles_from_data=True,from_rows=True);ch.set_categories(cats);ws.add_chart(ch,f"A{tr+2}")
    ws.column_dimensions["A"].width=26
    for j in range(n):ws.column_dimensions[L(2+j)].width=7
    ws.sheet_view.showGridLines=False;ws.freeze_panes="B5"

sweep_sheet("Take by AUM","Each role's share of the take as AUM grows $100m → $2bn",
            "Share = role weight × headcount ÷ total weighted units. Real weights from Compensation.xlsx. Columns sum to 100%.",
            role_share,PCT,"Total (check)")
sweep_sheet("Take per Person","Per-person share of the take as AUM grows $100m → $2bn",
            "Per-person share = role weight ÷ total weighted units. One individual's slice; declines for every role as the firm hires.",
            pp_share,PCT2,"Total units (denominator)")
wb.save(WB)

# ================================================================= PNGs
cmap=plt.get_cmap("tab20");cols=[cmap(i) for i in range(nR)]
# role share stacked area
fig,ax=plt.subplots(figsize=(13,7))
ax.stackplot(aum_grid,[role_share[i]*100 for i in range(nR)],labels=[r[0] for r in roles],colors=cols,alpha=0.9,edgecolor="white",linewidth=0.3)
ax.set_xlim(100,2000);ax.set_ylim(0,100);ax.set_xlabel("AUM (millions)");ax.set_ylabel("Share of the take (%)")
ax.set_title("How the performance-fee take is split by role, $100m → $2bn  (real weights)",fontsize=13,fontweight="bold")
ax.legend(loc="center left",bbox_to_anchor=(1.01,0.5),fontsize=9,frameon=False);ax.set_xticks(range(100,2001,200))
plt.tight_layout();plt.savefig(PNG_ROLE,dpi=130,bbox_inches="tight")
# per-person line
fig,ax=plt.subplots(figsize=(13,7))
for i,(nm,_,_,_,_) in enumerate(roles):ax.plot(aum_grid,pp_share[i]*100,color=cols[i],lw=2,marker="o",ms=3,label=f"{nm} (wt {roles[i][3]:g})")
ax.set_xlim(100,2000);ax.set_ylim(0,None);ax.set_xlabel("AUM (millions)");ax.set_ylabel("Share of the take PER PERSON (%)")
ax.set_title("Per-person share of the performance-fee take, by role ($100m → $2bn)  (real weights)",fontsize=13,fontweight="bold")
ax.grid(True,alpha=0.3);ax.set_xticks(range(100,2001,200));ax.legend(loc="center left",bbox_to_anchor=(1.01,0.5),fontsize=9,frameon=False)
plt.tight_layout();plt.savefig(PNG_PP,dpi=130,bbox_inches="tight")

# console reconciliation
print("Sheets:",wb.sheetnames)
print("\nReconcile unit value at real scenarios (should match Static H34:M34):")
for col,A in zip(SC,scn_aum):
    u=sum(wt*hc[scn_aum.index(A)] for(_,_,_,wt,hc) in roles)
    pool=A*RET*FEE*TAKE
    print(f"  AUM {A:>5}: units={u:>6.1f}  unit_val={pool/u:.5f}")
print("\nPer-person share (%) at selected AUM:")
pick=[100,300,500,1000,2000];idx=[aum_grid.index(a) for a in pick]
print("Role".ljust(14)+"wt".rjust(5)+"".join(f"{a:>8}" for a in pick))
for i,(nm,_,_,wt,_) in enumerate(roles):
    print(nm.ljust(14)+f"{wt:>5g}"+"".join(f"{pp_share[i,j]*100:>7.2f}%" for j in idx))
