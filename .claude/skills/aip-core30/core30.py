#!/usr/bin/env python3
"""aip-core30 — build the Constrained Quality Compounder Index from an aip-value
screener.

Values every company in the sheet with the **aip-value** engine (per-company WACC
from an equity hurdle, the de-risking lever-glide, and the distributed-cash equity
IRR), then assembles a portfolio under the risk framework:

  Phase 1  Universe & screens : market-cap floor, moat floor, exclude financials,
                                TAM-exhaustion reinvestment-deceleration screen.
  Phase 2  11 risk tags       : read from the sheet's tag columns (scored 1-4);
                                "material exposure" = score >= --tag-material (4).
  Phase 3  Construction       : rank by an equity-return hurdle, then greedily
                                select N names such that no single risk tag is held
                                by more than --tag-cap stocks (the 20% rule: at
                                3.33%/name, 6 stocks = 19.98%). Equal-weighted.

The portfolio is ranked by ER at --rank-re (default 12%); ER at --alt-re (default
7%) is shown alongside. Pass --er-gate to additionally require ER at the ranking
hurdle exceed a threshold (e.g. --er-gate 0 = "positive at 12%").

Usage:
  python3 core30.py <file.xlsx> [--rank-re 0.12] [--alt-re 0.07] [--mcap-floor 5e8]
                    [--n 30] [--tag-cap 6] [--tag-material 4] [--moat-min 7]
                    [--er-gate FLOAT] [--country-base "USD=0.0406,JPY=0.0265,..."]
                    [--no-lever-glide] [--csv out.csv]
"""
import argparse
import csv
import importlib.util
import os
import sys


def load_engine():
    """Import the aip-value engine (roiic_dcf.py)."""
    here = os.path.dirname(os.path.abspath(__file__))
    for path in (os.path.join(here, "..", "aip-value", "roiic_dcf.py"),
                 os.path.expanduser("~/.claude/skills/aip-value/roiic_dcf.py")):
        if os.path.exists(path):
            spec = importlib.util.spec_from_file_location("aipval", path)
            mod = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(mod)
            return mod
    sys.exit("Could not find the aip-value engine (roiic_dcf.py). Install aip-value "
             "alongside this skill (../aip-value/ or ~/.claude/skills/aip-value/).")


# The 11 standardized risk tags (exact sheet headers) and short labels.
TAGS = ["Customer & End-Market Demand", "Literal Customer Concentration",
        "Macro & Factor Sensitivity", "Currency & FX Exposure",
        "Geographic & Supply Chain Reliance", "Capital Structure & Funding Risk",
        "Technological & Platform Dependency",
        "Product Lifecycle & Intellectual Property Risk",
        "Regulatory & Compliance Risk", "Input Cost Structure",
        "Key Person & Management Risk"]
SHORT = ["Cust", "CustConc", "Macro", "FX", "Geo", "CapStr", "TechDep", "IP",
         "Reg", "InpCost", "KeyPpl"]


def equity_er(m, re, nopat0, roiic0, rr0, mktcap, netdebt, gross, tax, country,
              ind, moat, sales0, overrides, crp_over, lever_glide=True):
    """Expected equity return at hurdle `re` via the aip-value engine."""
    cbase, _ = m.currency_base(country, overrides)
    t = tax if (tax is not None and tax < 1) else 0.25
    rd = m.synthetic_rd(nopat0 / (1 - t), gross, mktcap, cbase)[0]
    crp = m.country_risk_premium(country, crp_over)
    if lever_glide:
        L = m.target_leverage(ind)
        rd_m = m.mature_cost_of_debt(L, cbase, mktcap)[0]
        w0 = min(max(m.firm_wacc_taxed(re, rd, mktcap, netdebt, t) + crp, 0.04), re + crp)
        wm = min(max(m.mature_wacc(re, rd_m, L, t)[0] + crp, 0.04), 0.25)
        wp, wt = m.wacc_glide(w0, wm, *_split(m, moat)), wm
        lever_L = L
    else:
        w0 = min(max(m.firm_wacc(re, rd, mktcap, netdebt) + crp, 0.04), 0.25)
        wp, wt, lever_L = None, None, None
    n1, n2 = _split(m, moat)
    base = m.base_rate_for(ind)[0]
    mp = m.moat_to_cap_persistence(moat)
    phi = mp[1] if mp else 0.75
    res = m.value_company(nopat0, roiic0, rr0, w0, 0.025, n1, n2, phi, base,
                          sales0=sales0, gics_industry=ind,
                          wacc_path=wp, wacc_terminal=wt)
    val, method = m.equity_return(res, mktcap, netdebt, 5, tax=t, lever_L=lever_L)[:2]
    return val, method


def _split(m, moat):
    life = m.moat_to_life(moat) or 15
    n1 = min(3, max(1, life - 1))
    return n1, max(1, life - n1)


def main():
    m = load_engine()
    import openpyxl
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("file")
    ap.add_argument("--rank-re", type=float, default=0.12, help="equity hurdle to RANK by (default 0.12)")
    ap.add_argument("--alt-re", type=float, default=0.07, help="second equity hurdle to display (default 0.07)")
    ap.add_argument("--mcap-floor", type=float, default=5e8, help="minimum market cap (default 500M)")
    ap.add_argument("--moat-min", type=float, default=7.0, help="minimum Moat Score (default 7.0)")
    ap.add_argument("--moat-trend-min", type=float, default=None,
                    help="exclude names whose 7y moat change (Moat Score - 'Moat Score - 7') is below "
                         "this, e.g. -0.3 drops material decliners, 0 keeps only flat/improving. "
                         "Names with no 7y history are kept. Off by default.")
    ap.add_argument("--n", type=int, default=30, help="portfolio size (default 30)")
    ap.add_argument("--tag-cap", type=int, default=6, help="max stocks sharing a material tag (default 6 = 20%%)")
    ap.add_argument("--tag-material", type=int, default=4, help="tag score >= this counts as material exposure (default 4)")
    ap.add_argument("--er-gate", type=float, default=None, help="require ER at the ranking hurdle > this (e.g. 0)")
    ap.add_argument("--tam-reinv-drop", type=float, default=0.25, help="TAM screen: exclude if RR3 < (1-this)*RR7 (default 0.25)")
    ap.add_argument("--no-lever-glide", action="store_true", help="value without the de-risking WACC glide")
    ap.add_argument("--country-base", default=None, help='refresh risk-free bases, e.g. "USD=0.0406,JPY=0.0265"')
    ap.add_argument("--country-crp", default=None, help="override country risk premiums")
    ap.add_argument("--sheet", default=None)
    ap.add_argument("--csv", default=None, help="write the portfolio to this CSV path")
    args = ap.parse_args()

    overrides = m.parse_kv_rates(args.country_base)
    crp_over = m.parse_kv_rates(args.country_crp)
    glide = not args.no_lever_glide
    wb = openpyxl.load_workbook(args.file, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb.worksheets[0]
    want = {"name": "Company Name", "nopat": "New Operating Income", "roiic": "ROICm 7",
            "rr": "RR 7", "moat": "Moat Score", "industry": "GICS Industry Group Name",
            "country": "Country of Headquarters", "gross": "Gross debt",
            "tax": "Income Tax Rate - Instrument", "mktcap": "Market Cap",
            "netdebt": "Net debt", "sales": "Sales", "ev": "EV",
            "ticker": "Instrument", "rr3": "RR 3", "moat7": "Moat Score - 7"}
    for t in TAGS:
        want[t] = t
    C = m.find_columns(ws, want)
    if C["name"] is None:
        sys.exit('Could not find a "Company Name" column.')
    if args.moat_trend_min is not None and C.get("moat7") is None:
        print("WARNING: --moat-trend-min set but no 'Moat Score - 7' column — trend filter skipped.",
              file=sys.stderr)
    missing_tags = [t for t in TAGS if C[t] is None]
    if missing_tags:
        print(f"WARNING: {len(missing_tags)} tag column(s) missing — the 20% rule "
              f"can't be enforced for them: {missing_tags}", file=sys.stderr)

    exc = {"data": 0, "mcap": 0, "moat": 0, "moat_trend": 0, "tags": 0, "tam": 0, "noval": 0, "gate": 0}
    cand = []
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=C["name"]).value
        if not name:
            continue
        g = lambda k: m.num(ws, row, C[k])
        nopat0, roiic0, rr0 = g("nopat"), g("roiic"), g("rr")
        mktcap, netdebt, ev = g("mktcap"), g("netdebt"), g("ev")
        ind = " ".join(str(ws.cell(row=row, column=C["industry"]).value or "").split())
        if ind in m.FINANCIAL_SECTORS:
            continue
        if None in (nopat0, roiic0, rr0) or not mktcap or ev in (None, 0) or netdebt is None:
            exc["data"] += 1; continue
        if mktcap < args.mcap_floor:
            exc["mcap"] += 1; continue
        moat = g("moat")
        if moat is None or moat < args.moat_min:
            exc["moat"] += 1; continue
        moat7 = g("moat7") if C.get("moat7") else None
        moat_chg = (moat - moat7) if (moat7 is not None) else None
        if args.moat_trend_min is not None and moat_chg is not None and moat_chg < args.moat_trend_min:
            exc["moat_trend"] += 1; continue            # drop materially-declining moats
        tagv = {t: (m.num(ws, row, C[t]) if C[t] else None) for t in TAGS}
        if any(C[t] is not None and tagv[t] is None for t in TAGS):
            exc["tags"] += 1; continue
        rr3, rr7 = g("rr3"), rr0
        if rr7 and rr7 > 0 and rr3 is not None and rr3 < (1 - args.tam_reinv_drop) * rr7:
            exc["tam"] += 1; continue
        country = ws.cell(row=row, column=C["country"]).value
        gross, tax = g("gross"), g("tax")
        er_rank, meth = equity_er(m, args.rank_re, nopat0, roiic0, rr0, mktcap, netdebt,
                                  gross, tax, country, ind, moat, g("sales"), overrides, crp_over, glide)
        if er_rank is None:
            exc["noval"] += 1; continue
        if args.er_gate is not None and er_rank <= args.er_gate:
            exc["gate"] += 1; continue
        er_alt = equity_er(m, args.alt_re, nopat0, roiic0, rr0, mktcap, netdebt,
                           gross, tax, country, ind, moat, g("sales"), overrides, crp_over, glide)[0]
        mat = set(t for t in TAGS if C[t] is not None and tagv[t] is not None and tagv[t] >= args.tag_material)
        cand.append({"name": str(name), "tick": ws.cell(row=row, column=C["ticker"]).value,
                     "co": str(country), "ind": ind, "moat": moat, "moat_chg": moat_chg, "mc": mktcap,
                     "er_rank": er_rank, "er_alt": er_alt, "meth": meth, "mat": mat})

    cand.sort(key=lambda d: d["er_rank"], reverse=True)
    cnt = {t: 0 for t in TAGS}
    port, bumped = [], []
    for d in cand:
        if len(port) >= args.n:
            break
        if all(cnt[t] < args.tag_cap for t in d["mat"]):
            port.append(d)
            for t in d["mat"]:
                cnt[t] += 1
        else:
            bumped.append(d)

    _report(args, cand, port, bumped, cnt, exc, glide)
    if args.csv:
        _write_csv(args, port)


def _report(args, cand, port, bumped, cnt, exc, glide):
    w = 1.0 / args.n * 100
    fr = lambda x: f"{x*100:6.1f}%" if x is not None else "   n/a"
    fmc = lambda x: f"{x/1e9:6.1f}b" if x >= 1e9 else f"{x/1e6:5.0f}m"
    print(f"\nAIP CORE-{args.n} — ranked by ER@{int(args.rank_re*100)}%"
          + (f" (gate >{args.er_gate*100:.0f}%)" if args.er_gate is not None else "")
          + f"  | mcap>=${args.mcap_floor/1e6:.0f}M, moat>={args.moat_min:g}, "
          + f"material tag>={args.tag_material}, cap {args.tag_cap}/tag"
          + (f", moat-trend>={args.moat_trend_min:+g}" if args.moat_trend_min is not None else "")
          + ("" if glide else ", no-lever-glide"))
    print(f"candidates {len(cand)}   selected {len(port)}   "
          f"excluded {exc}")
    fd = lambda x: (f"{x:+5.2f}" if x is not None else "  n/a")
    hdr = (f"{'#':>2} {'Company':30}{'Ctry':5}{'Moat':>5}{'Δ7y':>6}{'MCap':>8}"
           f"{'ER@'+str(int(args.rank_re*100)):>7}{'ER@'+str(int(args.alt_re*100)):>7}  Material tags")
    print(hdr); print("-" * len(hdr))
    for i, d in enumerate(port, 1):
        mt = ",".join(SHORT[TAGS.index(t)] for t in TAGS if t in d["mat"])
        flag = "c" if d["meth"] == "CAGR" else " "
        print(f"{i:>2} {d['name'][:29]:30}{_cc(d['co']):5}{d['moat']:5.1f}{fd(d.get('moat_chg')):>6}{fmc(d['mc'])}"
              f"{fr(d['er_rank'])}{flag}{fr(d['er_alt'])}  {mt}")
    pos = sum(1 for d in port if d["er_rank"] is not None and d["er_rank"] > 0)
    import statistics as st
    if port:
        print(f"\nEqual-weight {w:.2f}%/name.  avg ER@{int(args.rank_re*100)}% "
              f"{st.mean(d['er_rank'] for d in port)*100:.1f}%  median "
              f"{st.median([d['er_rank'] for d in port])*100:.1f}%  (positive {pos}/{len(port)})")
        from collections import Counter
        print("country mix:", dict(Counter(_cc(d['co']) for d in port).most_common()))
        chg = [d["moat_chg"] for d in port if d.get("moat_chg") is not None]
        if chg:
            up = sum(1 for x in chg if x > 0.1); dn = sum(1 for x in chg if x < -0.1)
            print(f"moat trend (7y): improving {up}, declining {dn}, flat {len(chg)-up-dn}, "
                  f"n/a {len(port)-len(chg)}  (mean Δ {st.mean(chg):+.2f})")
    capw = args.tag_cap / args.n * 100
    print(f"\nTAG EXPOSURE (material>= {args.tag_material}; cap {args.tag_cap} = {capw:.0f}%):")
    print("  " + "  ".join(f"{SHORT[i]}:{cnt[t]}{'*' if cnt[t] >= args.tag_cap else ''}"
                           for i, t in enumerate(TAGS)))
    if bumped:
        print(f"\nbumped by the {capw:.0f}% rule (rank in but breach a full tag):")
        for d in bumped[:8]:
            full = ",".join(SHORT[TAGS.index(t)] for t in TAGS if t in d["mat"] and cnt[t] >= args.tag_cap)
            print(f"   {d['name'][:30]:31}{_cc(d['co']):5} ER {fr(d['er_rank'])}  full: {full}")
    print("\nNote: TAM screen runs only the reinvestment-deceleration leg (RR3 vs RR7); "
          "the give-up ratio and top-line-gravity legs need FCF/payout/revenue-history "
          "columns. Analytical framework output, not investment advice.")


_CC = {'United States of America': 'US', 'Japan': 'JP', 'United Kingdom': 'UK',
       'Sweden': 'SE', 'Germany': 'DE', 'Korea; Republic (S. Korea)': 'KR',
       'Switzerland': 'CH', 'France': 'FR', 'Canada': 'CA', 'Netherlands': 'NL',
       'Ireland': 'IE', 'Italy': 'IT', 'Denmark': 'DK', 'Finland': 'FI',
       'Spain': 'ES', 'Mexico': 'MX', 'Uruguay': 'UY', 'Brazil': 'BR',
       'Luxembourg': 'LU', 'Norway': 'NO', 'Poland': 'PL', 'Greece': 'GR'}


def _cc(c):
    return _CC.get(c, (c[:4] if c else "?"))


def _write_csv(args, port):
    with open(args.csv, "w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["Rank", "Company", "Ticker", "Country", "Industry", "Moat", "Moat7yChg",
                    "MktCap_bn", "Weight", f"ER@{int(args.rank_re*100)}%",
                    f"ER@{int(args.alt_re*100)}%", "Method"] + SHORT)
        for i, d in enumerate(port, 1):
            w.writerow([i, d["name"], d["tick"], d["co"], d["ind"], f"{d['moat']:.2f}",
                        f"{d['moat_chg']:+.2f}" if d.get("moat_chg") is not None else "n/a",
                        f"{d['mc']/1e9:.2f}", f"{1.0/args.n*100:.2f}%",
                        f"{d['er_rank']*100:.1f}%" if d['er_rank'] is not None else "n/a",
                        f"{d['er_alt']*100:.1f}%" if d['er_alt'] is not None else "n/a",
                        d["meth"]] + [("1" if t in d["mat"] else "0") for t in TAGS])
    print(f"\nsaved -> {args.csv}")


if __name__ == "__main__":
    main()
