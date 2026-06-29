#!/usr/bin/env python3
"""revalue.py — re-run the valuation engine on fresh prices.

The only thing that goes stale in this model is PRICE. Fundamentals (NOPAT,
ROIC/RR, tax, sales), the researched moats, and the long-run history are stable.
So to revalue a name today an analyst supplies just its current **Market Cap**
and **Enterprise Value**; this module patches those onto the existing screener
fundamentals, re-runs `aip.value_and_return`, and returns the expected return
*and its carry / re-rate decomposition* (see decompose.py).

Analyst upload (xlsx or csv) — one row per stock, header row required:

    Instrument | Market Cap | EV            (Net debt = EV - Market Cap)
    7203.T     | 38500000   | 41000000
    NOVOb.CO   | 250000000  | 248000000
    ...

Optional columns `Net debt` / `Gross debt` override the derived values. Tickers
must match the screener's `Instrument` ids. Names absent from the upload keep
their existing price (or are skipped with --only-uploaded).

Usage
-----
  AIP_VALUE_ENGINE=/path/to/roiic_dcf.py python3 revalue.py \
      --ltm   LTM_baseline.xlsx \
      --prices analyst_prices.xlsx \
      --out    revalued.xlsx \
      [--moats Universe_final.xlsx]   # use researched CompanyMoat instead of screener Moat Score

Programmatic
------------
  from revalue import load_prices, revalue
  rows = revalue("LTM_baseline.xlsx", load_prices("analyst_prices.xlsx"))
"""
from __future__ import annotations
import argparse
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import openpyxl
import aip
from decompose import decompose


def _num(v):
    try:
        return float(v) if v not in (None, "") else None
    except (TypeError, ValueError):
        return None


# ------------------------------------------------------------------- inputs

def load_prices(path):
    """Read an analyst price upload (.xlsx or .csv) -> {ticker: {mktcap, ev, netdebt, grossdebt}}."""
    rows = []
    if str(path).lower().endswith(".csv"):
        import csv
        with open(path, newline="") as f:
            rows = list(csv.reader(f))
        rows = [list(r) for r in rows]
    else:
        ws = openpyxl.load_workbook(path, data_only=True, read_only=True).worksheets[0]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
    if not rows:
        return {}
    hdr = [str(h).strip() if h is not None else "" for h in rows[0]]
    idx = {h: i for i, h in enumerate(hdr)}

    def col(*names):
        for n in names:
            if n in idx:
                return idx[n]
        return None

    ti = col("Instrument", "Ticker", "RIC")
    mi = col("Market Cap", "MktCap", "Market Capitalization")
    ei = col("EV", "Enterprise Value", "Enterprise_Value")
    ni = col("Net debt", "Net Debt")
    gi = col("Gross debt", "Gross Debt")
    if ti is None or mi is None:
        raise ValueError("price file needs at least 'Instrument' and 'Market Cap' columns")
    out = {}
    for r in rows[1:]:
        if ti >= len(r) or not r[ti]:
            continue
        out[r[ti]] = {
            "mktcap": _num(r[mi]) if mi is not None and mi < len(r) else None,
            "ev": _num(r[ei]) if ei is not None and ei < len(r) else None,
            "netdebt": _num(r[ni]) if ni is not None and ni < len(r) else None,
            "grossdebt": _num(r[gi]) if gi is not None and gi < len(r) else None,
        }
    return out


def load_moats(scored_book):
    """{ticker: researched CompanyMoat} from a scored book's Scored sheet (optional)."""
    ws = openpyxl.load_workbook(scored_book, data_only=True, read_only=True)["Scored"]
    rows = list(ws.iter_rows(values_only=True)); ci = {h: i for i, h in enumerate(rows[0])}
    out = {}
    for r in rows[1:]:
        t = r[ci["Ticker"]] if "Ticker" in ci else None
        m = r[ci["CompanyMoat(v3.2)"]] if "CompanyMoat(v3.2)" in ci else None
        if t and isinstance(m, (int, float)):
            out[t] = m
    return out


def _base_financials(ltm_path):
    ws = openpyxl.load_workbook(ltm_path, data_only=True, read_only=True).worksheets[0]
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    li = {h: i for i, h in enumerate(hdr)}
    text = ("Company Name", "Instrument", "GICS Industry Group Name", "Country of Headquarters")
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        t = row[li["Instrument"]] if "Instrument" in li else None
        if not t:
            continue
        out[t] = {fld: (row[li[fld]] if fld in text else _num(row[li[fld]]))
                  for fld in aip.FIELDS.values() if fld in li}
    return out


# ---------------------------------------------------------------- revaluation

def revalue(ltm_path, prices, *, moats=None, re=0.07, re2=0.12, only_uploaded=True):
    """Revalue names on fresh prices. Returns a list of result dicts (one per name).

    prices: {ticker: {mktcap, ev, netdebt, grossdebt}} (load_prices()).
    moats:  optional {ticker: moat} to override the screener Moat Score with the
            researched moat. only_uploaded: value just the uploaded names."""
    base = _base_financials(ltm_path)
    names = list(prices) if only_uploaded else list(base)
    results = []
    for t in names:
        fin = base.get(t)
        if not fin:
            results.append({"ticker": t, "error": "not in screener"}); continue
        fin = dict(fin)
        p = prices.get(t, {})
        if p.get("mktcap") is not None:
            fin["Market Cap"] = p["mktcap"]
        # net debt: explicit > derived from EV > existing
        if p.get("netdebt") is not None:
            fin["Net debt"] = p["netdebt"]
        elif p.get("ev") is not None and fin.get("Market Cap") is not None:
            fin["Net debt"] = p["ev"] - fin["Market Cap"]
        if p.get("grossdebt") is not None:
            fin["Gross debt"] = p["grossdebt"]
        if moats and t in moats:
            fin["Moat Score"] = moats[t]
        d = decompose(fin, re=re, re2=re2)
        if not d:
            results.append({"ticker": t, "error": "engine could not value"}); continue
        ev = (fin.get("Market Cap") or 0) + (fin.get("Net debt") or 0)
        results.append({
            "ticker": t, "company": fin.get("Company Name"),
            "industry": fin.get("GICS Industry Group Name"),
            "market_cap": fin.get("Market Cap"), "ev": ev,
            "wacc": d["wacc"], "rating": d["rating"],
            "er_7": d["er"], "er_12": d["er2"],
            "carry": d["carry"], "rerate": d["rerate"],
            "carry_share": d["carry_share"], "driver": d["driver"],
            "moat_used": fin.get("Moat Score"),
        })
    return results


def write_results(path, rows, re=0.07, re2=0.12):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Revalued"
    ws.append(["Ticker", "Company", "Industry", "Market Cap", "EV", "WACC", "AIP Rating",
               f"ER@{re:.0%}", f"ER@{re2:.0%}", "Carry (ROIIC)", "Re-rate", "Carry % of ER",
               "Driver", "Moat used", "Note"])
    ok = [r for r in rows if "error" not in r]
    for r in rows:
        if "error" in r:
            ws.append([r["ticker"], "", "", "", "", "", "", "", "", "", "", "", "", "", r["error"]])
        else:
            ws.append([r["ticker"], r["company"], r["industry"], r["market_cap"], r["ev"],
                       r["wacc"], r["rating"], r["er_7"], r["er_12"], r["carry"], r["rerate"],
                       r["carry_share"], r["driver"], r["moat_used"], ""])
    if ok:
        avg = lambda k: sum(r[k] for r in ok if isinstance(r[k], (int, float))) / len(ok)
        ws.append([])
        ws.append(["EQUAL-WT", f"({len(ok)} valued)", "", "", "", "", "", avg("er_7"), avg("er_12"),
                   avg("carry"), avg("rerate"), avg("carry") / avg("er_7") if avg("er_7") else None, "", "", ""])
    for row in ws.iter_rows(min_row=2):
        for j in (5, 7, 8, 9, 10, 11):
            if isinstance(row[j].value, float):
                row[j].number_format = "0.0%" if j != 5 else "#,##0"
        if isinstance(row[3].value, (int, float)):
            row[3].number_format = "#,##0"
        if isinstance(row[4].value, (int, float)):
            row[4].number_format = "#,##0"
    for col, w in zip("ABCDEFGHIJKLMNO", [11, 26, 26, 14, 14, 8, 10, 8, 8, 12, 9, 13, 12, 10, 18]):
        ws.column_dimensions[col].width = w
    wb.save(path)
    return len(ok), len(rows) - len(ok)


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--ltm", required=True, help="baseline LTM screener (fundamentals)")
    ap.add_argument("--prices", required=True, help="analyst price upload (.xlsx/.csv)")
    ap.add_argument("--out", required=True)
    ap.add_argument("--moats", help="scored book to source researched CompanyMoat (optional)")
    ap.add_argument("--re", type=float, default=0.07)
    ap.add_argument("--re2", type=float, default=0.12)
    ap.add_argument("--all", action="store_true", help="value the whole screener (default: only uploaded names)")
    args = ap.parse_args()

    prices = load_prices(args.prices)
    moats = load_moats(args.moats) if args.moats else None
    rows = revalue(args.ltm, prices, moats=moats, re=args.re, re2=args.re2,
                   only_uploaded=not args.all)
    ok, bad = write_results(args.out, rows, re=args.re, re2=args.re2)
    print(f"revalued {ok} names ({bad} skipped) -> {args.out}")


if __name__ == "__main__":
    main()
