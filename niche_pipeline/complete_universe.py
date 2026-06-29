#!/usr/bin/env python3
"""
complete_universe.py — merge screener+history inputs into one universe and run the
full book with live qualitative research.

The pipeline's Pass-2 research (moat v3.2 chapters, ownership, risk tags) only runs
when analyst.py finds ANTHROPIC_API_KEY (and the `anthropic` SDK). With the key set,
every Gate-1 survivor that lacks a cached record is researched via server-side web
search and cached, so a re-run is then free. Without the key, deterministic scoring
still runs and researched fields are left blank (NEEDS_RESEARCH) — same as before.

This runner stitches several LTM screeners and several history panels (newer rows win
for overlapping tickers), serves the merged panel to the pipeline, and writes the
book — so "complete the universe" is one command:

    ANTHROPIC_API_KEY=sk-ant-...  python3 complete_universe.py \
        --ltm  cec4ea6d-LTM.xlsx a549e877-LTM.xlsx 040bac3e-LTM.xlsx \
        --hist Moat_5_binary.xlsb df73f86e-hist.xlsx 79e5bd4b-hist.xlsx \
        --out  Universe_complete_book.xlsx

The FIRST --ltm / --hist is treated as the base universe; later files extend/refresh
it. Pass --research-all to research every gate-1 survivor (default) or --no-research
to skip Pass 2 entirely.
"""
from __future__ import annotations
import argparse
import openpyxl
import history
import pipeline


def read_ltm(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    li = {h: i for i, h in enumerate(hdr) if h is not None}
    fins = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        t = r[li["Instrument"]] if "Instrument" in li else None
        if t:
            fins[t] = {h: r[li[h]] for h in li}
    return hdr, fins


def merge_ltms(paths, out_path):
    """Merge screeners onto the first file's column header; later files win."""
    base_hdr, combined = read_ltm(paths[0])
    cols = [h for h in base_hdr if h is not None]
    for p in paths[1:]:
        _, f = read_ltm(p)
        combined.update(f)
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(cols)
    for f in combined.values():
        ws.append([f.get(c) for c in cols])
    wb.save(out_path)
    return len(combined)


def merge_hists(paths):
    """Merge history panels into one {ticker: rows}; later files win. Returns (by, idx)."""
    by, idx = history.load_panel(paths[0])
    for p in paths[1:]:
        by_n, _ = history.load_panel(p)
        by.update(by_n)
    return by, idx


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--ltm", nargs="+", required=True, help="LTM screener xlsx files (first = base)")
    ap.add_argument("--hist", nargs="+", required=True, help="history panel files (xlsb/xlsx; first = base)")
    ap.add_argument("--out", required=True, help="output book xlsx")
    ap.add_argument("--merged-ltm", default="/tmp/_merged_universe_LTM.xlsx")
    ap.add_argument("--re", type=float, default=0.07)
    ap.add_argument("--re2", type=float, default=0.12)
    ap.add_argument("--no-research", action="store_true", help="skip the qualitative research pass")
    ap.add_argument("--research-all", action="store_true",
                    help="research every gate-1 survivor (default researches survivors only)")
    ap.add_argument("--no-overearning", action="store_true")
    args = ap.parse_args()

    import analyst
    ready = analyst._have_sdk_and_key()
    print(f"research backend: {'LIVE (ANTHROPIC_API_KEY found)' if ready else 'OFF (no key — cached + deterministic only)'}")

    n = merge_ltms(args.ltm, args.merged_ltm)
    print(f"merged {len(args.ltm)} screeners -> {n} names ({args.merged_ltm})")
    by, idx = merge_hists(args.hist)
    print(f"merged {len(args.hist)} history panels -> {len(by)} names")

    history.load_panel = lambda *a, **k: (by, idx)        # serve the merged panel to the pipeline
    pipeline.run(args.merged_ltm, args.out, re=args.re, re2=args.re2,
                 research=not args.no_research, research_all=args.research_all,
                 history_path="merged", apply_overearning=not args.no_overearning)


if __name__ == "__main__":
    main()
