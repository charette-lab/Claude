"""Add a 'Take Over Time' sheet + chart: each role's % of the take (comp pool) by year."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np

WB = "/home/user/Claude/Performance_Fee_Framework.xlsx"
PNG = "/home/user/Claude/Take_share_over_time.png"

years = list(range(2006, 2027))
ret = [0.625,-0.0806,-0.4140,0.8213,0.28313,-0.019554,0.049944,0.375842,0.206523,
0.093046,0.266437,0.044292,-0.127671,-0.059219,0.169445,2.004571,0.003004,-0.005452,
0.250207,0.040814,-0.025809]

START_AUM = 100.0  # USD millions, editable assumption
scn_aum = [100, 300, 500, 1000, 2000, 5000]
roles = [
    ("Managing Partner / CIO",  100, [1,1,1,1,1,1]),
    ("Partner / Co-PM",          60, [1,1,2,2,3,4]),
    ("Portfolio Manager",        40, [0,1,1,2,3,5]),
    ("Head of Research",         30, [0,1,1,1,1,1]),
    ("Senior Analyst",           20, [1,1,2,3,4,6]),
    ("Analyst",                  12, [1,2,3,4,6,10]),
    ("Junior Analyst",            6, [0,1,2,3,4,8]),
    ("Head of Trading",          18, [0,0,1,1,1,1]),
    ("Trader / Execution",       10, [0,1,1,2,2,3]),
    ("COO / Head of Operations", 15, [1,1,1,1,1,1]),
    ("CFO / Finance",            12, [0,1,1,1,1,1]),
    ("Compliance / Legal",        8, [0,0,1,1,1,2]),
    ("Investor Relations / Ops", 5, [1,1,1,2,3,5]),
]

# AUM path
aum_path = []
a = START_AUM
for r in ret:
    a *= (1 + r); aum_path.append(a)

def interp_hc(hc, A):
    A = max(scn_aum[0], min(scn_aum[-1], A))
    for k in range(len(scn_aum) - 1):
        if scn_aum[k] <= A <= scn_aum[k+1]:
            t = (A - scn_aum[k]) / (scn_aum[k+1] - scn_aum[k])
            return hc[k] + t * (hc[k+1] - hc[k])
    return hc[-1]

# share matrix: rows=roles, cols=years  (% of take)
share = np.zeros((len(roles), len(years)))
for j, A in enumerate(aum_path):
    units = [w * interp_hc(hc, A) for (_, w, hc) in roles]
    tot = sum(units)
    for i, u in enumerate(units):
        share[i, j] = u / tot

# ---------- write sheet ----------
wb = openpyxl.load_workbook(WB)
if "Take Over Time" in wb.sheetnames:
    del wb["Take Over Time"]
ws = wb.create_sheet("Take Over Time", 1)

NAVY="1F3864"; BLUE="2E5496"; GREY="F2F2F2"; LIGHT="EAF0FA"; GOLD="FFF2CC"
title_f=Font(size=14,bold=True,color="FFFFFF"); sub_f=Font(size=10,italic=True,color="FFFFFF")
hdr_f=Font(size=9,bold=True,color="FFFFFF"); bold=Font(size=9,bold=True); norm=Font(size=9)
ital=Font(size=9,italic=True,color="595959"); inp_f=Font(size=10,bold=True,color="C00000")
fnavy=PatternFill("solid",fgColor=NAVY); fhdr=PatternFill("solid",fgColor=BLUE)
flight=PatternFill("solid",fgColor=LIGHT); fgrey=PatternFill("solid",fgColor=GREY); fgold=PatternFill("solid",fgColor=GOLD)
ctr=Alignment("center","center",wrap_text=True); lft=Alignment("left","center"); rgt=Alignment("right","center")
thin=Side("thin",color="BFBFBF"); box=Border(thin,thin,thin,thin)

n_years = len(years)
last_col = 2 + n_years  # col B is the start of year columns... we'll put role in A, years B..
def L(i): return openpyxl.utils.get_column_letter(i)

ws.merge_cells(f"A1:{L(1+n_years)}1")
c=ws["A1"]; c.value="Each role's share of the take (comp pool), over time"; c.font=title_f; c.fill=fnavy; c.alignment=ctr
ws.merge_cells(f"A2:{L(1+n_years)}2")
c=ws["A2"]; c.value="Share = role weight × headcount ÷ total weighted units. Headcount scales with the AUM path; % sums to 100% each year."; c.font=sub_f; c.fill=fnavy; c.alignment=ctr
ws.row_dimensions[1].height=22

# assumption cell
ws["A4"].value="Starting AUM (USD m), compounded by Athanase net returns:"; ws["A4"].font=norm
ws["A4"].alignment=rgt
ws.merge_cells("A4:C4")
ws["D4"].value=START_AUM; ws["D4"].font=inp_f; ws["D4"].fill=fgold; ws["D4"].alignment=ctr; ws["D4"].border=box

# header rows
hr=6
ws.cell(hr,1,"Year").font=hdr_f; ws.cell(hr,1).fill=fhdr; ws.cell(hr,1).alignment=ctr; ws.cell(hr,1).border=box
for j,y in enumerate(years):
    cc=ws.cell(hr,2+j,y); cc.font=hdr_f; cc.fill=fhdr; cc.alignment=ctr; cc.border=box
# AUM row
ws.cell(hr+1,1,"AUM (USD m)").font=bold; ws.cell(hr+1,1).fill=fgrey; ws.cell(hr+1,1).alignment=lft; ws.cell(hr+1,1).border=box
for j,A in enumerate(aum_path):
    cc=ws.cell(hr+1,2+j,round(A)); cc.font=norm; cc.fill=fgrey; cc.alignment=ctr; cc.number_format='#,##0'; cc.border=box
# role share rows
first_data = hr+2
for i,(name,_,_) in enumerate(roles):
    r=first_data+i
    cc=ws.cell(r,1,name); cc.font=norm; cc.alignment=lft; cc.border=box
    if i%2: cc.fill=flight
    for j in range(n_years):
        c2=ws.cell(r,2+j,float(share[i,j])); c2.font=norm; c2.alignment=ctr; c2.number_format='0.0%'; c2.border=box
        if i%2: c2.fill=flight
# total check row
tr=first_data+len(roles)
ws.cell(tr,1,"Total").font=bold; ws.cell(tr,1).fill=fhdr; ws.cell(tr,1).font=hdr_f; ws.cell(tr,1).alignment=lft; ws.cell(tr,1).border=box
for j in range(n_years):
    col=L(2+j)
    c2=ws.cell(tr,2+j); c2.value=f"=SUM({col}{first_data}:{col}{tr-1})"; c2.font=hdr_f; c2.fill=fhdr; c2.alignment=ctr; c2.number_format='0%'; c2.border=box

# native Excel line chart
chart=LineChart(); chart.title="Role share of the take over time (%)"; chart.style=2
chart.y_axis.numFmt='0%'; chart.y_axis.title="% of take"; chart.x_axis.title="Year"
chart.height=11; chart.width=26
data=Reference(ws,min_col=1,min_row=first_data,max_row=tr-1,max_col=1+n_years)
cats=Reference(ws,min_col=2,min_row=hr,max_col=1+n_years,max_row=hr)
chart.add_data(data,titles_from_data=True,from_rows=True)
chart.set_categories(cats)
ws.add_chart(chart,f"A{tr+2}")

ws.column_dimensions["A"].width=26
for j in range(n_years): ws.column_dimensions[L(2+j)].width=7
ws.sheet_view.showGridLines=False
ws.freeze_panes="B7"
wb.save(WB)

# ---------- PNG stacked area ----------
fig,ax=plt.subplots(figsize=(13,7))
cmap=plt.get_cmap("tab20")
colors=[cmap(i) for i in range(len(roles))]
ax.stackplot(years,[share[i]*100 for i in range(len(roles))],
             labels=[r[0] for r in roles],colors=colors,alpha=0.9,edgecolor="white",linewidth=0.3)
ax.set_xlim(years[0],years[-1]); ax.set_ylim(0,100)
ax.set_ylabel("Share of the take (%)"); ax.set_xlabel("Year")
ax.set_title("How the performance-fee take is split by role, over time\n(Athanase net-return AUM path, start $%dm)"%START_AUM,
             fontsize=13,fontweight="bold")
ax.legend(loc="center left",bbox_to_anchor=(1.01,0.5),fontsize=8,frameon=False)
ax2=ax.twinx(); ax2.plot(years,aum_path,color="black",lw=1.6,ls="--",label="AUM")
ax2.set_ylabel("AUM (USD m)"); ax2.set_ylim(0,max(aum_path)*1.1)
ax.set_xticks(years[::2])
plt.tight_layout()
plt.savefig(PNG,dpi=130,bbox_inches="tight")
print("done", PNG)

# print a compact table for chat
print("\nRole % of take — selected years:")
hdr_y=[2006,2009,2013,2017,2021,2026]
idx=[years.index(y) for y in hdr_y]
print("Role".ljust(28)+"".join(f"{y:>8}" for y in hdr_y))
for i,(name,_,_) in enumerate(roles):
    print(name.ljust(28)+"".join(f"{share[i,j]*100:>7.1f}%" for j in idx))
print("AUM(m)".ljust(28)+"".join(f"{aum_path[j]:>8.0f}" for j in idx))
