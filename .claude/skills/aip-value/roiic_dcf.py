#!/usr/bin/env python3
"""AIP Value — ROIIC persistence-fade DCF, run against an attached Excel file.

Economic premise: growth requires capital (g = ROIIC x RR), and a company's
*excess* return (ROIIC - WACC) does not persist — it decays geometrically toward
the cost of capital at a speed set by the durability of the moat. This directly
handles a cyclically-inflated starting ROIIC (e.g. a semiconductor firm posting
60% ROIIC at a demand peak): the surge is mean-reverted away over the empirical
Competitive Advantage Period (CAP) rather than held flat.

Reinvestment-driven engine. Growth is whatever the firm's own reinvestment can
fund — g = ROIIC * RR — NOT an externally imposed industry rate. RR is the
structural reinvestment rate (RR_0, held); as ROIIC fades toward the CFROI base,
g = ROIIC * RR falls of its own accord. A SALES-GROWTH FLOOR (the Mauboussin
size x industry MEDIAN sales-growth base rate) stops g — and so the implied RR —
from going artificially low: a firm needs to reinvest at least enough to keep its
sales growing with its industry. When the floor binds, RR is lifted to fund it.
    Phase 1, t = 1..n1 :  ROIIC_t = ROIIC_0
    Phase 2, k = 1..n2 :  ROIIC_t = base + (ROIIC_0 - base)*phi**k
    g_t   = max( ROIIC_t * RR_0 , sales_base_median(size_t) )   # funded, but >= floor
    RR_t  = min( g_t / ROIIC_t , 1 )                            # lifted when floored
    NOPAT_t = NOPAT_(t-1)*(1+g_t) ;  FCF_t = NOPAT_t*(1-RR_t)
    PV = sum_t FCF_t / (1+WACC)**t
n1 is a fixed 3y hold; n2 = (moat-score life - 3), so n1 + n2 = the total moat
life (< 6 -> <10y, 6-7.5 -> 10-20y, > 7.5 -> 50y). phi is the moat-tier fade.
Note: where the faded ROIIC sits BELOW WACC, the floor forces value-DESTROYING
reinvestment (RR rises to fund growth that earns < cost of capital) — this is
correct and conservative: it prevents an artificially-low RR from inflating FCF.
Terminal — competitive equilibrium (standard continuing value). By the end of the
CAP the return on NEW invested capital (RONIC) has competed down to the cost of
capital, so terminal growth is value-NEUTRAL: it runs at a GDP-like rate g_term
but adds nothing. RONIC = WACC, so RR_term = g_term/WACC and TV collapses to
NOPAT_N*(1+g)/WACC:
    g_eff = min(g_term, 0.99r) ;  RONIC = WACC ;  RR_term = g_eff / WACC
    TV    = NOPAT_N*(1+g_eff)*(1 - RR_term) / (WACC - g_eff)  ==  NOPAT_N*(1+g_eff)/WACC
    PV_TV = TV / (1+WACC)**N

Empirical CAP durations & persistence, mapped from the Moat Score:
    Moat > 7.5   Superior / Wide moat  -> CAP 10-20y, persistence 0.85-0.95
    6.0 - 7.5    Narrow / Standard     -> CAP  5-10y, persistence 0.70-0.80
    Moat < 6.0   No moat / cyclical    -> CAP  1-5y,  persistence 0.50-0.60
(interpolated within each band by score; override with --cap / --persistence).

NOTE on notation: the persistence factor is written as 'r' in the source
framework, but here 'r' is the discount rate / WACC. The persistence factor is
called 'persistence' (phi) to avoid the clash.

DEFAULT EXCEL COLUMN MAPPING (override with --col-* flags):
  NOPAT_0  <- "New Operating Income"   ROIIC_0 <- "ROICm 7"   RR <- "RR 7"
  moat     <- "Moat Score"
  (comparison) "EV", "Market Cap", "Net debt",
               "Shares used to calculate Diluted EPS - Total", "Close Price"

DEFAULT PARAMETERS (override with flags):
  r (WACC) = 0.12     g_term = 0.025 (< r)     horizon = 5 (for the IRR)

Usage:
  python3 roiic_dcf.py <file.xlsx> "<company>" [options]
  python3 roiic_dcf.py <file.xlsx> --list
Options:
  --r FLOAT --gterm FLOAT --cap INT --persistence FLOAT --horizon INT
  --payout-total FLOAT --sheet NAME --col-* STR
"""

import argparse
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")


def moat_to_cap_persistence(score):
    """Map a Moat Score to (CAP years, persistence factor, tier label).

    Empirical CAP durations:
      score > 7.5      Superior/Wide  -> CAP 10-20y, persistence 0.85-0.95
      6.0 <= s <= 7.5  Narrow/Std     -> CAP  5-10y, persistence 0.70-0.80
      score < 6.0      No moat/cyclic -> CAP  1-5y,  persistence 0.50-0.60
    Values are interpolated linearly within each band. Returns None if missing.
    """
    if score is None:
        return None
    if score > 7.5:                                  # Superior / Wide
        f = min(max((score - 7.5) / (9.0 - 7.5), 0.0), 1.0)
        cap = 10 + f * (20 - 10)
        phi = 0.85 + f * (0.95 - 0.85)
        tier = "Wide"
    elif score >= 6.0:                               # Narrow / Standard
        f = (score - 6.0) / (7.5 - 6.0)
        cap = 5 + f * (10 - 5)
        phi = 0.70 + f * (0.80 - 0.70)
        tier = "Narrow"
    else:                                            # No moat / cyclical
        f = max(0.0, score) / 6.0
        cap = 1 + f * (5 - 1)
        phi = 0.50 + f * (0.60 - 0.50)
        tier = "Cyclical"
    return max(1, int(round(cap))), phi, tier


def find_columns(ws, wanted):
    """Map each wanted header string to its 1-based column index (exact, then
    case-insensitive, then prefix match on the first row)."""
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is not None:
            headers[str(v).strip()] = c
    lower = {k.lower(): idx for k, idx in headers.items()}
    out = {}
    for key, label in wanted.items():
        if label in headers:
            out[key] = headers[label]
        elif label.lower() in lower:
            out[key] = lower[label.lower()]
        else:
            match = next((idx for h, idx in lower.items()
                          if h.startswith(label.lower())), None)
            out[key] = match
    return out


def find_company_row(ws, name_col, query):
    q = query.strip().lower()
    matches = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=name_col).value
        if v and q in str(v).lower():
            matches.append((r, str(v)))
    return matches


def num(ws, row, col):
    if col is None:
        return None
    v = ws.cell(row=row, column=col).value
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def money(x):
    if x is None:
        return "n/a"
    a = abs(x)
    if a >= 1e9:
        return f"{x/1e9:,.2f} bn"
    if a >= 1e6:
        return f"{x/1e6:,.1f} mn"
    return f"{x:,.0f}"


def pct(x):
    return "n/a" if x is None else f"{x*100:.2f}%"


# ROIIC base rate = long-term sector CFROI median (Mauboussin, The Base Rate
# Book, 1950-2015). The sheet's GICS industry groups are mapped up to the sector
# whose CFROI applies. NOTE: CFROI is inflation-adjusted (real); WACC here is
# nominal. Use --base-inflation to gross the base rate up to nominal if desired.
SECTOR_CFROI_MEDIAN = {        # sector -> (CFROI median, 5y persistence)
    "Information Technology": (0.085, 0.50),
    "Consumer Staples": (0.081, 0.78),
    "Consumer Discretionary": (0.080, 0.67),
    "Health Care": (0.083, 0.64),
    "Financials": (0.075, 0.43),
    "Industrials": (0.067, 0.62),
    "Telecommunication Services": (0.057, 0.55),
    "Energy": (0.050, 0.35),
    "Materials": (0.046, 0.41),
    "Utilities": (0.035, 0.57),
}
# GICS Industry Group (as it appears in the sheet) -> sector above.
INDUSTRY_TO_SECTOR = {
    "Software & Services": "Information Technology",
    "Technology Hardware & Equipment": "Information Technology",
    "Semiconductors & Semiconductor Equipment": "Information Technology",
    "Health Care Equipment & Services": "Health Care",
    "Pharmaceuticals, Biotechnology & Life Sciences": "Health Care",
    "Capital Goods": "Industrials",
    "Commercial & Professional Services": "Industrials",
    "Transportation": "Industrials",
    "Materials": "Materials",
    "Automobiles & Components": "Consumer Discretionary",
    "Consumer Durables & Apparel": "Consumer Discretionary",
    "Consumer Discretionary Distribution & Retail": "Consumer Discretionary",
    "Consumer Services": "Consumer Discretionary",
    "Food, Beverage & Tobacco": "Consumer Staples",
    "Media & Entertainment": "Telecommunication Services",  # Comm. Services proxy
    "Utilities": "Utilities",
}
INDUSTRY_BASE_RATE = {
    ind: SECTOR_CFROI_MEDIAN[sec][0] for ind, sec in INDUSTRY_TO_SECTOR.items()
}
DEFAULT_BASE_RATE = 0.055     # full-universe long-term CFROI baseline (~5-6%)


def base_rate_for(industry, override=None):
    """Resolve the ROIIC base rate: explicit override > industry table > default.
    Returns (rate, source_label)."""
    if override is not None:
        return override, "override"
    if industry:
        key = " ".join(str(industry).split())          # collapse double spaces
        for k, v in INDUSTRY_BASE_RATE.items():
            if " ".join(k.split()).lower() == key.lower():
                return v, f"{key} (table)"
    return DEFAULT_BASE_RATE, "default"


# ---------------------------------------------------------------------------
# Cost of debt / WACC  (firm-specific synthetic rating + country risk-free)
# ---------------------------------------------------------------------------
# Cached 10y risk-free BASE by currency (govt 10y; USD already swap-adjusted
# ~UST-40bp). REFRESH these live each run (pass --country-base "EUR=0.0303,...").
# As of early June 2026.
CURRENCY_BASE = {
    "JPY": 0.0267, "SEK": 0.0273, "EUR": 0.0303, "KRW": 0.0412,
    "GBP": 0.0485, "USD": 0.0405, "CHF": 0.0050, "NOK": 0.0441, "DKK": 0.0285,
}
DEFAULT_CCY_BASE = 0.04
# Country of Headquarters -> currency.
COUNTRY_TO_CCY = {
    "United States": "USD", "USA": "USD", "Japan": "JPY", "South Korea": "KRW",
    "Korea": "KRW", "Korea, Republic of": "KRW", "Sweden": "SEK",
    "United Kingdom": "GBP", "Switzerland": "CHF", "Norway": "NOK",
    "Denmark": "DKK",
    # Eurozone members -> EUR
    "Germany": "EUR", "France": "EUR", "Netherlands": "EUR", "Finland": "EUR",
    "Italy": "EUR", "Spain": "EUR", "Ireland": "EUR", "Belgium": "EUR",
    "Austria": "EUR", "Portugal": "EUR", "Greece": "EUR", "Luxembourg": "EUR",
}
# Synthetic-rating default spread by interest coverage (Damodaran, ~2024 level).
# REFRESH the level live (ICE BofA OAS). Large-firm and small-firm breakpoints.
RATING_LARGE = [(8.5, "AAA", .0059), (6.5, "AA", .0078), (5.5, "A+", .0098),
                (4.25, "A", .0108), (3.0, "A-", .0122), (2.5, "BBB", .0156),
                (2.25, "BB+", .0200), (2.0, "BB", .0240), (1.75, "B+", .0351),
                (1.5, "B", .0421), (1.25, "B-", .0515), (0.8, "CCC", .0820),
                (0.65, "CC", .0864), (0.2, "C", .1134), (-9e9, "D", .1512)]
RATING_SMALL = [(12.5, "AAA", .0059), (9.5, "AA", .0078), (7.5, "A+", .0098),
                (6.0, "A", .0108), (4.5, "A-", .0122), (4.0, "BBB", .0156),
                (3.5, "BB+", .0200), (3.0, "BB", .0240), (2.5, "B+", .0351),
                (2.0, "B", .0421), (1.5, "B-", .0515), (1.25, "CCC", .0820),
                (0.8, "CC", .0864), (0.5, "C", .1134), (-9e9, "D", .1512)]
FINANCIAL_SECTORS = {"Financials", "Banks", "Diversified Financials", "Insurance"}

# Country risk premium (≈ Damodaran country default spread, ~2025-26). Added to
# the discount rate for a USD-denominated valuation of a non-AAA-sovereign firm.
# 0 for developed AAA/AA sovereigns. REFRESH live; override with --country-crp.
COUNTRY_RISK_PREMIUM = {
    "United States of America": 0.0, "USA": 0.0, "United States": 0.0,
    "Germany": 0.0, "Netherlands": 0.0, "Switzerland": 0.0, "Sweden": 0.0,
    "Denmark": 0.0, "Norway": 0.0, "Finland": 0.0, "Austria": 0.0,
    "Ireland": 0.0, "Singapore": 0.0, "Canada": 0.0, "New Zealand": 0.0,
    "United Kingdom": 0.003, "France": 0.003,
    "Japan": 0.004, "Korea; Republic (S. Korea)": 0.004, "South Korea": 0.004,
    "Korea": 0.004, "Taiwan": 0.004, "Israel": 0.010, "Italy": 0.015,
    "China": 0.007, "India": 0.025, "Indonesia": 0.018, "Philippines": 0.018,
    "Vietnam": 0.027, "Brazil": 0.030, "Turkey": 0.060,
}


def country_risk_premium(country, overrides=None):
    """Country/sovereign risk premium for a USD valuation (0 for DM AAA/AA)."""
    if not country:
        return 0.0
    key = " ".join(str(country).split())
    ov = overrides or {}
    if key in ov:
        return ov[key]
    return COUNTRY_RISK_PREMIUM.get(key, 0.0)


def currency_base(country, overrides=None):
    """Risk-free base for a company's country. overrides: dict currency->rate."""
    ccy = COUNTRY_TO_CCY.get(" ".join(str(country).split()), None) if country else None
    if ccy is None:
        return DEFAULT_CCY_BASE, "default"
    rate = (overrides or {}).get(ccy, CURRENCY_BASE.get(ccy, DEFAULT_CCY_BASE))
    return rate, ccy


def synthetic_rd(ebit, gross_debt, mktcap, base):
    """Iterate interest-coverage -> synthetic rating -> spread -> R_d (cost of
    debt = base + spread). Returns (R_d, rating, coverage)."""
    tbl = RATING_SMALL if (mktcap is not None and mktcap < 2e9) else RATING_LARGE
    if not gross_debt or gross_debt <= 0 or ebit is None:
        return base + tbl[0][2], tbl[0][1], float("inf")
    rd = base + 0.015
    rat, cov = tbl[0][1], float("inf")
    for _ in range(8):
        cov = ebit / (gross_debt * rd)
        sp, rat = tbl[-1][2], tbl[-1][1]
        for mc, r_, s_ in tbl:
            if cov >= mc:
                sp, rat = s_, r_
                break
        nrd = base + sp
        if abs(nrd - rd) < 1e-5:
            rd = nrd
            break
        rd = nrd
    return rd, rat, cov


def firm_wacc(re, rd, mktcap, netdebt, lo=0.04, hi=0.12):
    """WACC from an equity hurdle re and cost of debt rd, weighted by capital
    structure (EV = MktCap + NetDebt). Capped to [lo, hi] for sane extremes."""
    ev = mktcap + netdebt
    w = (mktcap / ev) * re + (netdebt / ev) * rd if ev > 0 else 1.5 * re
    return min(max(w, lo), hi)


def parse_kv_rates(s):
    """Parse 'EUR=0.0303,USD=0.0405' into a dict."""
    out = {}
    for part in (s or "").split(","):
        if "=" in part:
            k, v = part.split("=", 1)
            try:
                out[k.strip().upper()] = float(v)
            except ValueError:
                pass
    return out


# ---------------------------------------------------------------------------
# Sales-growth base rates (Mauboussin, "The Impact of Intangibles on Base
# Rates", 2021, Exhibit 3 — 10y median CAGR and std, by SIZE x INDUSTRY,
# 1984-2020). Used to cap the model's growth: a firm can't out-grow its size.
# ---------------------------------------------------------------------------
SALES_GROWTH_BASE_10Y = {   # size bucket -> industry -> (median, std)
    "<1":    {"Healthcare": (.119, .265), "Technology": (.084, .136), "Consumer": (.074, .121), "Manufacturing": (.072, .111), "Other": (.076, .134), "All": (.078, .144)},
    "1-5":   {"Healthcare": (.077, .070), "Technology": (.057, .120), "Consumer": (.059, .078), "Manufacturing": (.051, .081), "Other": (.056, .101), "All": (.056, .091)},
    "5-10":  {"Healthcare": (.083, .050), "Technology": (.025, .104), "Consumer": (.052, .061), "Manufacturing": (.045, .062), "Other": (.037, .120), "All": (.045, .085)},
    "10-25": {"Healthcare": (.063, .061), "Technology": (.044, .072), "Consumer": (.045, .066), "Manufacturing": (.027, .061), "Other": (.032, .065), "All": (.036, .065)},
    "25-50": {"Healthcare": (.053, .046), "Technology": (.049, .085), "Consumer": (.043, .063), "Manufacturing": (.022, .071), "Other": (.035, .060), "All": (.037, .066)},
    "50-100":{"Healthcare": (.016, .044), "Technology": (.088, .073), "Consumer": (.040, .060), "Manufacturing": (.028, .096), "Other": (.052, .052), "All": (.039, .070)},
    ">100":  {"Healthcare": (.079, .004), "Technology": (.020, .036), "Consumer": (.034, .048), "Manufacturing": (.010, .076), "Other": (-.013, .044), "All": (.022, .059)},
}
GICS_TO_MAUBOUSSIN = {
    "Capital Goods": "Manufacturing", "Materials": "Manufacturing",
    "Commercial & Professional Services": "Other", "Transportation": "Other",
    "Automobiles & Components": "Consumer", "Consumer Durables & Apparel": "Consumer",
    "Consumer Discretionary Distribution & Retail": "Consumer", "Consumer Services": "Consumer",
    "Food, Beverage & Tobacco": "Consumer",
    "Health Care Equipment & Services": "Healthcare",
    "Pharmaceuticals, Biotechnology & Life Sciences": "Healthcare",
    "Software & Services": "Technology", "Technology Hardware & Equipment": "Technology",
    "Semiconductors & Semiconductor Equipment": "Technology",
    "Media & Entertainment": "Other", "Telecommunication Services": "Other",
    "Utilities": "Other", "Energy": "Other", "Financials": "Other",
}
GROWTH_Z = {"median": 0.0, "p75": 0.6745, "p90": 1.2816}   # normal-approx percentile


def _size_bucket(sales):
    for hi, key in ((1e9, "<1"), (5e9, "1-5"), (10e9, "5-10"), (25e9, "10-25"),
                    (50e9, "25-50"), (100e9, "50-100")):
        if sales < hi:
            return key
    return ">100"


def sales_growth_base(sales, gics_industry, z=0.6745):
    """Long-run sales-growth ceiling for a firm of this size & industry
    (Exhibit 3 10y median + z·std). z selects the percentile (0=median)."""
    if sales is None or sales <= 0:
        return None
    mau = GICS_TO_MAUBOUSSIN.get(" ".join(str(gics_industry).split()), "Other")
    cell = SALES_GROWTH_BASE_10Y[_size_bucket(sales)]
    med, std = cell.get(mau, cell["All"])
    return med + z * std


def moat_to_life(score):
    """Total competitive period (years) from the Moat Score:
       < 6.0 -> linear 0->10 ; 6.0-7.5 -> linear 10->20 ; > 7.5 -> 50."""
    if score is None:
        return None
    if score < 6.0:
        yrs = score / 6.0 * 10.0
    elif score <= 7.5:
        yrs = 10.0 + (score - 6.0) / 1.5 * 10.0
    else:
        yrs = 50.0
    return max(2, int(round(yrs)))


def split_life(life):
    """Split a total competitive period into n1 = 1/3 (hold), n2 = 2/3 (fade)."""
    n1 = max(1, int(round(life / 3.0)))
    n2 = max(1, int(round(life * 2.0 / 3.0)))
    return n1, n2


def value_company(nopat0, roiic0, rr0, r, g_term, n1, n2, phi, base,
                  sales0=None, gics_industry=None, sales_floor=True):
    """Reinvestment-driven DCF. Phase 1 (n1 yrs) holds ROIIC at ROICm7; Phase 2
    (n2 yrs) mean-reverts ROIIC to the CFROI base at persistence phi. RR_0 is the
    structural reinvestment rate, held flat, so growth g = ROIIC * RR_0 falls as
    ROIIC fades. A SALES-GROWTH FLOOR (Mauboussin size x industry MEDIAN base
    rate, recomputed as sales compound) keeps g — and the implied RR — from going
    artificially low: g_t = max(ROIIC_t*RR_0, sales_base_median); when the floor
    binds, RR_t = g_t/ROIIC_t is lifted to fund it. Where the faded ROIIC < WACC,
    that forced reinvestment is value-destroying (correct & conservative). With no
    sales/gics, the floor falls back to g_term. Returns schedule, PVs, total, FCF
    fn."""
    use_floor = bool(sales_floor)
    sales_run = [sales0 if (sales0 and sales0 > 0) else None]

    def floor_at(t):
        """Sales-growth median floor for the firm's size in year t (g_term if no sales)."""
        s = sales_run[t - 1] if t - 1 < len(sales_run) else None
        if not use_floor:
            return 0.0
        if s is not None:
            gb = sales_growth_base(s, gics_industry, z=0.0)   # MEDIAN
            if gb is not None:
                return gb
        return g_term

    nopat_path = [nopat0]           # index t -> NOPAT at end of year t (0 = today)
    fcf_path = [None]
    sched = []                      # (t, roiic_t, rr_t, g_t, phase)
    pv_explicit = 0.0

    def step(t, roiic_t, phase):
        nonlocal pv_explicit
        g_reinv = roiic_t * rr0
        g_t = max(g_reinv, floor_at(t))
        rr_t = min(g_t / roiic_t, 1.0) if roiic_t > 1e-9 else 1.0
        nopat_t = nopat_path[t - 1] * (1 + g_t)
        fcf_t = nopat_t * (1 - rr_t)
        nopat_path.append(nopat_t)
        fcf_path.append(fcf_t)
        if sales_run[0] is not None:
            sales_run.append(sales_run[t - 1] * (1 + g_t))
        sched.append((t, roiic_t, rr_t, g_t, phase))
        pv_explicit += fcf_t / (1 + r) ** t

    for t in range(1, n1 + 1):                       # Phase 1 — hold ROICm7
        step(t, roiic0, "hold")
    for k in range(1, n2 + 1):                       # Phase 2 — fade ROIIC to base
        roiic_k = base + (roiic0 - base) * (phi ** k)
        step(n1 + k, roiic_k, "fade")

    cap = n1 + n2
    nopat_n = nopat_path[cap]

    # Terminal — competitive equilibrium (McKinsey/Mauboussin continuing value).
    # By the end of the CAP the return on NEW invested capital (RONIC) has competed
    # down to the cost of capital, so terminal growth is value-NEUTRAL: it runs at a
    # GDP-like rate (g_term) but adds nothing to value. With RONIC = WACC the cost
    # of that growth is RR_term = g_eff / WACC, and TV reduces to NOPAT_N*(1+g)/r.
    g_eff = min(g_term, 0.99 * r)                 # GDP-like terminal growth
    roiic_term = r                                # RONIC = cost of capital
    rr_term = g_eff / roiic_term if roiic_term > 0 else 0.0
    cf_term = nopat_n * (1 + g_eff) * (1 - rr_term)
    tv = cf_term / (r - g_eff) if r > g_eff else nopat_n / r
    pv_tv = tv / (1 + r) ** cap
    total = pv_explicit + pv_tv

    def cf_for_year(t):
        if 1 <= t <= cap:
            return fcf_path[t]
        nopat_t = nopat_n * (1 + g_eff) ** (t - cap)
        return nopat_t * (1 - rr_term)

    return {"sched": sched, "pv_explicit": pv_explicit, "tv": tv, "pv_tv": pv_tv,
            "total": total, "nopat_n": nopat_n, "base": base, "n1": n1, "n2": n2,
            "rr_term": rr_term, "g_eff": g_eff, "roiic_term": roiic_term,
            "cf_for_year": cf_for_year}


def main():
    ap = argparse.ArgumentParser(add_help=True, description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("file")
    ap.add_argument("company", nargs="?", default=None)
    ap.add_argument("--list", action="store_true", help="list companies and exit")
    ap.add_argument("--r", type=float, default=0.12, help="flat WACC / discount rate")
    ap.add_argument("--re", type=float, default=None,
                    help="required EQUITY return; if set, the discount rate becomes a "
                         "per-company WACC = wE*re + wD*Rd, with Rd from a synthetic "
                         "credit rating (interest coverage) + the country risk-free base")
    ap.add_argument("--country-base", default=None,
                    help='refresh risk-free bases, e.g. "EUR=0.0303,USD=0.0405,JPY=0.0267"')
    ap.add_argument("--country-crp", default=None,
                    help='override country risk premiums, e.g. "India=0.025,China=0.007"')
    ap.add_argument("--col-country", default="Country of Headquarters")
    ap.add_argument("--col-gross", default="Gross debt")
    ap.add_argument("--col-tax", default="Income Tax Rate - Instrument")
    ap.add_argument("--gterm", type=float, default=0.025)
    ap.add_argument("--n1", type=int, default=None,
                    help="hold years (ROICm7 flat); if omitted = 1/3 of the moat life")
    ap.add_argument("--n2", type=int, default=None,
                    help="fade years (revert to base); if omitted = 2/3 of the moat life")
    ap.add_argument("--persistence", type=float, default=None,
                    help="persistence factor phi (0-1); if omitted, from the Moat Score")
    ap.add_argument("--base-rate", type=float, default=None,
                    help="ROIIC base rate the return reverts to; if omitted, from "
                         "the industry/sector CFROI table keyed on the GICS column")
    ap.add_argument("--base-inflation", type=float, default=0.0,
                    help="added to the real CFROI base rate to make it nominal; "
                         "default 0 = use real CFROI (mirrors our growth-capitalized, "
                         "lower ROICm). Set e.g. 0.025 to gross up to nominal.")
    ap.add_argument("--horizon", type=int, default=5,
                    help="holding period in years for the expected-return / IRR (default 5)")
    ap.add_argument("--payout-total", type=float, default=0.0,
                    help="cumulative dividends + net buybacks over the horizon, "
                         "subtracted from the cash sweep (default 0 = all FCF de-levers)")
    ap.add_argument("--sheet", default=None)
    ap.add_argument("--col-nopat", default="New Operating Income")
    ap.add_argument("--col-roiic", default="ROICm 7")
    ap.add_argument("--col-rr", default="RR 7")
    ap.add_argument("--col-name", default="Company Name")
    ap.add_argument("--col-moat", default="Moat Score")
    ap.add_argument("--col-industry", default="GICS Industry Group Name")
    ap.add_argument("--col-sales", default="Sales")
    ap.add_argument("--no-sales-floor", action="store_true",
                    help="disable the Mauboussin sales-growth MEDIAN floor on g "
                         "(by default g is floored so RR can't be artificially low)")
    args = ap.parse_args()

    if args.gterm >= args.r:
        sys.exit(f"g_term ({args.gterm}) must be < r ({args.r}).")

    wb = openpyxl.load_workbook(args.file, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb.worksheets[0]

    cols = find_columns(ws, {
        "name": args.col_name, "nopat": args.col_nopat,
        "roiic": args.col_roiic, "rr": args.col_rr, "moat": args.col_moat,
        "industry": args.col_industry, "country": args.col_country,
        "gross": args.col_gross, "tax": args.col_tax, "sales": args.col_sales,
        "ev": "EV", "mktcap": "Market Cap", "netdebt": "Net debt",
        "shares": "Shares used to calculate Diluted EPS - Total",
        "price": "Close Price", "ticker": "Instrument",
    })
    if cols["name"] is None:
        sys.exit(f'Could not find a "{args.col_name}" column in row 1.')

    if args.list or not args.company:
        print(f"Companies in '{ws.title}':")
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=cols["name"]).value
            if v:
                print(f"  {v}")
        if not args.company and not args.list:
            print("\nPass a company name substring to value one.")
        return

    matches = find_company_row(ws, cols["name"], args.company)
    if not matches:
        sys.exit(f'No company matching "{args.company}". Use --list to see options.')
    if len(matches) > 1:
        names = ", ".join(n for _, n in matches)
        sys.exit(f'Ambiguous "{args.company}" matches: {names}. Be more specific.')
    row, company = matches[0]

    nopat0 = num(ws, row, cols["nopat"])
    roiic0 = num(ws, row, cols["roiic"])
    rr0 = num(ws, row, cols["rr"])
    if None in (nopat0, roiic0, rr0):
        sys.exit(f"Missing driver inputs for {company}: "
                 f"NOPAT={nopat0}, ROIIC={roiic0}, RR={rr0}")

    g_term = args.gterm
    mktcap = num(ws, row, cols["mktcap"])
    netdebt = num(ws, row, cols["netdebt"]) or 0.0

    # Discount rate: flat --r, or a per-company WACC from an equity hurdle --re.
    wacc_note = ""
    if args.re is not None:
        country = ws.cell(row=row, column=cols["country"]).value if cols["country"] else None
        cbase, ccy = currency_base(country, parse_kv_rates(args.country_base))
        gross = num(ws, row, cols["gross"])
        tax = num(ws, row, cols["tax"])
        if tax is None or tax >= 1:
            tax = 0.25
        ebit = nopat0 / (1 - tax)
        rd, rating, cov = synthetic_rd(ebit, gross, mktcap, cbase)
        industry_raw = ws.cell(row=row, column=cols["industry"]).value if cols["industry"] else ""
        is_fin = " ".join(str(industry_raw).split()) in FINANCIAL_SECTORS
        crp = country_risk_premium(country, parse_kv_rates(args.country_crp))
        r = firm_wacc(args.re, rd, mktcap, netdebt) if (mktcap and mktcap > 0) else args.re
        r = min(max(r + crp, 0.04), 0.25)
        cv = ">99" if cov == float("inf") else f"{cov:.1f}x"
        wacc_note = (f"  WACC (re {pct(args.re)}, {ccy} base {pct(cbase)}, cov {cv} -> "
                     f"{rating} Rd {pct(rd)}" + (f", +CRP {pct(crp)}" if crp else "")
                     + f") = {pct(r)}"
                     + ("   [FINANCIAL: rating/coverage unreliable]" if is_fin else ""))
    else:
        r = args.r
    if g_term >= r:
        sys.exit(f"g_term ({g_term}) must be < r ({r:.4f}).")

    # Competitive period from the Moat Score: total life = n1 + n2. n1 is a
    # fixed 3y hold of ROICm7; n2 is the rest, fading to the base rate at the
    # moat-tier persistence phi.
    moat = num(ws, row, cols["moat"])
    life = moat_to_life(moat)
    mp = moat_to_cap_persistence(moat)
    if life is not None and mp is not None:
        tier, d_phi = mp[2], mp[1]
        src = f"Moat {moat:.2f} [{tier}], life {life}y"
    else:
        life, d_phi, tier = 15, 0.75, "Narrow(default)"
        src = "no Moat Score; default life 15y"
    dn1 = min(3, max(1, life - 1))        # 3y hold (shrunk only if life is tiny)
    dn2 = max(1, life - dn1)
    n1 = args.n1 if args.n1 is not None else dn1
    n2 = args.n2 if args.n2 is not None else dn2
    phi = args.persistence if args.persistence is not None else d_phi

    industry = ws.cell(row=row, column=cols["industry"]).value if cols["industry"] else None
    base, base_src = base_rate_for(industry, args.base_rate)
    if args.base_rate is None and args.base_inflation:
        base += args.base_inflation
        base_src += f" +{pct(args.base_inflation)} infl"

    sales0 = num(ws, row, cols["sales"]) if cols["sales"] else None
    use_floor = not args.no_sales_floor
    res = value_company(nopat0, roiic0, rr0, r, g_term, n1, n2, phi, base,
                        sales0=sales0, gics_industry=industry, sales_floor=use_floor)
    growth_note = ""
    if use_floor and sales0:
        growth_note = (f"  sales-growth floor [median]: sales {money(sales0)} "
                       f"[{_size_bucket(sales0)}, "
                       f"{GICS_TO_MAUBOUSSIN.get(' '.join(str(industry).split()),'Other')}]"
                       f" -> g_floor {pct(sales_growth_base(sales0, industry, z=0.0))}")
    total = res["total"]

    ticker = ws.cell(row=row, column=cols["ticker"]).value if cols["ticker"] else ""
    print(f"\nAIP VALUE — two-phase ROIIC DCF — {company} {f'({ticker})' if ticker else ''}")
    print(f"WACC r = {pct(r)}   n1(hold)={n1}y  n2(fade)={n2}y  phi={phi:.2f}   "
          f"g_term = {pct(g_term)}   ({src})")
    if wacc_note:
        print(wacc_note)
    if growth_note:
        print(growth_note)
    print("=" * 70)
    print(f"  NOPAT_0 (New Operating Income) .... {money(nopat0)}")
    print(f"  ROIIC_0 (ROICm 7, held in phase 1)  {pct(roiic0)}")
    print(f"  ROIIC base rate (revert target) ... {pct(base)}   [{base_src}]")
    print(f"  RR_0 (RR 7, held; lifted by floor)  {pct(rr0)}")

    print(f"\n  SCHEDULE (hold {n1}y, then fade ROIIC->{pct(base)} at phi={phi:.2f};  "
          f"g = ROIIC*RR_0, floored at the median sales base)")
    sched = res["sched"]
    marks = sorted(set([1, n1, n1 + max(1, n2 // 2), n1 + n2]))
    for t, roiic_t, rr_t, g_t, phase in sched:
        if t in marks:
            print(f"    year {t:>2} [{phase}]:  ROIIC {pct(roiic_t):>8}  RR {pct(rr_t):>7}  g {pct(g_t):>8}")

    print(f"\n  Terminal (competitive equilibrium): RONIC = WACC {pct(r)}, "
          f"g_eff {pct(res['g_eff'])} (GDP), RR_term {pct(res['rr_term'])} "
          f"-> growth value-neutral")
    print(f"  PV(explicit FCF, yrs 1-{n1+n2}) ...... {money(res['pv_explicit'])}")
    print(f"  Terminal value (at yr {n1+n2}) ........ {money(res['tv'])}")
    print(f"  PV(terminal) ...................... {money(res['pv_tv'])}")
    print("  " + "-" * 50)
    print(f"  TOTAL OPERATING VALUE ............. {money(total)}")
    if total:
        print(f"    explicit {res['pv_explicit']/total*100:5.1f}%   "
              f"terminal {res['pv_tv']/total*100:5.1f}%")

    # --- Market comparison ---
    ev = num(ws, row, cols["ev"])
    netdebt = num(ws, row, cols["netdebt"])
    shares = num(ws, row, cols["shares"])
    price = num(ws, row, cols["price"])
    mktcap = num(ws, row, cols["mktcap"])
    print("\n  MARKET COMPARISON")
    if ev is not None:
        upside = total / ev - 1 if ev else None
        print(f"    Enterprise Value (market) ..... {money(ev)}")
        if ev:
            print(f"    Model / EV .................... {total/ev:.2f}x  "
                  f"({'+' if upside>=0 else ''}{upside*100:.1f}% vs EV)")
    if netdebt is not None:
        equity = total - netdebt
        print(f"    Net debt (neg = net cash) ..... {money(netdebt)}")
        print(f"    Implied equity value .......... {money(equity)}")
        if shares:
            iv_ps = equity / shares
            line = f"    Implied value / share ......... {iv_ps:,.2f}"
            if price:
                line += f"   (market {price:,.2f}, {'+' if iv_ps>=price else ''}{(iv_ps/price-1)*100:.1f}%)"
            print(line)
    if mktcap is not None:
        print(f"    Market cap .................... {money(mktcap)}")

    # --- Expected return (IRR) over the holding horizon ---
    cf_for_year = res["cf_for_year"]
    n = args.horizon
    print(f"\n  EXPECTED RETURN  (horizon n = {n}y)")
    cf_sum = sum(cf_for_year(t) for t in range(1, n + 1))
    print(f"    EV_target (operating value) ... {money(total)}")
    print(f"    Sum FCF yrs 1..{n} ............. {money(cf_sum)}")
    if args.payout_total:
        print(f"    - Dividends/buybacks (horizon)  {money(args.payout_total)}")
    if netdebt is not None:
        nd5 = netdebt - (cf_sum - args.payout_total)
        eqv_target = total - nd5
        print(f"    ND_0 (current net debt) ....... {money(netdebt)}")
        print(f"    ND_{n} (after cash sweep) ....... {money(nd5)}")
        print(f"    EqV_target = EV_target - ND_{n} . {money(eqv_target)}")
        if mktcap and mktcap > 0 and eqv_target > 0:
            eq_irr = (eqv_target / mktcap) ** (1 / n) - 1
            print(f"    EqV_0 (current market cap) .... {money(mktcap)}")
            print(f"    >> Expected equity return (IRR) {pct(eq_irr)} / yr")
        elif eqv_target <= 0:
            print("    >> Projected equity value <= 0; equity IRR undefined.")
    if ev is not None and ev > 0 and total > 0:
        unlev = (total / ev) ** (1 / n) - 1
        print(f"    >> Unlevered return (EV_target/EV_0) {pct(unlev)} / yr")
        print("    (If equity IRR >> unlevered, the thesis leans on leverage.)")

    print("\n  Analytical framework output, not investment advice.")


if __name__ == "__main__":
    main()
