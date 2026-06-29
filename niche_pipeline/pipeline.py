"""
pipeline.py — orchestrator.

Reads a list of companies from an Excel screener and, for each one, produces the
full analytic record using the three frameworks plus the AIP valuation:

  * Company moat        — Unified Niche Compounder v3.2 (blended, 10 chapters)
  * Core moat           — identify the durable core, score the core only
  * Ownership block test — can an engaged owner force change, or will a holder veto?
  * AIP valuation       — operating value, expected return at two hurdles
  * Implied moat length — reverse DCF (years of excess return the price implies)
  * Risk tags           — the 11 binary exposure tags + the two entry gates
  * Engaged-Ownership   — Keep Score / Detachability / action for the core

"Fill the missing information" — qualitative inputs (moat chapter scores, risk
tags) are read from the sheet when present and only researched when absent.
Deterministic modules (AIP value, implied moat, gates, sizing) always run.

Output: a new workbook with a per-company "Scored" sheet and a gated, tag-capped,
return-sized "Satellite Book" sheet.
"""
from __future__ import annotations
import argparse
import os

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import frameworks as F
import aip
import analyst
import history

# ---- input column resolution ----------------------------------------------
NCC_COMPANY_COLS = [f"NCC_{c}" for c in F.NCC_CHAPTERS]          # optional pre-scored
NCC_CORE_COLS = [f"Core_{c}" for c in F.NCC_CHAPTERS]            # optional pre-scored


def _index(ws):
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    return {h: i for i, h in enumerate(hdr) if h is not None}, hdr


def _get(row, idx, name):
    i = idx.get(name)
    if i is None or i >= len(row):
        return None
    return row[i]


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _present_scores(row, idx, cols):
    """Return a list of 10 ints if all chapter columns are present & numeric, else None."""
    if not all(c in idx for c in cols):
        return None
    vals = [_num(_get(row, idx, c)) for c in cols]
    if any(v is None for v in vals):
        return None
    return [int(round(v)) for v in vals]


def _present_tags(row, idx):
    """Return 11 binary tags if the named tag columns are present & 0/1, else None."""
    if not all(t in idx for t in F.RISK_TAGS):
        return None
    vals = [_num(_get(row, idx, t)) for t in F.RISK_TAGS]
    if any(v is None for v in vals):
        return None
    return [1 if v >= 0.5 else 0 for v in vals]


def _coerce_vec(v, n):
    """Sanitize a researched score vector to exactly n ints, else None.
    Guards against an analyst returning 9/11 chapters or non-numeric values —
    a malformed vector is dropped (NEEDS_RESEARCH) rather than silently scored."""
    if not isinstance(v, (list, tuple)):
        return None
    out = [_num(x) for x in v]
    out = [x for x in out if x is not None]
    if len(out) < n:
        return None
    return [int(round(x)) for x in out[:n]]


def deterministic_part(row, idx, *, re, re2, country_base, apply_gate2=False,
                       gate_basis="re"):
    """Cheap, no-network: AIP valuation, implied moat, both gates. Always runs.

    gate_basis picks which cost-of-capital read is the company's *expected return*
    for the gate and sizing: "re" = the ER@`re` read (default; the realistic read —
    with a ~2.65% JPY risk-free a 7% cost of equity is appropriate, and a 12%
    discount would imply a ~9% equity-risk premium), "re2" = the conservative
    ER@`re2` read. Gate 1 always tests this return against the 12% HURDLE; the
    other read is still computed and displayed."""
    name = _get(row, idx, "Company Name")
    ticker = _get(row, idx, "Instrument")
    if not name:
        return None
    fin = {fld: _num(_get(row, idx, fld)) if fld not in
           ("Company Name", "Instrument", "GICS Industry Group Name",
            "Country of Headquarters") else _get(row, idx, fld)
           for fld in aip.FIELDS.values()}
    rec = {"ticker": ticker, "company": name,
           "industry": _get(row, idx, "GICS Industry Group Name"),
           "country": _get(row, idx, "Country of Headquarters"),
           "_row": row, "_apply_gate2": apply_gate2, "_gb": gate_basis}
    rec["_fin"] = fin; rec["_re"] = re; rec["_re2"] = re2; rec["_cb"] = country_base
    val = aip.value_and_return(fin, re=re, re2=re2, country_base=country_base)
    if val:
        rec.update({"wacc": val["wacc"], "aip_rating": val["rating"],
                    "op_value": val["op_value"], "ev": val["ev"],
                    "er_lo": val.get("er1"), "er_hi": val.get("er2"),
                    "op_over_ev": (val["op_value"] / val["ev"]) if val.get("op_value") and val.get("ev") else None})
        rec["implied_moat"] = aip.implied_moat(fin, r=re2 or 0.12, country_base=country_base)
    rec["irr12"] = rec.get("er_hi")           # ER at the 12% discount (conservative read; display)
    bkey = "er1" if gate_basis == "re" else "er2"
    # The as-is expected return that drives the gate, the artifact screen and sizing.
    rec["as_is_ret"] = rec.get("er_lo") if gate_basis == "re" else rec.get("er_hi")
    rec["er_effective"] = rec.get("as_is_ret")   # may be raised to the separated return later
    # Cheap pre-screen: would this clear the hurdle IF its core were a top compounder?
    vc = aip.value_and_return(fin, re=re, re2=re2, country_base=country_base,
                              moat_override=F.RESTRUCTURE_MIN_CORE + 1.3)   # ~7.8 "Compounder"
    rec["er_ceiling"] = vc.get(bkey) if vc else None
    floor_ratio = _num(_get(row, idx, "Value per share without growth/share price"))
    rec["gate2_floor"] = floor_ratio
    rec["gate2_pass"] = F.gate2_pass(floor_ratio)
    _finalize_gates(rec)
    # Research-worthy if it clears Gate 1 as-is OR could clear once a high-quality
    # core is recognised (the trapped-jewel case) — both deterministic, no spend.
    rec["research_worthy"] = bool(
        (F.gate1_pass(rec.get("as_is_ret")) or F.gate1_pass(rec.get("er_ceiling")))
        and not F.er_is_artifact(rec.get("as_is_ret")))
    return rec


def _gate2_effective(rec):
    """Gate 2 is disabled by default (its no-growth-floor proxy is not a sound
    downside measure). The floor is still computed and reported; it only gates
    selection when --gate2 is passed."""
    return rec.get("gate2_pass") if rec.get("_apply_gate2") else True


def _finalize_gates(rec):
    """Set Gate 1, the ER-artifact flag and the gauntlet verdict. Gate 1 uses the
    effective return (as-is, or the separated return for a restructuring candidate);
    the artifact flag screens the AS-IS return only — that is where a depressed-
    earnings trap shows up, whereas a high separated return is a legitimate moat
    re-rating, not an artifact."""
    rec["gate1_irr12"] = F.gate1_pass(rec.get("er_effective"))
    rec["er_artifact"] = F.er_is_artifact(rec.get("as_is_ret"))
    rec["clears_gauntlet"] = bool(rec["gate1_irr12"] and _gate2_effective(rec)
                                  and not rec["er_artifact"])


def attach_qualitative(rec, idx, research_rec):
    """Add moat / core / ownership / risk to a deterministic record. Uses scores
    already in the sheet when present; otherwise the (sanitized) research record."""
    row = rec["_row"]
    fin_net_debt = _num(_get(row, idx, "Net debt"))
    comp_ch = _present_scores(row, idx, NCC_COMPANY_COLS)
    core_ch = _present_scores(row, idx, NCC_CORE_COLS)
    tags = _present_tags(row, idx)
    detach = {}
    if research_rec:
        comp_ch = comp_ch or _coerce_vec(research_rec.get("chapters_company"), 10)
        core_ch = core_ch or _coerce_vec(research_rec.get("chapters_core"), 10)
        tags = tags or _coerce_vec(research_rec.get("risk_tags"), 11)
        rec["core_business"] = research_rec.get("core_business")
        rec["owner_bloc_pct"] = research_rec.get("largest_bloc_pct")
        rec["owner_notes"] = research_rec.get("ownership_notes")
        detach = research_rec.get("core_detach") or {}

    # ---- Niche Compounder v3.2 ----
    if comp_ch and len(comp_ch) == 10:
        rec["company_moat"] = F.ncc_score(comp_ch)
        rec["company_band"] = F.ncc_band(rec["company_moat"])
        rec["company_chapters"] = comp_ch
    if core_ch and len(core_ch) == 10:
        rec["core_moat"] = F.ncc_score(core_ch)
        rec["core_band"] = F.ncc_band(rec["core_moat"])
        rec["core_chapters"] = core_ch
        # ---- Engaged Ownership v3.0 for the core ----
        rec["keep_score"] = F.keep_score(core_ch)
        ent = detach.get("entanglement"); capi = detach.get("capital_independence")
        cori = detach.get("core_independence")
        if None not in (ent, capi, cori):
            rec["detachability"] = F.detachability_score(ent, capi, cori)
            rec["eo_action"] = F.eo_action(rec["keep_score"], rec["detachability"], cori, capi)

    # ---- Risk tags (+ funding tag computable from financials) ----
    if tags and len(tags) == 11:
        ft = F.funding_tag_from_financials(fin_net_debt,
                                           _num(_get(row, idx, "EBITA_Avg_Last_7yr")))
        if ft is not None:
            tags[5] = ft                         # override #6 with the hard number
        rec["risk_tags"] = tags
        rec["risk_tag_names"] = [F.RISK_TAGS_SHORT[i] for i, v in enumerate(tags) if v]

    # ---- Ownership block verdict ----
    rec["owner_verdict"] = F.ownership_verdict(rec.get("owner_bloc_pct"), rec.get("country"))

    # ---- Restructuring lens + two-path selection ----
    cm, km = rec.get("company_moat"), rec.get("core_moat")
    if cm is not None and km is not None:
        rec["moat_gap"] = round(km - cm, 2)
        # "Has a strong, freeable core" — informational; flagged even when the name
        # already clears as-is (the separation is then optional upside, not needed).
        rec["restructuring"] = F.is_restructuring_candidate(cm, km, rec["owner_verdict"])
        # Value the durable core on its OWN moat (a longer competitive-advantage
        # period) — the separated re-rating the whole-company return misses. Use
        # the same cost-of-capital basis as the as-is return so they compare like
        # for like.
        bkey = "er1" if rec.get("_gb", "re") == "re" else "er2"
        vc = aip.value_and_return(rec["_fin"], re=rec["_re"], re2=rec["_re2"],
                                  country_base=rec["_cb"], moat_override=km)
        rec["er_core"] = vc.get(bkey) if vc else None
        asis = rec.get("as_is_ret")
        if asis is not None and rec.get("er_core") is not None:
            rec["separation_uplift"] = round(rec["er_core"] - asis, 4)
        # AS-IS winner (any ownership) OR freed-core (separable only).
        clears, basis, eff = F.select(
            asis, rec.get("er_core"), cm, km, rec["owner_verdict"],
            _gate2_effective(rec), F.er_is_artifact(asis))
        rec["clears_gauntlet"] = clears
        rec["return_basis"] = basis
        rec["er_effective"] = eff
        rec["gate1_irr12"] = F.gate1_pass(asis)
        rec["er_artifact"] = F.er_is_artifact(asis)

    # ---- Implied-moat regime: is the return cheapness, or borrowed from the discount? ----
    moat_for_life = rec.get("core_moat") if rec.get("core_moat") is not None else rec.get("company_moat")
    if moat_for_life is not None:
        iy = F.implied_moat_years(rec.get("implied_moat"))
        warr = aip.warranted_life(moat_for_life)
        rec["implied_years"] = iy
        rec["warranted_life"] = warr
        rec["valuation_cushion"] = F.valuation_cushion(iy, warr)
        rec["priced_for"] = "durability" if F.priced_for_perfection(iy, warr) else "fade"
        # ER@re (realistic, = as_is_ret when gate_basis='re') vs ER@re2 (conservative, irr12)
        rec["return_leans_on_discount"] = F.return_leans_on_discount(
            rec.get("as_is_ret"), rec.get("irr12"), rec.get("valuation_cushion"))
    return rec


# ---------------------------------------------------------------------------
def build_satellite(records, exclude_tags=(), max_names=12, segment_tags=True):
    """From the names that clear the gauntlet, build the concentrated book:
    return-sized, 20%-tag-capped via the Trim Protocol, capped at `max_names`.

    Eligibility (the moat floor and the HARD-BLOCK rule) is decided per-name in
    frameworks.select(): an AS-IS winner qualifies on its own return regardless of
    ownership, while a FREED-CORE pick must be separable. So a founder-controlled
    name with a high as-is return belongs here; one that needs a break-up to work
    does not.

    `segment_tags` (default True) qualifies each business risk tag by the name's
    industry regime, so the 20% rule only caps genuinely correlated exposure
    (pharma drug-pricing risk does not collide with gambling-licence risk); the
    market-wide tags Macro and Funding stay un-segmented. Without it, one coarse
    "Regulatory"/"Demand" tag shared across most quality names collapses the
    sized book to a single position.
    `exclude_tags` (short names) are not counted toward the 20% rule — e.g. pass
    ("FX",) for an all-domestic/heavy-exporter book where FX is shared, not idiosyncratic."""
    elig = [r for r in records
            if r.get("clears_gauntlet") and r.get("er_effective") and r.get("risk_tags")
            and r.get("moat_vs_history") != "CONTRADICTED"]   # economics veto the moat
    if not elig:
        return [], {}, []
    elig.sort(key=lambda r: -r["er_effective"])
    ex = set(exclude_tags)
    # All business tags are regime-specific: aluminium-price, interest-rate and
    # auto-cycle "Macro" exposures are uncorrelated, as are pharma-pricing vs
    # gambling-licence "Regulatory". Segment every tag by industry so the 20% rule
    # caps only genuinely correlated wipeout risk.
    systemic = set()

    def tagset(r):
        ind = " ".join(str(r.get("industry") or "?").split())
        out = set()
        for t in r.get("risk_tag_names", []):
            if t in ex:
                continue
            out.add(f"{t}:{ind}" if (segment_tags and t not in systemic) else t)
        return out

    # Size on the effective return (separated re-rating for a trapped jewel, else
    # as-is), capped at the plausibility ceiling so an optimistic separated return
    # cannot dominate the interpolation.
    irr = {r["ticker"]: min(r["er_effective"], F.MAX_PLAUSIBLE_IRR) for r in elig}
    tags_by = {r["ticker"]: tagset(r) for r in elig}

    # Fixed return-interpolated weights anchored to the 12% hurdle (5% floor) and
    # the best eligible return (20% ceiling). Fixing the scale — rather than
    # re-deriving R_min from the book each step — keeps a name's weight independent
    # of which others are held, so adding a name never re-weights the rest.
    rmax = max(irr.values())
    rmin = F.GATE1_IRR

    def wt(er):
        if rmax <= rmin:
            return F.W_MIN
        return min(F.W_MAX, max(F.W_MIN,
                   F.W_MIN + (er - rmin) / (rmax - rmin) * (F.W_MAX - F.W_MIN)))

    # Greedy return-ranked fill: walk highest-return first and add each name only
    # if it keeps every (industry-segmented) tag bucket within the 20% cap, until
    # the book reaches `max_names`. Buckets only grow, so the fill is stable and
    # builds a full, diversified 8-12 name book — a name that would breach a maxed
    # tag-regime is skipped for the next that fits.
    book, log, weights, buckets = [], [], {}, {}
    for r in elig:                                    # already sorted by return
        t = r["ticker"]; w = wt(irr[t])
        breach = next((tag for tag in tags_by[t]
                       if buckets.get(tag, 0.0) + w > F.TAG_CAP + 1e-9), None)
        if breach is None:
            book.append(t); weights[t] = w
            for tag in tags_by[t]:
                buckets[tag] = buckets.get(tag, 0.0) + w
        else:
            log.append((breach, round(buckets.get(breach, 0.0) + w, 4), t, round(irr[t], 4)))
        if len(book) >= max_names:
            break
    # The concentrated book cannot exceed 100% of capital. When the return-sized
    # weights sum above 1.0 (many high-return ideas), scale them to a 100% ceiling,
    # preserving their return-based relative emphasis; otherwise leave them and the
    # remainder sweeps to the Core index.
    tot = sum(weights.values())
    if tot > 1.0:
        weights = {t: w / tot for t, w in weights.items()}
    return book, weights, log


# ---------------------------------------------------------------------------
def build_core(records, n=F.CORE_N, slot_cap=F.CORE_SLOT_CAP, exclude_tags=(),
               segment_tags=True):
    """The Core Index — the diversified, EQUAL-weighted book (vs the return-sized
    Satellite). Same eligibility as the Satellite (clears the gauntlet, has tags,
    not history-CONTRADICTED), ranked by effective return, filled under the 20%
    tag rule as a slot cap: no tag in more than `slot_cap` of the `n` names.

    `segment_tags` (default) qualifies each tag by industry so the cap bounds only
    correlated wipeout risk; pass segment_tags=False for the STRICT raw-tag budget
    (the framework's literal reading), which in a quality universe binds hard and
    typically yields far fewer than `n` names — that shortfall is the real signal.
    Returns (tickers, weight_each, skipped_log, achievable_n)."""
    elig = [r for r in records
            if r.get("clears_gauntlet") and r.get("er_effective") and r.get("risk_tags")
            and r.get("moat_vs_history") != "CONTRADICTED"]
    if not elig:
        return [], 0.0, [], 0
    elig.sort(key=lambda r: -min(r["er_effective"], F.MAX_PLAUSIBLE_IRR))
    ex = set(exclude_tags)

    def tagset(r):
        ind = " ".join(str(r.get("industry") or "?").split())
        return {(f"{t}:{ind}" if segment_tags else t)
                for t in r.get("risk_tag_names", []) if t not in ex}

    book, skipped = F.fill_under_slot_cap(elig, tagset, n, slot_cap)
    w = 1.0 / len(book) if book else 0.0
    log = [(tag, r["ticker"], round(r["er_effective"], 4)) for tag, r in skipped]
    return [r["ticker"] for r in book], w, log, len(book)


def _style_header(ws):
    fill = PatternFill("solid", fgColor="1F4E78"); font = Font(color="FFFFFF", bold=True, size=10)
    thin = Side(style="thin", color="CCCCCC"); border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in ws[1]:
        c.fill = fill; c.font = font; c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "C2"


def write_output(records, path, exclude_tags=(), segment_tags=True,
                 core_n=F.CORE_N, core_strict=False):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Scored"
    cols = ["Ticker", "Company", "Industry", "Country",
            "CompanyMoat(v3.2)", "Band", "CoreBusiness", "CoreMoat(v3.2)", "CoreBand",
            "KeepScore", "Detachability", "EO_Action",
            "OwnerVerdict", "OwnerBloc%", "OwnerNotes",
            "AIP_WACC", "AIP_Rating", "OpValue", "OpVal/EV", "ER@7%", "ER@12%(as-is)", "ImpliedMoat",
            "MoatGap", "ER(separated)", "SeparationUplift", "ReturnBasis", "Restructuring",
            "Gate1(IRR>=12)", "Gate2(floor)", "Gate2pass", "ER_Artifact", "ClearsGauntlet", "RiskTags",
            "MargROIC7y", "MarginTrend", "NormEV/EBITA", "MoatVsHistory", "HistoryNote",
            "ImpliedYrs", "WarrantedYrs", "ValuationCushion", "PricedFor", "ReturnFromDiscount",
            "ROIC*(7y)", "OE_RevExcess", "OE_Barrier", "OE_Faded%", "OE_H(y)", "OE_ER_adj",
            "Dmd_Channel", "Dmd_EffMoat", "Dmd_Comp", "Dmd_Drop%"]
    ws.append(cols)
    for r in sorted(records, key=lambda r: (r.get("er_effective") is not None, r.get("er_effective") or -9), reverse=True):
        ws.append([
            r.get("ticker"), r.get("company"), r.get("industry"), r.get("country"),
            r.get("company_moat"), r.get("company_band"), r.get("core_business"),
            r.get("core_moat"), r.get("core_band"), r.get("keep_score"),
            r.get("detachability"), r.get("eo_action"),
            r.get("owner_verdict"), r.get("owner_bloc_pct"), r.get("owner_notes"),
            r.get("wacc"), r.get("aip_rating"), r.get("op_value"), r.get("op_over_ev"),
            r.get("er_lo"), r.get("irr12"), r.get("implied_moat"),
            r.get("moat_gap"), r.get("er_core"), r.get("separation_uplift"),
            r.get("return_basis"), "TRAPPED-JEWEL" if r.get("restructuring") else "",
            "PASS" if r.get("gate1_irr12") else "fail",
            r.get("gate2_floor"), "PASS" if r.get("gate2_pass") else "fail",
            "ARTIFACT" if r.get("er_artifact") else "",
            "YES" if r.get("clears_gauntlet") else "no",
            ", ".join(r.get("risk_tag_names", [])),
            r.get("hist_marginal_roic"), r.get("hist_margin_trend"),
            r.get("hist_norm_ev_ebita"), r.get("moat_vs_history"), r.get("hist_reasons"),
            r.get("implied_years"), r.get("warranted_life"), r.get("valuation_cushion"),
            r.get("priced_for"), "YES" if r.get("return_leans_on_discount") else "",
            r.get("roic_star"), r.get("oe_rev_excess"), r.get("oe_barrier"),
            r.get("oe_faded"), r.get("oe_H"), r.get("oe_er_adj"),
            r.get("dmd_channel"), r.get("dmd_eff_moat"), r.get("dmd_comp"), r.get("dmd_drop"),
        ])
    _style_header(ws)
    for row in ws.iter_rows(min_row=2):
        for j in (15, 18, 19, 20, 23, 24, 33, 34):   # wacc, opval/ev, ers, marg-roic, margin-trend
            if row[j].value is not None:
                row[j].number_format = "0.0%"
    widths = [9, 26, 24, 12, 14, 16, 30, 13, 14, 10, 12, 22, 13, 10, 40,
              9, 9, 12, 9, 8, 12, 9, 8, 14, 12, 11, 14, 13, 11, 10, 11, 14, 40,
              11, 11, 12, 14, 44, 10, 12, 16, 11, 17,
              10, 12, 11, 10, 8, 11]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ---- Satellite Book ----
    book, weights, log = build_satellite(records, exclude_tags, segment_tags=segment_tags)
    ws2 = wb.create_sheet("Satellite Book")
    ws2.append(["TargetWeight", "Ticker", "Company", "ER@12%(eff)", "Basis", "CoreMoat",
                "OwnerVerdict", "RiskTags"])
    by = {r["ticker"]: r for r in records}
    for t in sorted(book, key=lambda t: -weights[t]):
        r = by[t]
        ws2.append([weights[t], t, r["company"], r.get("er_effective"), r.get("return_basis"),
                    r.get("core_moat"), r.get("owner_verdict"), ", ".join(r.get("risk_tag_names", []))])
    tot = sum(weights.values())
    ws2.append([tot, "TOTAL CONCENTRATED", f"Core sweep {1-tot:.1%}", "", "", "", "", ""])
    _style_header(ws2)
    for row in ws2.iter_rows(min_row=2):
        for j in (0, 3):
            if isinstance(row[j].value, (int, float)):
                row[j].number_format = "0.0%"
    for i, w in enumerate([13, 9, 26, 11, 11, 9, 14, 40], 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    if log:
        ws2.append([]); ws2.append(["Trim Protocol log (tag, bucket%, liquidated, its IRR):"])
        for tag, bw, drop, dirr in log:
            ws2.append(["", tag, f"{bw:.0%}", by[drop]["company"], f"{dirr:.1%}"])

    # ---- Core Index (equal-weighted, history-vetoed, 6/30 tag rule) ----
    core, cw, _, achievable = build_core(records, n=core_n,
                                         exclude_tags=exclude_tags,
                                         segment_tags=not core_strict and segment_tags)
    ws3 = wb.create_sheet("Core Index")
    rule = "strict raw-tag" if core_strict else "industry-segmented"
    ws3.append(["Rank", "Weight", "Ticker", "Company", "Industry", "ER(eff)", "CoreMoat",
                "ImpliedMoat", "ValuationCushion", "PricedFor", "RetFromDiscount",
                "History", "Basis", "Ownership", "RiskTags"])
    core_recs = [by[t] for t in core]
    er_book = (sum(min(r["er_effective"], F.MAX_PLAUSIBLE_IRR) for r in core_recs)
               / len(core_recs)) if core_recs else 0.0
    for i, r in enumerate(core_recs, 1):
        ws3.append([i, cw, r["ticker"], r["company"], r.get("industry"),
                    r.get("er_effective"), r.get("core_moat"), r.get("implied_moat"),
                    r.get("valuation_cushion"), r.get("priced_for"),
                    "YES" if r.get("return_leans_on_discount") else "",
                    r.get("moat_vs_history"), r.get("return_basis"), r.get("owner_verdict"),
                    ", ".join(r.get("risk_tag_names", []))])
    ws3.append([])
    ws3.append(["", "", f"{achievable} names", f"({rule} 20% budget, target {core_n})",
                "", er_book, "", "", "", "", "", "", "", "EQUAL-WT EXPECTED RETURN", ""])
    _style_header(ws3)
    for row in ws3.iter_rows(min_row=2):
        for j in (1, 5):
            if isinstance(row[j].value, (int, float)):
                row[j].number_format = "0.0%"
    for i, w in enumerate([5, 8, 9, 26, 30, 9, 9, 11, 16, 11, 15, 11, 11, 12, 42], 1):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # ---- Risk Exposure of the Core Index (raw factor counts) ----
    ws4 = wb.create_sheet("Risk Exposure")
    nb = max(1, len(core_recs))
    counts = {t: 0 for t in F.RISK_TAGS_SHORT}
    for r in core_recs:
        for t in r.get("risk_tag_names", []):
            counts[t] += 1
    ws4.append(["Risk Tag", f"Names (of {len(core_recs)})", "% of book",
                "within 20%? (raw)", "Note"])
    ex = set(exclude_tags)
    for t in F.RISK_TAGS_SHORT:
        c = counts[t]; frac = c / nb
        within = "excluded" if t in ex else ("YES" if frac <= 0.2001 else "no (raw)")
        ws4.append([t, c, frac, within, ""])
    ws4.append([])
    seg_note = ("Selection enforces the 20% cap on INDUSTRY-SEGMENTED tags, so raw "
                "aggregates above 20% (Demand/Regulatory/KeyPerson) are spread across "
                "uncorrelated industries/founders — not a single correlated wipeout."
                if not core_strict else
                "STRICT mode: the 20% cap is enforced on RAW tags; the book size is "
                "capped wherever the universe cannot supply more names within budget.")
    ws4.append([seg_note])
    _style_header(ws4)
    for row in ws4.iter_rows(min_row=2, max_row=1 + len(F.RISK_TAGS_SHORT)):
        if isinstance(row[2].value, (int, float)):
            row[2].number_format = "0%"
    for i, w in enumerate([14, 14, 10, 18, 60], 1):
        ws4.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    wb.save(path)
    return book, weights, core, achievable


# ---------------------------------------------------------------------------
def run(input_path, output_path, *, sheet=None, re=0.07, re2=0.12,
        research=True, limit=None, country_base=None, exclude_tags=(),
        workers=6, research_all=False, history_path=None, apply_gate2=False,
        segment_tags=True, gate_basis="re", core_n=F.CORE_N, core_strict=False,
        apply_overearning=True):
    """Two-pass, hands-off run.

    Pass 1 (cheap, no network): value every company and apply the gates.
    Pass 2 (the only paid step): research the moat/ownership/risk for the names
    that survive Gate 1 (or all, with research_all) — concurrently, so a full
    universe completes unattended. Researching only gate-survivors is itself a
    guardrail: no spend on names that already fail the return hurdle.
    """
    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb[sheet] if sheet else wb.worksheets[0]
    idx, _ = _index(ws)

    # ---- Pass 1: deterministic ----
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = deterministic_part(row, idx, re=re, re2=re2, country_base=country_base,
                                 apply_gate2=apply_gate2, gate_basis=gate_basis)
        if rec:
            records.append(rec)
        if limit and len(records) >= limit:
            break
    survivors = [r for r in records if (research_all or r.get("research_worthy"))]
    jewels = sum(1 for r in survivors if not r["gate1_irr12"])
    print(f"Pass 1: valued {len(records)} companies | researching {len(survivors)} "
          f"(Gate-1 survivors + {jewels} potential trapped-jewels)"
          + (" [research-all]" if research_all else ""))

    # ---- Pass 2: concurrent research (only when enabled) ----
    fetched = {}
    if research:
        from concurrent.futures import ThreadPoolExecutor, as_completed
        def fetch(r):
            return r["ticker"], analyst.research(r["ticker"], r["company"],
                                                 hint=str(r.get("industry") or ""))
        with ThreadPoolExecutor(max_workers=max(1, workers)) as ex:
            futs = [ex.submit(fetch, r) for r in survivors]
            for i, fut in enumerate(as_completed(futs), 1):
                try:
                    t, rr = fut.result()
                    fetched[t] = rr
                except Exception as e:                      # one failure can't sink the run
                    print(f"  research error: {e}")
                if i % 10 == 0 or i == len(futs):
                    print(f"  researched {i}/{len(futs)}")

    # ---- Finalize qualitative scoring ----
    for r in records:
        attach_qualitative(r, idx, fetched.get(r["ticker"]))

    # ---- Empirical history cross-check (Domain B: actual ROIC beats narrative) ----
    if history_path:
        try:
            by, hidx = history.load_panel(history_path)
            for r in records:
                rows = by.get(r["ticker"])
                if not rows:
                    continue
                s = history.summarize(rows, hidx)
                v, reasons = history.verdict(s, r.get("core_moat"))
                r["hist_marginal_roic"] = s["roic_marginal_7y"]
                r["hist_margin_trend"] = s["margin_ratio"]
                r["hist_norm_ev_ebita"] = s["norm_ev_ebita"]
                r["moat_vs_history"] = v
                r["hist_reasons"] = "; ".join(reasons)
            tally = {k: sum(1 for r in records if r.get("moat_vs_history") == k)
                     for k in ("CONFIRMED", "SOFT", "CONTRADICTED")}
            print(f"History cross-check: {tally}  (CONTRADICTED names are vetoed from the book)")
        except Exception as e:
            print(f"history cross-check skipped: {e}")

        # ---- Supply + demand normalization (two-stage) ----
        # Two mirror models, both run inside overearning.two_stage_return:
        #  SUPPLY SIDE (capitalcycle): NOI applies a 7yr-averaged margin to CURRENT
        #    revenue, so a name over-earning on a transient supply shortage is priced
        #    off an inflated base; the scarcity rent is faded DOWN to the reproduction
        #    equilibrium over its gestation lead time.
        #  DEMAND SIDE (softwarecycle): a genAI-tagged software franchise's moat is
        #    being structurally compressed; its competitive-advantage period fades
        #    FASTER (a moat override on the valuation).
        # er_adj reflects both; the books rank on it. Needs the same panel as --history.
        if apply_overearning:
            try:
                import overearning
                n_faded = 0; n_compressed = 0
                for r in records:
                    rows = by.get(r["ticker"]); fin = r.get("_fin")
                    if not rows or not fin or r.get("er_effective") is None:
                        continue
                    f = dict(fin)
                    if r.get("core_moat") is not None:        # use the researched moat
                        f[aip.FIELDS["moat"]] = r["core_moat"]
                    sig = overearning.panel_signals(rows, hidx)
                    ts = overearning.two_stage_return(f, sig, re=re, re2=re2)
                    if not ts:
                        continue
                    r["roic_star"] = sig.get("roic_star")
                    r["oe_rev_excess"] = ts.get("rev_excess")
                    r["oe_barrier"] = ts.get("barrier")
                    r["oe_faded"] = ts.get("faded_frac")
                    r["oe_H"] = ts.get("H")
                    r["oe_er_adj"] = ts.get("er_adj")
                    # demand side (softwarecycle genAI moat compression)
                    r["dmd_channel"] = ts.get("demand_channel")
                    r["dmd_eff_moat"] = ts.get("eff_moat")
                    r["dmd_comp"] = ts.get("moat_comp")
                    r["dmd_drop"] = ts.get("demand_drop")
                    # er_adj reflects BOTH models, so this single downgrade captures the
                    # supply-side scarcity-rent fade AND the demand-side moat compression.
                    dn = ts["er_current"] - ts["er_adj"]
                    if dn > 1e-4:
                        r["er_effective"] = r["er_effective"] - dn
                        n_faded += 1
                        if ts.get("moat_comp", 0) > 1e-9:
                            n_compressed += 1
                        if not F.gate1_pass(r["er_effective"]):
                            r["gate1_irr12"] = False
                            r["clears_gauntlet"] = False
                print(f"Supply+demand normalization: adjusted {n_faded} names "
                      f"({n_compressed} via genAI demand-side moat compression); "
                      f"books rank on the adjusted return")
            except Exception as e:
                print(f"over-earning normalization skipped: {e}")

    for r in records:                                       # drop internal handles before writing
        for k in ("_row", "_fin", "_re", "_re2", "_cb", "_apply_gate2", "_gb"):
            r.pop(k, None)

    book, weights, core, achievable = write_output(
        records, output_path, exclude_tags, segment_tags=segment_tags,
        core_n=core_n, core_strict=core_strict)
    print(f"\nScored {len(records)} companies -> {output_path}")
    art = sum(1 for r in records if r.get("er_artifact"))
    print(f"Flagged {art} ER-artifact names (IRR > {F.MAX_PLAUSIBLE_IRR:.0%}, excluded from book)")
    lean = sum(1 for r in records if r.get("clears_gauntlet") and r.get("return_leans_on_discount"))
    print(f"Flagged {lean} priced-for-perfection names (return leans on the discount rate, no valuation cushion)")
    print(f"Satellite book ({len(book)} names): " +
          ", ".join(f"{by}={weights[by]:.0%}" for by in sorted(book, key=lambda t: -weights[t])))
    rule = "strict raw-tag" if core_strict else "segmented"
    print(f"Core Index ({rule} 20% budget, target {core_n}): {achievable} names, "
          f"equal-weighted {1.0/max(1,achievable):.1%} each")
    return records


def _fmt(x):
    return f"{x*100:.1f}%" if isinstance(x, (int, float)) else "n/a"


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("input", help="Excel screener with the list of companies")
    ap.add_argument("-o", "--output", default="scored_output.xlsx")
    ap.add_argument("--sheet", default=None)
    ap.add_argument("--re", type=float, default=0.07, help="low equity hurdle (ER1 / WACC)")
    ap.add_argument("--re2", type=float, default=0.12, help="high hurdle (ER2 = Gate-1 IRR)")
    ap.add_argument("--no-research", action="store_true",
                    help="skip the LLM analyst; use only scores already in the sheet")
    ap.add_argument("--limit", type=int, default=None)
    ap.add_argument("--country-base", default=None, help='e.g. "JPY=0.0265,USD=0.041"')
    ap.add_argument("--exclude-tags", default="", help="comma short tags excluded from the 20%% rule, e.g. FX")
    ap.add_argument("--workers", type=int, default=6, help="concurrent research calls")
    ap.add_argument("--research-all", action="store_true",
                    help="research every name, not just Gate-1 survivors (slower/costlier)")
    ap.add_argument("--history", default=None,
                    help="long-run panel (time-series) xlsx to cross-check moats vs actual ROIC")
    ap.add_argument("--gate2", action="store_true",
                    help="re-enable the Gate-2 no-growth-floor downside screen (off by default)")
    ap.add_argument("--no-segment-tags", action="store_true",
                    help="disable industry-regime tag segmentation for the 20%% rule")
    ap.add_argument("--conservative-er", action="store_true",
                    help="gate on the ER@re2 (12%% discount) read instead of the realistic "
                         "ER@re (7%%) read; the 12%% hurdle is unchanged either way")
    ap.add_argument("--core-n", type=int, default=F.CORE_N,
                    help="target size of the equal-weighted Core Index (default 30)")
    ap.add_argument("--core-strict", action="store_true",
                    help="enforce the 20%% tag rule on RAW (un-segmented) tags in the Core "
                         "Index — the stringent budget; usually yields fewer than --core-n names")
    ap.add_argument("--no-overearning", action="store_true",
                    help="disable the revenue over-earning normalization (requires --history); "
                         "by default, when a panel is supplied the books rank on the over-"
                         "earning-adjusted return")
    args = ap.parse_args()
    run(args.input, args.output, sheet=args.sheet, re=args.re, re2=args.re2,
        research=not args.no_research, limit=args.limit, country_base=args.country_base,
        exclude_tags=tuple(t.strip() for t in args.exclude_tags.split(',') if t.strip()),
        workers=args.workers, research_all=args.research_all, history_path=args.history,
        apply_gate2=args.gate2, segment_tags=not args.no_segment_tags,
        gate_basis="re2" if args.conservative_er else "re",
        core_n=args.core_n, core_strict=args.core_strict,
        apply_overearning=not args.no_overearning)


if __name__ == "__main__":
    main()
