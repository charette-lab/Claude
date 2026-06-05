#!/usr/bin/env python3
"""AIP — implied moat length (reverse of the aip-value DCF).

Instead of deriving the competitive-advantage period from the Moat Score, this
inverts the aip-value model: it holds the discount rate at YOUR REQUIRED RETURN
and solves for the total moat life n1 + n2 (years) that makes the model's
operating value equal the company's current enterprise value.

  "What competitive-advantage period is the market price baking in, if I am to
   earn exactly my required return?"

Same framework as aip-value (imported directly), with the same drivers:
  NOPAT_0 <- New Operating Income ; ROIIC_0 <- ROICm 7 ; RR_0 <- RR 7
  base    <- sector CFROI median (GICS) ; phi <- Moat-Score tier
  n1 = 3y hold (fixed); we solve for n2 so that total value == market EV.
Target EV = Market Cap + Net debt (net cash reduces EV).

Usage:
  python3 moat_length.py <file.xlsx> "<company>" [--r 0.12] [options]
  python3 moat_length.py <file.xlsx> --list
Options:
  --r FLOAT (required return, default 0.12) --gterm FLOAT --n1 INT
  --persistence FLOAT --base-rate FLOAT --base-inflation FLOAT
  --sheet NAME --col-* STR
"""

import argparse
import importlib.util
import os
import sys


def load_aip_value():
    """Import the aip-value engine (roiic_dcf.py) from a sibling skill dir."""
    here = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.path.join(here, "..", "aip-value", "roiic_dcf.py"),
        os.path.expanduser("~/.claude/skills/aip-value/roiic_dcf.py"),
    ]
    for path in candidates:
        if os.path.exists(path):
            spec = importlib.util.spec_from_file_location("aipval", path)
            mod = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(mod)
            return mod
    sys.exit("Could not find the aip-value skill (roiic_dcf.py). Install it "
             "alongside this skill (../aip-value/ or ~/.claude/skills/aip-value/).")


def main():
    m = load_aip_value()
    ap = argparse.ArgumentParser(add_help=True, description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("file")
    ap.add_argument("company", nargs="?", default=None)
    ap.add_argument("--list", action="store_true", help="list companies and exit")
    ap.add_argument("--r", type=float, default=0.12, help="flat required return (firm-level)")
    ap.add_argument("--re", type=float, default=None,
                    help="required EQUITY return; if set, the discount rate becomes a "
                         "per-company WACC (synthetic credit rating + country risk-free)")
    ap.add_argument("--country-base", default=None,
                    help='refresh risk-free bases, e.g. "EUR=0.0303,USD=0.0405"')
    ap.add_argument("--col-country", default="Country of Headquarters")
    ap.add_argument("--col-gross", default="Gross debt")
    ap.add_argument("--col-tax", default="Income Tax Rate - Instrument")
    ap.add_argument("--gterm", type=float, default=0.025)
    ap.add_argument("--n1", type=int, default=3, help="fixed hold years (default 3)")
    ap.add_argument("--persistence", type=float, default=None,
                    help="persistence phi; if omitted, from the Moat Score tier")
    ap.add_argument("--base-rate", type=float, default=None)
    ap.add_argument("--base-inflation", type=float, default=0.0)
    ap.add_argument("--max-n2", type=int, default=200, help="search ceiling for n2")
    ap.add_argument("--sheet", default=None)
    ap.add_argument("--col-nopat", default="New Operating Income")
    ap.add_argument("--col-roiic", default="ROICm 7")
    ap.add_argument("--col-rr", default="RR 7")
    ap.add_argument("--col-name", default="Company Name")
    ap.add_argument("--col-moat", default="Moat Score")
    ap.add_argument("--col-industry", default="GICS Industry Group Name")
    args = ap.parse_args()

    import openpyxl
    wb = openpyxl.load_workbook(args.file, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb.worksheets[0]
    cols = m.find_columns(ws, {
        "name": args.col_name, "nopat": args.col_nopat, "roiic": args.col_roiic,
        "rr": args.col_rr, "moat": args.col_moat, "industry": args.col_industry,
        "country": args.col_country, "gross": args.col_gross, "tax": args.col_tax,
        "ev": "EV", "mktcap": "Market Cap", "netdebt": "Net debt", "ticker": "Instrument",
    })
    if cols["name"] is None:
        sys.exit(f'Could not find a "{args.col_name}" column in row 1.')

    if args.list or not args.company:
        print(f"Companies in '{ws.title}':")
        for rr in range(2, ws.max_row + 1):
            v = ws.cell(row=rr, column=cols["name"]).value
            if v:
                print(f"  {v}")
        return

    matches = m.find_company_row(ws, cols["name"], args.company)
    if not matches:
        sys.exit(f'No company matching "{args.company}". Use --list.')
    if len(matches) > 1:
        sys.exit(f'Ambiguous: {", ".join(n for _, n in matches)}.')
    row, company = matches[0]

    nopat0 = m.num(ws, row, cols["nopat"])
    roiic0 = m.num(ws, row, cols["roiic"])
    rr0 = m.num(ws, row, cols["rr"])
    mktcap = m.num(ws, row, cols["mktcap"])
    netdebt = m.num(ws, row, cols["netdebt"]) or 0.0
    if None in (nopat0, roiic0, rr0) or mktcap in (None, 0):
        sys.exit(f"Missing inputs for {company}: NOPAT={nopat0}, ROIIC={roiic0}, "
                 f"RR={rr0}, MktCap={mktcap}")

    g_term, n1 = args.gterm, args.n1
    rd_note = ""
    if args.re is not None:
        country = ws.cell(row=row, column=cols["country"]).value if cols["country"] else None
        cbase, ccy = m.currency_base(country, m.parse_kv_rates(args.country_base))
        gross = m.num(ws, row, cols["gross"]); tax = m.num(ws, row, cols["tax"])
        if tax is None or tax >= 1:
            tax = 0.25
        rd, rating, cov = m.synthetic_rd(nopat0 / (1 - tax), gross, mktcap, cbase)
        r = m.firm_wacc(args.re, rd, mktcap, netdebt)
        cv = ">99" if cov == float("inf") else f"{cov:.1f}x"
        rd_note = f"WACC (re {m.pct(args.re)}, {ccy} {m.pct(cbase)}, cov {cv} {rating} Rd {m.pct(rd)})"
    else:
        r = args.r
    if g_term >= r:
        sys.exit(f"g_term ({g_term}) must be < r ({r:.4f}).")

    moat = m.num(ws, row, cols["moat"])
    mp = m.moat_to_cap_persistence(moat)
    phi = args.persistence if args.persistence is not None else (mp[1] if mp else 0.75)
    tier = mp[2] if mp else "n/a"
    score_life = m.moat_to_life(moat)            # moat-score-implied length, for comparison

    industry = ws.cell(row=row, column=cols["industry"]).value if cols["industry"] else None
    base, base_src = m.base_rate_for(industry, args.base_rate)
    if args.base_rate is None and args.base_inflation:
        base += args.base_inflation
        base_src += f" +{m.pct(args.base_inflation)} infl"

    # Target: the operating (enterprise) value implied by today's market cap.
    target_ev = mktcap + netdebt

    def total_for(n2):
        return m.value_company(nopat0, roiic0, rr0, r, g_term, n1, n2, phi, base)["total"]

    # value_company total is monotonic increasing in n2. Scan integer n2, then
    # linearly interpolate the crossing to a fractional year.
    t0 = total_for(0)
    implied_n2 = None
    note = ""
    if t0 >= target_ev:
        note = (f"hold-only value ({m.money(t0)}) already >= market EV; the price "
                f"implies a moat at/below the {n1}y hold.")
        implied_total = float(n1)               # <= n1 years
        capped = "<="
    else:
        prev_tot = t0
        capped = None
        for n2 in range(1, args.max_n2 + 1):
            tot = total_for(n2)
            if tot >= target_ev:
                frac = (target_ev - prev_tot) / (tot - prev_tot)
                implied_n2 = (n2 - 1) + frac
                break
            prev_tot = tot
        if implied_n2 is None:
            note = (f"even n2={args.max_n2}y (total {m.money(prev_tot)}) is below "
                    f"market EV; price not justified by the model at r={m.pct(r)}.")
            implied_total = float(n1 + args.max_n2)
            capped = ">="
        else:
            implied_total = n1 + implied_n2
            capped = "="

    ticker = ws.cell(row=row, column=cols["ticker"]).value if cols["ticker"] else ""
    print(f"\nAIP IMPLIED MOAT LENGTH — {company} {f'({ticker})' if ticker else ''}")
    print(f"discount r = {m.pct(r)}   phi = {phi:.2f} [{tier}]   "
          f"g_term = {m.pct(g_term)}   n1(hold) = {n1}y")
    if rd_note:
        print(f"  {rd_note}")
    print("=" * 66)
    print(f"  NOPAT_0 ........................ {m.money(nopat0)}")
    print(f"  ROIIC_0 (ROICm 7) .............. {m.pct(roiic0)}")
    print(f"  RR_0 (RR 7) ................... {m.pct(rr0)}")
    print(f"  ROIIC base rate ............... {m.pct(base)}   [{base_src}]")
    print(f"  Market cap .................... {m.money(mktcap)}")
    print(f"  + Net debt .................... {m.money(netdebt)}")
    print(f"  = Target EV (to solve to) ..... {m.money(target_ev)}")
    print("  " + "-" * 50)
    print(f"  >> IMPLIED MOAT LENGTH (n1+n2) {capped} {implied_total:.1f} years")
    if implied_n2 is not None and capped == "=":
        print(f"     (n1 {n1}y hold + n2 {implied_n2:.1f}y fade)")
    if note:
        print(f"     note: {note}")

    print("\n  COMPARISON TO MOAT SCORE")
    if score_life is not None:
        print(f"    Moat Score {moat:.2f} [{tier}] -> framework life {score_life}y")
        diff = implied_total - score_life
        if capped == "=":
            verdict = ("market prices in a LONGER moat than the score supports "
                       "(rich)" if diff > 1 else
                       "market prices in a SHORTER moat than the score supports "
                       "(cheap)" if diff < -1 else
                       "market-implied length ~ the score-implied length (fair)")
            print(f"    Market-implied {implied_total:.1f}y vs score {score_life}y "
                  f"({'+' if diff>=0 else ''}{diff:.1f}y) -> {verdict}")
    else:
        print("    No Moat Score in the sheet for a comparison.")

    print("\n  Analytical framework output, not investment advice.")


if __name__ == "__main__":
    main()
