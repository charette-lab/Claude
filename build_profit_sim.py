"""Profit simulation by AUM level + ownership distribution, added to the framework workbook.

Firm P&L (mirrors Compensation.xlsx Static EBIT):
  Revenue   = Management fee (1.5%xAUM) + Equity take (40% of performance fee)
  Opex      = Loaded salaries (team) + Other costs
  EBIT      = Revenue - Opex
Owners: CIO 63.6%, PM1 10%, PM2 10%, Other 16.4%.
(The 60% team take of the performance fee is paid to staff and is NOT firm revenue.)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
import matplotlib; matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np

WB="/home/user/Claude/Performance_Fee_Framework.xlsx"
PNG="/home/user/Claude/Profit_simulation.png"

RET=0.12; FEE=0.20; TAKE=0.60; MGMT=0.015; THR=300; LOAD=1.62
scn=[20,300,500,1000,2000,5000]
roles=[  # (name, sal<300, sal>300, weight, headcount)
 ("CIO",0.20,0.30,16,[1,1,1,1,1,1]),("PM",0.13,0.20,8,[2,2,3,3,4,4]),
 ("Junior PM",0.10,0.20,6,[0,0,1,1,2,2]),("Senior Analyst",0.05,0.20,4,[0,1,1,2,3,3]),
 ("Analyst",0.00,0.20,2.5,[0,2,3,4,8,8]),("CEO",0.15,0.30,8,[1,1,1,1,1,1]),
 ("CFO",0.10,0.15,3,[0.2,0.2,0.2,1,1,1]),("COO",0.12,0.20,3,[0.5,0.5,0.5,1,1,1]),
 ("IR",0.12,0.12,1,[1,2,2,2,2,2]),("Support",0.12,0.12,1,[0.5,0.5,0.5,1,1,1]),
 ("Execution",0.00,0.20,2,[0,0,0,1,1,1]),("Risk",0.00,0.20,2,[0,0,0,1,1,1]),
 ("Compliance",0.00,0.20,1,[0,0,0,1,1,1])]
OWN=[("CIO",0.636),("Portfolio Manager 1",0.10),("Portfolio Manager 2",0.10),("Other owners",0.164)]

def salary(j):
    use_lo = scn[j] < THR
    return sum((lo if use_lo else hi)*hc[j] for (_,lo,hi,_,hc) in roles)*LOAD
sal=[salary(j) for j in range(6)]
other=[0.5,1.0,1.5,3.0,sal[4]/2,sal[5]/2]
mgmt=[MGMT*a for a in scn]
perf=[a*RET*FEE for a in scn]
team=[p*TAKE for p in perf]
equity=[p*(1-TAKE) for p in perf]
rev=[mgmt[j]+equity[j] for j in range(6)]
opex=[sal[j]+other[j] for j in range(6)]
ebit=[rev[j]-opex[j] for j in range(6)]

print("AUM        ",["%8d"%a for a in scn])
print("Mgmt fee   ",["%8.3f"%x for x in mgmt])
print("Perf fee   ",["%8.3f"%x for x in perf])
print(" team(60%) ",["%8.3f"%x for x in team])
print(" equity(40)",["%8.3f"%x for x in equity])
print("Revenue    ",["%8.3f"%x for x in rev])
print("Salaries   ",["%8.3f"%x for x in sal])
print("Other      ",["%8.3f"%x for x in other])
print("Opex       ",["%8.3f"%x for x in opex])
print("EBIT       ",["%8.3f"%x for x in ebit])
print("(source EBIT @20 should be -1.4174):", round(ebit[0],4))
for nm,sh in OWN:
    print(f"  {nm:20} {sh:5.1%}",["%8.3f"%(e*sh) for e in ebit])

# ---------------- workbook sheet ----------------
wb=openpyxl.load_workbook(WB)
if "Profit Simulation" in wb.sheetnames: del wb["Profit Simulation"]
ws=wb.create_sheet("Profit Simulation")
NAVY="1F3864";BLUE="2E5496";LIGHT="D6E0F0";LIGHTER="EAF0FA";GREY="F2F2F2";GOLD="FFF2CC";GREEN="E2EFDA";RED="F8CBAD"
fnavy=PatternFill("solid",fgColor=NAVY);fhdr=PatternFill("solid",fgColor=BLUE);flight=PatternFill("solid",fgColor=LIGHT)
flighter=PatternFill("solid",fgColor=LIGHTER);fgrey=PatternFill("solid",fgColor=GREY);fgold=PatternFill("solid",fgColor=GOLD)
fgreen=PatternFill("solid",fgColor=GREEN);fred=PatternFill("solid",fgColor=RED)
title_f=Font(size=14,bold=True,color="FFFFFF");sub_f=Font(size=10,italic=True,color="FFFFFF")
hdr_f=Font(size=10,bold=True,color="FFFFFF");bold=Font(size=10,bold=True);norm=Font(size=10)
ital=Font(size=9,italic=True,color="595959");inp_f=Font(size=10,bold=True,color="C00000");wfw=Font(size=10,bold=True,color="FFFFFF")
ctr=Alignment("center","center");lft=Alignment("left","center",wrap_text=True);rgt=Alignment("right","center")
thin=Side("thin",color="BFBFBF");box=Border(thin,thin,thin,thin)
SC=["H","I","J","K","L","M"];MM='#,##0.000';PCT='0.0%'
def s(coord,v,f=norm,fill=None,al=None,fmt=None,bd=True):
    c=ws[coord];c.value=v;c.font=f
    if fill:c.fill=fill
    if al:c.alignment=al
    if fmt:c.number_format=fmt
    if bd:c.border=box
    return c
ws.merge_cells("A1:M1");s("A1","Profit simulation by AUM level, with ownership distribution",title_f,fnavy,ctr,bd=False)
ws.merge_cells("A2:M2");s("A2","Firm EBIT = management fee + equity take (40% of perf fee) − operating costs. Distributed to owners. Millions; links to Static.",sub_f,fnavy,ctr,bd=False)
ws.row_dimensions[1].height=22
s("A4","P&L line  (millions)",hdr_f,fhdr,lft)
for i,col in enumerate(SC):s(f"{col}4",f"AUM ${scn[i]:,}m",hdr_f,fhdr,ctr)
def row(r,label,fn,fill=None,f=norm,fmt=MM,al=lft):
    s(f"A{r}",label,f,fill,al)
    for col in SC:s(f"{col}{r}",fn(col),f,fill,ctr,fmt)
row(5,"AUM",lambda c:f"=Static!{c}15",flight,bold,'#,##0')
row(6,"Management fee  (1.5% × AUM)",lambda c:f"=Static!{c}15*Static!$F$8",flighter)
row(7,"Performance fee  (20% × profit)",lambda c:f"=Static!{c}15*Static!$F$7*Static!$F$5",flighter)
row(8,"  of which team take (60%) → staff",lambda c:f"={c}7*Static!$F$6",None,ital)
row(9,"  of which equity take (40%) → firm",lambda c:f"={c}7*(1-Static!$F$6)",flighter)
row(10,"Firm revenue  (mgmt fee + equity take)",lambda c:f"={c}6+{c}9",fgreen,bold)
row(11,"Operating cost: loaded salaries",lambda c:f"=Static!{c}36",flighter)
s("A12","Operating cost: other costs",norm,fgold,lft)
for col,o in zip(SC,other):s(f"{col}12",round(o,3),inp_f,fgold,ctr,MM)
row(13,"Total operating costs",lambda c:f"={c}11+{c}12",fred,bold)
row(14,"EBIT  (profit to owners)",lambda c:f"={c}10-{c}13",fhdr,wfw)
ws["A14"].font=wfw
row(15,"EBIT margin (on firm revenue)",lambda c:f"=IF({c}10=0,0,{c}14/{c}10)",None,ital,PCT)
s("A17","Ownership distribution of EBIT",hdr_f,fhdr,lft)
for col in SC:s(f"{col}17","",hdr_f,fhdr,ctr)
own_first=18
for i,(nm,sh) in enumerate(OWN):
    r=own_first+i;fr=flighter if i%2 else None
    s(f"A{r}",f"{nm}",norm,fr,lft)
    s(f"G{r}",sh,inp_f,fgold,ctr,PCT)
    for col in SC:s(f"{col}{r}",f"={col}14*$G{r}",norm,fr,ctr,MM)
chk=own_first+len(OWN)
s(f"A{chk}","Total distributed",bold,fgreen,lft)
s(f"G{chk}",f"=SUM(G{own_first}:G{chk-1})",bold,fgreen,ctr,PCT)
for col in SC:s(f"{col}{chk}",f"=SUM({col}{own_first}:{col}{chk-1})",bold,fgreen,ctr,MM)
bc=BarChart();bc.type="col";bc.title="Firm EBIT by AUM level (millions)";bc.height=8;bc.width=18
dd=Reference(ws,min_col=1,min_row=14,max_col=7,max_row=14)
cats=Reference(ws,min_col=8,min_row=4,max_col=13,max_row=4)
bc.add_data(dd,titles_from_data=True,from_rows=True);bc.set_categories(cats);ws.add_chart(bc,f"A{chk+2}")
lc=LineChart();lc.title="Profit to each owner (millions)";lc.height=8;lc.width=18;lc.y_axis.title="millions"
dd2=Reference(ws,min_col=1,min_row=own_first,max_col=7,max_row=chk-1)
lc.add_data(dd2,titles_from_data=True,from_rows=True);lc.set_categories(cats);ws.add_chart(lc,f"J{chk+2}")
ws.column_dimensions["A"].width=36
for c in "BCDEF":ws.column_dimensions[c].width=6
ws.column_dimensions["G"].width=9
for col in SC:ws.column_dimensions[col].width=12
ws.sheet_view.showGridLines=False;ws.freeze_panes="B5"
wb.save(WB)
print("\nSheets:",wb.sheetnames)

# ---------------- PNG ----------------
x=np.arange(6);labels=[f"${a:,}m" for a in scn]
fig,(ax1,ax2)=plt.subplots(1,2,figsize=(15,6.2))
colors=['#C00000' if e<0 else '#2E7D32' for e in ebit]
ax1.bar(x,ebit,color=colors,edgecolor="black",linewidth=0.5)
ax1.axhline(0,color="black",lw=0.8);ax1.set_xticks(x);ax1.set_xticklabels(labels)
ax1.set_title("Firm EBIT (profit) by AUM level",fontweight="bold");ax1.set_ylabel("Millions")
for xi,e in zip(x,ebit):ax1.text(xi,e+(0.8 if e>=0 else -1.4),f"{e:.1f}",ha="center",fontsize=9,fontweight="bold")
ax1.grid(True,axis="y",alpha=0.3)
w=0.2;oc=['#1F3864','#2E75B6','#9DC3E6','#BFBFBF']
for i,(nm,sh) in enumerate(OWN):
    ax2.bar(x+(i-1.5)*w,[e*sh for e in ebit],w,label=f"{nm} ({sh:.1%})",color=oc[i],edgecolor="black",linewidth=0.3)
ax2.axhline(0,color="black",lw=0.8);ax2.set_xticks(x);ax2.set_xticklabels(labels)
ax2.set_title("Profit distributed to owners",fontweight="bold");ax2.set_ylabel("Millions")
ax2.legend(fontsize=9,frameon=False);ax2.grid(True,axis="y",alpha=0.3)
fig.suptitle("Profit simulation: management fee + equity take − operating costs",fontsize=14,fontweight="bold")
plt.tight_layout();plt.savefig(PNG,dpi=130,bbox_inches="tight");print("PNG done")
