#!/usr/bin/env python3
"""decompose.py — split each company's expected return into CARRY vs RE-RATE.

Where does a name's modelled return come from — the internal compounding of the
business, or the price closing to intrinsic value?

  CARRY   = the IRR if you EXIT AT THE PRICE YOU PAID (entry enterprise value held
            constant): interim cash distributions + ROIIC-funded growth. The
            internal return — what you earn even if the multiple never re-rates.
  RE-RATE = ER - carry = the bonus from the price converging to intrinsic value
            over the horizon. The revaluation / cheapness component.

  ER = CARRY + RE-RATE.

The split comes straight from the valuation engine: `aip.value_and_return`
returns `er1_carry` / `er1_rerate` (and the 12%-discount `er2_*` pair), computed
by re-solving the equity IRR with the terminal exit swapped from intrinsic value
to the entry EV (see aip.py). This module applies it across a scored book and adds:

  * a per-name decomposition table (carry / re-rate / carry-share / driver tag);
  * a carry-ranked alternative Core index — the "compounder" book — built under
    the SAME 6/30 segmented-tag rule the pipeline uses (frameworks.fill_under_slot_cap),
    with a selectable eligibility gate.

Financials are read from the LTM screener (the same columns aip uses); the engine
is the proven roiic_dcf.py via aip.py (set AIP_VALUE_ENGINE if it isn't found).

Usage
-----
  # Decompose the Core Index of a scored book:
  AIP_VALUE_ENGINE=/path/to/roiic_dcf.py python3 decompose.py \
      --book Universe_final.xlsx --ltm LTM_current.xlsx --out decomposition.xlsx

  # Decompose every gauntlet-clearer (--scope clears) or every row (--scope all):
  ... --scope clears

  # Build the carry-ranked compounder Core index instead:
  ... --mode compounder --gate value   --out compounder.xlsx
  ... --mode compounder --gate loose --carry-min 0.12 --out compounder_loose.xlsx

Programmatic
------------
  from decompose import load_financials, load_book, decompose
  fins = load_financials("LTM_current.xlsx")
  recs, _ = load_book("Universe_final.xlsx")
  d = decompose(fins[ticker])        # -> {"er","carry","rerate","carry_share","driver",...}
"""
from __future__ import annotations
import argparse
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import openpyxl
import aip
import frameworks as F

# ----------------------------------------------------------------------------- I/O

def _num(v):
    return v if isinstance(v, (int, float)) else None


def load_financials(ltm_path, sheet=None):
    """{ticker: fin-dict} keyed by the field names aip.value_and_return expects."""
    wb = openpyxl.load_workbook(ltm_path, data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb.worksheets[0]
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    li = {h: i for i, h in enumerate(hdr)}
    text_fields = ("Company Name", "Instrument", "GICS Industry Group Name",
                   "Country of Headquarters")
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        t = row[li["Instrument"]] if "Instrument" in li else None
        if not t:
            continue
        out[t] = {fld: (row[li[fld]] if fld in text_fields else _num(row[li[fld]]))
                  for fld in aip.FIELDS.values() if fld in li}
    return out


def load_book(book_path):
    """Return (records, value_core_tickers) from a scored book.

    records: list of dicts with the qualitative/eligibility fields the decomposition
    and the carry-ranked selection need. value_core: the tickers in the book's own
    'Core Index' sheet (for the 'also in the value book?' flag)."""
    wb = openpyxl.load_workbook(book_path, data_only=True, read_only=True)
    ws = wb["Scored"]
    rows = list(ws.iter_rows(values_only=True))
    ci = {h: i for i, h in enumerate(rows[0])}

    def g(r, c):
        return r[ci[c]] if c in ci else None

    recs = []
    for r in rows[1:]:
        t = g(r, "Ticker")
        if not t:
            continue
        tags = [x.strip() for x in (g(r, "RiskTags") or "").split(",") if x.strip()]
        recs.append({
            "ticker": t, "company": g(r, "Company"), "industry": g(r, "Industry"),
            "core_moat": g(r, "CoreMoat(v3.2)"), "company_moat": g(r, "CompanyMoat(v3.2)"),
            "hist": g(r, "MoatVsHistory"), "artifact": g(r, "ER_Artifact") == "ARTIFACT",
            "clears": g(r, "ClearsGauntlet") == "YES", "owner": g(r, "OwnerVerdict"),
            "risk_tag_names": tags,
        })

    value_core = set()
    if "Core Index" in wb.sheetnames:
        for r in wb["Core Index"].iter_rows(min_row=2, values_only=True):
            if r and r[2] and r[0] is not None:
                value_core.add(r[2])
    return recs, value_core


# ------------------------------------------------------------------- decomposition

def decompose(fin, re=0.07, re2=0.12, country_base=None, moat_override=None):
    """Carry / re-rate split for one company.

    Returns a dict with the realistic read (er1, at `re`) and the conservative
    read (er2, at `re2`): er, carry, rerate, carry_share (carry/er), driver.
    `moat_override` values the name on a chosen moat (e.g. the researched CORE
    moat for a freed-core thesis) instead of the screener Moat Score.
    None is returned only if the engine could not value the name."""
    v = aip.value_and_return(fin, re=re, re2=re2, country_base=country_base,
                             moat_override=moat_override)
    if not v or v.get("er1") is None or v.get("er1_carry") is None:
        return None
    er, carry, rerate = v["er1"], v["er1_carry"], v.get("er1_rerate")
    share = (carry / er) if er else None
    out = {"er": er, "carry": carry, "rerate": rerate, "carry_share": share,
           "driver": driver_label(share), "wacc": v.get("wacc"), "rating": v.get("rating"),
           "er2": v.get("er2"), "carry2": v.get("er2_carry"), "rerate2": v.get("er2_rerate")}
    return out


def driver_label(carry_share, hi=0.60, lo=0.40):
    """Tag the return's source from the carry share of total ER."""
    if carry_share is None:
        return "n/a"
    if carry_share >= hi:
        return "ROIIC-carry"
    if carry_share <= lo:
        return "RE-RATE"
    return "mixed"


# ----------------------------------------------------------- carry-ranked selection

def _eligible(rec, gate, core_moat_min=6.5):
    """Eligibility for the carry-ranked book.

    gate='value' : same pool as the pipeline's Core Index (clears the gauntlet,
                   not history-CONTRADICTED, tagged) — isolates the effect of the
                   ranking key (carry vs total ER).
    gate='loose' : quality bars only (core moat >= min, not CONTRADICTED, not an
                   ER-artifact, tagged) but NO total-ER gauntlet — admits fairly-
                   and richly-priced compounders the value gate would reject."""
    if rec["hist"] == "CONTRADICTED" or not rec["risk_tag_names"]:
        return False
    if gate == "value":
        return bool(rec["clears"])
    if gate == "loose":
        cm = rec["core_moat"]
        return (isinstance(cm, (int, float)) and cm >= core_moat_min
                and not rec["artifact"])
    raise ValueError(f"unknown gate {gate!r}")


def build_carry_book(recs, fins, *, re=0.07, re2=0.12, gate="value",
                     carry_min=None, n=F.CORE_N, slot_cap=F.CORE_SLOT_CAP,
                     core_moat_min=6.5):
    """Carry-ranked Core index under the 6/30 segmented-tag rule.

    Decomposes every eligible name, optionally drops names below `carry_min`,
    ranks by carry (capped at MAX_PLAUSIBLE_IRR like the pipeline), and fills under
    the slot cap. Returns (book, all_scored) where each entry carries its
    decomposition fields merged in."""
    scored = []
    for rec in recs:
        if not _eligible(rec, gate, core_moat_min):
            continue
        fin = fins.get(rec["ticker"])
        if not fin:
            continue
        d = decompose(fin, re=re, re2=re2)
        if not d:
            continue
        scored.append({**rec, **d})
    cand = [r for r in scored if carry_min is None or (r["carry"] is not None and r["carry"] >= carry_min)]
    cand.sort(key=lambda r: -min(r["carry"], F.MAX_PLAUSIBLE_IRR))

    def tagset(r):
        ind = " ".join(str(r.get("industry") or "?").split())
        return {f"{t}:{ind}" for t in r["risk_tag_names"]}

    book, _skipped = F.fill_under_slot_cap(cand, tagset, n, slot_cap)
    return book, scored


# ----------------------------------------------------------------------- reporting

def _avg(xs):
    xs = [x for x in xs if isinstance(x, (int, float))]
    return sum(xs) / len(xs) if xs else None


def write_decomposition(path, rows, value_core=frozenset()):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Return decomposition"
    ws.append(["Ticker", "Company", "Industry", "ER", "Carry (ROIIC)", "Re-rate",
               "Carry % of ER", "Driver", "CoreMoat", "History", "In Core Index?"])
    for r in rows:
        ws.append([r["ticker"], r["company"], r["industry"], r["er"], r["carry"],
                   r["rerate"], r["carry_share"], r["driver"], r["core_moat"],
                   r["hist"], "yes" if r["ticker"] in value_core else ""])
    ws.append([])
    ws.append(["EQUAL-WT", f"({len(rows)} names)", "", _avg([r["er"] for r in rows]),
               _avg([r["carry"] for r in rows]), _avg([r["rerate"] for r in rows]),
               (_avg([r["carry"] for r in rows]) / _avg([r["er"] for r in rows]))
               if rows and _avg([r["er"] for r in rows]) else None, "", "", ""])
    for row in ws.iter_rows(min_row=2):
        for j in (3, 4, 5, 6):
            if isinstance(row[j].value, float):
                row[j].number_format = "0.0%"
    for col, w in zip("ABCDEFGHIJK", [11, 26, 26, 8, 14, 9, 13, 12, 9, 11, 14]):
        ws.column_dimensions[col].width = w
    wb.save(path)


def write_compounder(path, book, scored, value_core, *, gate, carry_min):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Compounder Core (by carry)"
    ws.append(["Rank", "Ticker", "Company", "Industry", "Carry", "Total ER", "Re-rate",
               "Carry % of ER", "CoreMoat", "History", "Ownership", "Clears value gate?",
               "In Core Index?"])
    for i, r in enumerate(book, 1):
        ws.append([i, r["ticker"], r["company"], r["industry"], r["carry"], r["er"],
                   r["rerate"], r["carry_share"], r["core_moat"], r["hist"], r["owner"],
                   "yes" if r["clears"] else "NO", "yes" if r["ticker"] in value_core else ""])
    ws.append([])
    ws.append(["EQ-WT", "", f"({len(book)})", "", _avg([r["carry"] for r in book]),
               _avg([r["er"] for r in book]), _avg([r["rerate"] for r in book])])
    for row in ws.iter_rows(min_row=2):
        for j in (4, 5, 6, 7):
            if isinstance(row[j].value, float):
                row[j].number_format = "0.0%"
    for col, w in zip("ABCDEFGHIJKLM", [5, 11, 26, 28, 10, 10, 9, 13, 9, 11, 13, 16, 14]):
        ws.column_dimensions[col].width = w

    info = wb.create_sheet("Finding")
    qualifiers = [r for r in scored if carry_min is None or
                  (r["carry"] is not None and r["carry"] >= (carry_min or 0))]
    fail_gate = sum(1 for r in qualifiers if not r["clears"])
    for line in [
        ["Carry-ranked compounder book"],
        [""],
        ["gate", gate], ["carry_min", carry_min],
        ["eligible pool valued", len(scored)],
        [f"...with carry >= {carry_min}", len(qualifiers)],
        ["...of those, FAIL the value gate (total ER < hurdle)", fail_gate],
        [""],
        ["book carry (eq-wt)", round(_avg([r["carry"] for r in book]) or 0, 4)],
        ["book total ER (eq-wt)", round(_avg([r["er"] for r in book]) or 0, 4)],
    ]:
        info.append(line)
    info.column_dimensions["A"].width = 48
    info.column_dimensions["B"].width = 16
    wb.save(path)


# ---------------------------------------------------------------------------- main

def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--book", required=True, help="scored book xlsx (Scored sheet)")
    ap.add_argument("--ltm", required=True, help="LTM screener xlsx (financials/prices)")
    ap.add_argument("--out", required=True, help="output xlsx")
    ap.add_argument("--mode", choices=["decompose", "compounder"], default="decompose")
    ap.add_argument("--scope", choices=["core", "clears", "all"], default="core",
                    help="decompose mode: which names (default: the book's Core Index)")
    ap.add_argument("--gate", choices=["value", "loose"], default="value",
                    help="compounder mode: eligibility gate")
    ap.add_argument("--carry-min", type=float, default=None,
                    help="compounder mode: drop names with carry below this (e.g. 0.12)")
    ap.add_argument("--re", type=float, default=0.07)
    ap.add_argument("--re2", type=float, default=0.12)
    ap.add_argument("--core-n", type=int, default=F.CORE_N)
    args = ap.parse_args()

    fins = load_financials(args.ltm)
    recs, value_core = load_book(args.book)
    by_ticker = {r["ticker"]: r for r in recs}

    if args.mode == "decompose":
        if args.scope == "core":
            names = [by_ticker[t] for t in value_core if t in by_ticker]
        elif args.scope == "clears":
            names = [r for r in recs if r["clears"]]
        else:
            names = recs
        out_rows = []
        for rec in names:
            fin = fins.get(rec["ticker"])
            if not fin:
                continue
            d = decompose(fin, re=args.re, re2=args.re2)
            if d:
                out_rows.append({**rec, **d})
        out_rows.sort(key=lambda r: -(r["er"] or -9))
        write_decomposition(args.out, out_rows, value_core)
        eq_er = _avg([r["er"] for r in out_rows]); eq_c = _avg([r["carry"] for r in out_rows])
        print(f"decomposed {len(out_rows)} names ({args.scope}) -> {args.out}")
        if eq_er:
            print(f"  equal-wt ER={eq_er*100:.1f}%  carry={eq_c*100:.1f}%  "
                  f"re-rate={(eq_er-eq_c)*100:.1f}%  carry-share={eq_c/eq_er*100:.0f}%")
    else:
        book, scored = build_carry_book(recs, fins, re=args.re, re2=args.re2,
                                        gate=args.gate, carry_min=args.carry_min,
                                        n=args.core_n)
        write_compounder(args.out, book, scored, value_core,
                         gate=args.gate, carry_min=args.carry_min)
        eq_c = _avg([r["carry"] for r in book]); eq_er = _avg([r["er"] for r in book])
        print(f"compounder book: {len(book)} names (gate={args.gate}, "
              f"carry_min={args.carry_min}) -> {args.out}")
        if eq_er:
            print(f"  equal-wt carry={eq_c*100:.1f}%  ER={eq_er*100:.1f}%  "
                  f"carry-share={eq_c/eq_er*100:.0f}%")


if __name__ == "__main__":
    main()
