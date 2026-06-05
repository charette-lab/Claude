#!/usr/bin/env python3
"""AIP Value — batch ranker. Runs the whole sheet in one command: per-company
WACC (from an equity hurdle), expected return at one or two hurdles, and the
implied moat length, ranked. Reuses the aip-value engine (roiic_dcf.py).

Usage:
  python3 rank.py <file.xlsx> --re 0.07 [--re2 0.12] [--country-base "EUR=0.0303,..."]
  python3 rank.py <file.xlsx> --r 0.12            # flat firm discount rate (no WACC)

Refresh --country-base with live 10y govt yields (USD = UST - ~40bp).
"""
import argparse
import importlib.util
import os
import sys


def load_engine():
    here = os.path.dirname(os.path.abspath(__file__))
    for path in (os.path.join(here, "roiic_dcf.py"),
                 os.path.expanduser("~/.claude/skills/aip-value/roiic_dcf.py")):
        if os.path.exists(path):
            spec = importlib.util.spec_from_file_location("aipval", path)
            mod = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(mod)
            return mod
    sys.exit("Could not find roiic_dcf.py (the aip-value engine).")


def main():
    m = load_engine()
    import openpyxl
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("file")
    ap.add_argument("--re", type=float, default=None, help="required equity return (WACC mode)")
    ap.add_argument("--re2", type=float, default=None, help="optional 2nd equity hurdle column")
    ap.add_argument("--r", type=float, default=0.12, help="flat firm rate if --re not set")
    ap.add_argument("--country-base", default=None)
    ap.add_argument("--country-crp", default=None, help="override country risk premiums")
    ap.add_argument("--gterm", type=float, default=0.025)
    ap.add_argument("--horizon", type=int, default=5)
    ap.add_argument("--max-n2", type=int, default=150)
    ap.add_argument("--sheet", default=None)
    args = ap.parse_args()

    overrides = m.parse_kv_rates(args.country_base)
    ws = (openpyxl.load_workbook(args.file, data_only=True))[args.sheet] if args.sheet \
        else openpyxl.load_workbook(args.file, data_only=True).worksheets[0]
    C = m.find_columns(ws, {
        "name": "Company Name", "nopat": "New Operating Income", "roiic": "ROICm 7",
        "rr": "RR 7", "moat": "Moat Score", "industry": "GICS Industry Group Name",
        "country": "Country of Headquarters", "gross": "Gross debt",
        "tax": "Income Tax Rate - Instrument", "mktcap": "Market Cap",
        "netdebt": "Net debt", "ticker": "Instrument"})

    def discount_rate(re, nopat0, mktcap, netdebt, gross, tax, country):
        if re is None:
            return args.r, None, None
        cbase, ccy = m.currency_base(country, overrides)
        t = tax if (tax is not None and tax < 1) else 0.25
        rd, rating, cov = m.synthetic_rd(nopat0 / (1 - t), gross, mktcap, cbase)
        crp = m.country_risk_premium(country, m.parse_kv_rates(args.country_crp))
        return min(max(m.firm_wacc(re, rd, mktcap, netdebt) + crp, 0.04), 0.25), rd, rating

    def exp_ret(nopat0, roiic0, rr0, r, phi, base, life, mktcap, netdebt):
        n1 = min(3, max(1, life - 1)); n2 = max(1, life - n1)
        res = m.value_company(nopat0, roiic0, rr0, r, args.gterm, n1, n2, phi, base)
        cf = sum(res["cf_for_year"](t) for t in range(1, args.horizon + 1))
        eqv = res["total"] - (netdebt - cf)
        return ((eqv / mktcap) ** (1 / args.horizon) - 1) if (eqv > 0 and mktcap > 0) else None

    def imp_moat(nopat0, roiic0, rr0, r, phi, base, mktcap, netdebt):
        target = mktcap + netdebt
        tot = lambda n2: m.value_company(nopat0, roiic0, rr0, r, args.gterm, 3, n2, phi, base)["total"]
        t0 = tot(0)
        if t0 >= target:
            return "<=3y"
        prev = t0
        for n2 in range(1, args.max_n2 + 1):
            tt = tot(n2)
            if tt >= target:
                return f"{((n2 - 1) + (target - prev) / (tt - prev) + 3):.0f}y"
            prev = tt
        return ">150y"

    rows = []
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=C["name"]).value
        if not name:
            continue
        nopat0 = m.num(ws, row, C["nopat"]); roiic0 = m.num(ws, row, C["roiic"])
        rr0 = m.num(ws, row, C["rr"]); mktcap = m.num(ws, row, C["mktcap"])
        netdebt = m.num(ws, row, C["netdebt"]) or 0.0
        if None in (nopat0, roiic0, rr0) or mktcap in (None, 0):
            continue
        gross = m.num(ws, row, C["gross"]); tax = m.num(ws, row, C["tax"])
        moat = m.num(ws, row, C["moat"]); country = ws.cell(row=row, column=C["country"]).value
        ind = " ".join(str(ws.cell(row=row, column=C["industry"]).value or "").split())
        mp = m.moat_to_cap_persistence(moat)
        phi = mp[1] if mp else 0.75
        life = m.moat_to_life(moat) or 15
        base = m.base_rate_for(ind, None)[0]
        r1, rd, rating = discount_rate(args.re, nopat0, mktcap, netdebt, gross, tax, country)
        er1 = exp_ret(nopat0, roiic0, rr0, r1, phi, base, life, mktcap, netdebt)
        im1 = imp_moat(nopat0, roiic0, rr0, r1, phi, base, mktcap, netdebt)
        er2 = None
        if args.re2 is not None:
            r2 = discount_rate(args.re2, nopat0, mktcap, netdebt, gross, tax, country)[0]
            er2 = exp_ret(nopat0, roiic0, rr0, r2, phi, base, life, mktcap, netdebt)
        rows.append({"name": str(name), "moat": moat, "mc": mktcap, "rat": rating or "-",
                     "rd": rd, "wacc": r1, "er1": er1, "er2": er2, "im1": im1,
                     "fin": "FIN*" if ind in m.FINANCIAL_SECTORS else ""})

    rows.sort(key=lambda d: (d["er1"] is not None, d["er1"] if d["er1"] is not None else -9),
              reverse=True)
    fmc = lambda x: f"{x/1e9:,.1f}bn" if x >= 1e9 else f"{x/1e6:,.0f}m"
    fer = lambda x: f"{x*100:5.1f}%" if x is not None else "  n/a "
    h = args.horizon
    re1 = m.pct(args.re) if args.re is not None else f"flat {m.pct(args.r)}"
    hdr = f"{'#':>2}  {'Company':27} {'Moat':>5} {'MCap':>7} {'Rat':>4} {'Rd':>5} {'WACC':>5} {'ER1':>7}"
    if args.re2 is not None:
        hdr += f" {'ER2':>7}"
    hdr += f" {'ImpMoat':>7}  fin"
    print(f"\nAIP batch — ER1=ER@{re1}" + (f", ER2=ER@{m.pct(args.re2)}" if args.re2 else "")
          + f", {h}y horizon" + (f"  | rates: {args.country_base}" if args.country_base else ""))
    print(hdr); print("-" * len(hdr))
    for i, d in enumerate(rows, 1):
        rd = m.pct(d["rd"]) if d["rd"] is not None else "  -  "
        line = (f"{i:>2}  {d['name'][:27]:27} {d['moat'] if d['moat'] is not None else 0:5.2f} "
                f"{fmc(d['mc']):>7} {d['rat']:>4} {rd:>5} {m.pct(d['wacc']):>5} {fer(d['er1']):>7}")
        if args.re2 is not None:
            line += f" {fer(d['er2']):>7}"
        line += f" {d['im1']:>7}  {d['fin']}"
        print(line)
    print(f"\nValued: {len(rows)}   (banks marked FIN* — coverage/rating unreliable)")


if __name__ == "__main__":
    main()
