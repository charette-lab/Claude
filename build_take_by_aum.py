"""Redo: role share of the take as AUM grows smoothly from $100m to $2bn (no historical returns)."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
import matplotlib; matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np

WB="/home/user/Claude/Performance_Fee_Framework.xlsx"
PNG="/home/user/Claude/Take_share_by_AUM.png"

scn_aum=[100,300,500,1000,2000,5000]
roles=[
 ("Managing Partner / CIO",100,[1,1,1,1,1,1]),
 ("Partner / Co-PM",         60,[1,1,2,2,3,4]),
 ("Portfolio Manager",       40,[0,1,1,2,3,5]),
 ("Head of Research",        30,[0,1,1,1,1,1]),
 ("Senior Analyst",          20,[1,1,2,3,4,6]),
 ("Analyst",                 12,[1,2,3,4,6,10]),
 ("Junior Analyst",           6,[0,1,2,3,4,8]),
 ("Head of Trading",         18,[0,0,1,1,1,1]),
 ("Trader / Execution",      10,[0,1,1,2,2,3]),
 ("COO / Head of Operations",15,[1,1,1,1,1,1]),
 ("CFO / Finance",           12,[0,1,1,1,1,1]),
 ("Compliance / Legal",       8,[0,0,1,1,1,2]),
 ("Investor Relations / Ops", 5,[1,1,1,2,3,5]),
]
aum_grid=list(range(100,2001,100))  # $100m -> $2bn, 20 steps

def interp_hc(hc,A):
    A=max(scn_aum[0],min(scn_aum[-1],A))
    for k in range(len(scn_aum)-1):
        if scn_aum[k]<=A<=scn_aum[k+1]:
            t=(A-scn_aum[k])/(scn_aum[k+1]-scn_aum[k])
            return hc[k]+t*(hc[k+1]-hc[k])
    return hc[-1]

share=np.zeros((len(roles),len(aum_grid)))
for j,A in enumerate(aum_grid):
    units=[w*interp_hc(hc,A) for (_,w,hc) in roles]
    tot=sum(units)
    for i,u in enumerate(units): share[i,j]=u/tot

# ---------- workbook sheet ----------
wb=openpyxl.load_workbook(WB)
for s in ("Take by AUM","Take Over Time"):
    if s in wb.sheetnames: del wb[s]
ws=wb.create_sheet("Take by AUM",1)
NAVY="1F3864";BLUE="2E5496";GREY="F2F2F2";LIGHT="EAF0FA"
title_f=Font(size=14,bold=True,color="FFFFFF");sub_f=Font(size=10,italic=True,color="FFFFFF")
hdr_f=Font(size=9,bold=True,color="FFFFFF");bold=Font(size=9,bold=True);norm=Font(size=9)
fnavy=PatternFill("solid",fgColor=NAVY);fhdr=PatternFill("solid",fgColor=BLUE)
flight=PatternFill("solid",fgColor=LIGHT);fgrey=PatternFill("solid",fgColor=GREY)
ctr=Alignment("center","center",wrap_text=True);lft=Alignment("left","center")
thin=Side("thin",color="BFBFBF");box=Border(thin,thin,thin,thin)
L=openpyxl.utils.get_column_letter
n=len(aum_grid)
ws.merge_cells(f"A1:{L(1+n)}1");c=ws["A1"];c.value="Each role's share of the take (comp pool) as AUM grows $100m → $2bn";c.font=title_f;c.fill=fnavy;c.alignment=ctr
ws.merge_cells(f"A2:{L(1+n)}2");c=ws["A2"];c.value="Share = role weight × headcount ÷ total weighted units. Headcount interpolated across the AUM scenarios. Columns sum to 100%.";c.font=sub_f;c.fill=fnavy;c.alignment=ctr
ws.row_dimensions[1].height=22
hr=4
cc=ws.cell(hr,1,"AUM (USD m) →");cc.font=hdr_f;cc.fill=fhdr;cc.alignment=lft;cc.border=box
for j,A in enumerate(aum_grid):
    cc=ws.cell(hr,2+j,A);cc.font=hdr_f;cc.fill=fhdr;cc.alignment=ctr;cc.border=box;cc.number_format='#,##0'
first=hr+1
for i,(name,_,_) in enumerate(roles):
    r=first+i
    cc=ws.cell(r,1,name);cc.font=norm;cc.alignment=lft;cc.border=box
    if i%2: cc.fill=flight
    for j in range(n):
        c2=ws.cell(r,2+j,float(share[i,j]));c2.font=norm;c2.alignment=ctr;c2.number_format='0.0%';c2.border=box
        if i%2: c2.fill=flight
tr=first+len(roles)
ws.cell(tr,1,"Total").font=hdr_f;ws.cell(tr,1).fill=fhdr;ws.cell(tr,1).alignment=lft;ws.cell(tr,1).border=box
for j in range(n):
    col=L(2+j);c2=ws.cell(tr,2+j);c2.value=f"=SUM({col}{first}:{col}{tr-1})";c2.font=hdr_f;c2.fill=fhdr;c2.alignment=ctr;c2.number_format='0%';c2.border=box
chart=LineChart();chart.title="Role share of the take vs AUM";chart.style=2
chart.y_axis.numFmt='0%';chart.y_axis.title="% of take";chart.x_axis.title="AUM (USD m)"
chart.height=11;chart.width=26
data=Reference(ws,min_col=1,min_row=first,max_row=tr-1,max_col=1+n)
cats=Reference(ws,min_col=2,min_row=hr,max_col=1+n,max_row=hr)
chart.add_data(data,titles_from_data=True,from_rows=True);chart.set_categories(cats)
ws.add_chart(chart,f"A{tr+2}")
ws.column_dimensions["A"].width=26
for j in range(n): ws.column_dimensions[L(2+j)].width=7
ws.sheet_view.showGridLines=False;ws.freeze_panes="B5"
wb.save(WB)

# ---------- PNG ----------
fig,ax=plt.subplots(figsize=(13,7))
cmap=plt.get_cmap("tab20");colors=[cmap(i) for i in range(len(roles))]
ax.stackplot(aum_grid,[share[i]*100 for i in range(len(roles))],labels=[r[0] for r in roles],
             colors=colors,alpha=0.9,edgecolor="white",linewidth=0.3)
ax.set_xlim(aum_grid[0],aum_grid[-1]);ax.set_ylim(0,100)
ax.set_xlabel("AUM (USD m)");ax.set_ylabel("Share of the take (%)")
ax.set_title("How the performance-fee take is split by role, as AUM grows $100m → $2bn",fontsize=13,fontweight="bold")
ax.legend(loc="center left",bbox_to_anchor=(1.01,0.5),fontsize=8,frameon=False)
ax.set_xticks(range(100,2001,200))
plt.tight_layout();plt.savefig(PNG,dpi=130,bbox_inches="tight");print("done")

cols=[100,300,500,1000,1500,2000];idx=[aum_grid.index(a) for a in cols]
print("\nRole % of take vs AUM (USD m):")
print("Role".ljust(28)+"".join(f"{a:>8}" for a in cols))
for i,(name,_,_) in enumerate(roles):
    print(name.ljust(28)+"".join(f"{share[i,j]*100:>7.1f}%" for j in idx))
