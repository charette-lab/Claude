#!/usr/bin/env python3
"""aip-value-sectors — industry-aware valuation on the rich moat-7 panel dataset.

Routes every company to the valuation model that matches its value driver, reusing
the aip-value engine (roiic_dcf.py) for the DCF primitives. Builds the NON-financial
models; financials are flagged and skipped (they need a residual-income/ROE model).

  Archetype       GICS groups                          Model
  --------------- ------------------------------------ --------------------------------
  operating       Capital Goods, Software, Pharma,     aip-value ROIIC DCF on spot
                  Semis, Health Care, Consumer, etc.   ROICm7 (intangibles already
                                                       capitalized in New Op Income)
  cyclical        Energy, Materials                    ROIIC DCF on CYCLE-NORMALISED
                                                       inputs (21/14yr ROICm, mid-cycle
                                                       NOPAT) + reproduction-cost floor
  regulated       Utilities, Telecommunications        ROIIC DCF with the excess return
                                                       faded to WACC (value ~ RAB),
                                                       utility leverage, dividend-weighted
  nav             Real Estate Mgmt & Development        Net asset value: book equity +
                                                       (reproduction cost - net PPE)
                                                       revaluation; price->NAV + yield
  financial       Banks, Insurance, Financial Services  SKIPPED (residual-income model)

Data is a 2011-2026 panel; the latest row per Instrument is used. Net debt = EV - Mkt
Cap. There is no Moat Score column (file is pre-filtered moat>=7) so a default tier is
assumed (--moat-assume, default 7.5).

Usage:
  python3 sector_value.py <file.xlsx> [--re 0.09] [--moat-assume 7.5]
          [--country-base "USD=0.0406,..."] [--only ARCHETYPE] [--csv out.csv]
"""
import argparse
import csv
import importlib.util
import os
import sys


def load_engine():
    here = os.path.dirname(os.path.abspath(__file__))
    for path in (os.path.join(here, "..", "aip-value", "roiic_dcf.py"),
                 os.path.expanduser("~/.claude/skills/aip-value/roiic_dcf.py")):
        if os.path.exists(path):
            spec = importlib.util.spec_from_file_location("aipval", path)
            mod = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(mod)
            return mod
    sys.exit("Could not find the aip-value engine (roiic_dcf.py).")


ARCHETYPE_BY_GICS = {
    "Capital Goods": "operating", "Commercial  & Professional Services": "operating",
    "Commercial & Professional Services": "operating", "Transportation": "operating",
    "Software & Services": "operating", "Technology Hardware & Equipment": "operating",
    "Semiconductors & Semiconductor Equipment": "operating",
    "Health Care Equipment & Services": "operating",
    "Pharmaceuticals, Biotechnology & Life Sciences": "operating",
    "Media & Entertainment": "operating", "Consumer Services": "operating",
    "Consumer Discretionary Distribution & Retail": "operating",
    "Consumer Durables & Apparel": "operating", "Food, Beverage & Tobacco": "operating",
    "Automobiles & Components": "operating",
    "Energy": "cyclical", "Materials": "cyclical",
    "Utilities": "regulated", "Telecommunication Services": "regulated",
    "Real Estate Management & Development": "nav",
    "Financial Services": "financial", "Banks": "financial", "Insurance": "financial",
}


def archetype_for(gics):
    return ARCHETYPE_BY_GICS.get(" ".join(str(gics or "").split()), "operating")


def clamp(x, lo, hi):
    return max(lo, min(hi, x))


# --------------------------------------------------------------------------- #
# WACC (shared) — equity hurdle -> de-risking lever-glide WACC (per aip-value) #
# --------------------------------------------------------------------------- #
def build_wacc(m, re, nopat0, gross_debt, mktcap, netdebt, tax, country, ind, n1, n2, overrides):
    cbase, _ = m.currency_base(country, overrides)
    t = tax if (tax is not None and 0 < tax < 1) else 0.25
    rd = m.synthetic_rd(nopat0 / (1 - t), gross_debt, mktcap, cbase)[0]
    crp = m.country_risk_premium(country)
    L = m.target_leverage(ind)
    rd_m = m.mature_cost_of_debt(L, cbase, mktcap)[0]
    w0 = clamp(m.firm_wacc_taxed(re, rd, mktcap, netdebt, t) + crp, 0.04, re + crp)
    wm = clamp(m.mature_wacc(re, rd_m, L, t)[0] + crp, 0.04, 0.25)
    path = m.wacc_glide(w0, wm, n1, n2)
    return w0, wm, path, L, t


def er_from(m, res, mktcap, netdebt, tax, lever_L, horizon=5):
    val, method = m.equity_return(res, mktcap, netdebt, horizon, tax=tax, lever_L=lever_L)[:2]
    return val, method


# --------------------------------------------------------------------------- #
# The four non-financial models                                               #
# --------------------------------------------------------------------------- #
def value_operating(m, d, re, moat, overrides, roiic=None, nopat=None, base=None, horizon=5):
    """Standard aip-value ROIIC DCF. roiic/nopat/base overridable (used by cyclical
    & regulated variants)."""
    n1, n2, phi = _life(m, moat)
    nopat0 = nopat if nopat is not None else d["nopat"]
    roiic0 = roiic if roiic is not None else d["roiic7"]
    rr0 = clamp(d["rr"] if d["rr"] is not None else 0.3, 0.0, 1.0)
    w0, wm, path, L, t = build_wacc(m, re, nopat0, d["gross"], d["mktcap"], d["netdebt"],
                                    d["tax"], d["country"], d["ind"], n1, n2, overrides)
    b = base if base is not None else m.base_rate_for(d["ind"])[0]
    res = m.value_company(nopat0, roiic0, rr0, w0, 0.025, n1, n2, phi, b,
                          sales0=d["sales"], gics_industry=d["ind"],
                          wacc_path=path, wacc_terminal=wm)
    er, method = er_from(m, res, d["mktcap"], d["netdebt"], t, L, horizon)
    return {"total": res["total"], "wacc0": w0, "wacc_m": wm, "er": er, "method": method,
            "roiic": roiic0, "nopat": nopat0, "base": b, "extra": ""}


def value_cyclical(m, d, re, moat, overrides, horizon=5):
    """Cycle-normalised ROIIC DCF + reproduction-cost asset floor (Energy, Materials)."""
    nopat_mid = (d["nopat7sum"] / 7.0) if d["nopat7sum"] else d["nopat"]      # mid-cycle NOPAT
    roiic_norm = d["roiic21"] or d["roiic7"] or m.base_rate_for(d["ind"])[0]   # long-run ROIC
    out = value_operating(m, d, re, moat, overrides, roiic=roiic_norm, nopat=nopat_mid, horizon=horizon)
    # Asset-value (reproduction cost) floor, discounted if the business earns < WACC.
    floor = None
    if d["repro"]:
        floor = d["repro"] * clamp(roiic_norm / out["wacc0"], 0.30, 1.0)
    if floor and floor > out["total"]:
        out["extra"] = f"asset-floored (repro {d['repro']/1e9:.1f}b)"
        out["total"] = floor
        # recompute ER on the floored operating value
        out["er"] = _er_on_total(m, out["total"], d, horizon)
        out["method"] = "ASSET"
    else:
        out["extra"] = f"EPV>asset (repro {d['repro']/1e9:.1f}b)" if d["repro"] else "no repro"
    out["norm"] = f"ROIC21 {roiic_norm*100:.0f}%, NOPAT mid {nopat_mid/1e9:.2f}b"
    return out


def value_regulated(m, d, re, moat, overrides, horizon=5):
    """Excess return faded to the cost of capital -> value ~ regulated asset base
    (Utilities, Telecom). RAB cross-check = Invested Capital."""
    n1, n2, phi = _life(m, moat)
    w0, wm, path, L, t = build_wacc(m, re, d["nopat"], d["gross"], d["mktcap"], d["netdebt"],
                                    d["tax"], d["country"], d["ind"], n1, n2, overrides)
    # Regulated returns sit AT (a small premium over) the cost of capital — cap the
    # noisy marginal ROIC so the value anchors on the asset base, not extrapolation.
    roiic_reg = clamp(d["roiic7"] or wm, 0.0, 1.5 * wm)
    out = value_operating(m, d, re, moat, overrides, roiic=roiic_reg, base=wm, horizon=horizon)
    rab = d["ic"]
    out["extra"] = f"RAB(IC) {rab/1e9:.1f}b, model/RAB {out['total']/rab:.2f}x" if rab else "no RAB"
    return out


def value_nav(m, d, re, overrides, horizon=5):
    """Net asset value for property developers: book equity + reproduction surplus."""
    surplus = 0.0
    if d["repro"] and d["ppe_net"] is not None:
        surplus = max(0.0, d["repro"] - d["ppe_net"])           # hidden asset value vs depreciated book
    nav_equity = (d["equity"] or 0.0) + surplus
    # operating EV-equivalent for the comparison table = NAV equity + net debt
    total = nav_equity + d["netdebt"]
    er = (nav_equity / d["mktcap"]) ** (1.0 / horizon) - 1 if (nav_equity > 0 and d["mktcap"]) else None
    dy = (d["div"] / d["mktcap"]) if (d["div"] and d["mktcap"]) else 0.0
    asset_light = d["repro"] is None or (d["ic"] and d["ppe_net"] is not None and d["ppe_net"] < 0.2 * d["ic"])
    return {"total": total, "nav_equity": nav_equity, "er": (er + dy) if er is not None else None,
            "method": "NAV", "wacc0": None, "wacc_m": None, "roiic": d["roiic7"], "nopat": d["nopat"],
            "base": None, "extra": (f"NAV/px {nav_equity/d['mktcap']:.2f}x, +{dy*100:.1f}% yld"
                                    + ("  [asset-light: likely OPERATING, not NAV]" if asset_light else ""))}


def _er_on_total(m, total, d, horizon):
    nav_eq = total - d["netdebt"]
    return (nav_eq / d["mktcap"]) ** (1.0 / horizon) - 1 if (nav_eq > 0 and d["mktcap"]) else None


def _life(m, moat):
    life = m.moat_to_life(moat) or 15
    n1 = min(3, max(1, life - 1))
    mp = m.moat_to_cap_persistence(moat)
    return n1, max(1, life - n1), (mp[1] if mp else 0.80)


# --------------------------------------------------------------------------- #
# Adapter: read the rich panel, latest row per company, derive drivers        #
# --------------------------------------------------------------------------- #
def latest_rows(ws, H):
    ni, ti, di = H["Company Name"], H["Instrument"], H.get("Date")
    best = {}
    for r in range(2, ws.max_row + 1):
        nm = ws.cell(row=r, column=ni).value
        if not nm:
            continue
        key = ws.cell(row=r, column=ti).value or nm
        dt = str(ws.cell(row=r, column=di).value) if di else str(r)
        if key not in best or dt > best[key][0]:
            best[key] = (dt, r)
    return [r for _, r in best.values()]


def drivers(ws, H, row):
    def n(name):
        c = H.get(name)
        if not c:
            return None
        v = ws.cell(row=row, column=c).value
        try:
            return float(v)
        except (TypeError, ValueError):
            return None
    mktcap = n("Market Capitalization")
    ev = n("EV")
    netdebt = (ev - mktcap) if (ev is not None and mktcap is not None) else None
    gross = sum(x for x in (n("Debt - Long-Term - Total"),
                            n("Short-Term Debt & Current Portion of Long-Term Debt"),
                            n("Capitalized Lease Obligations - Long-Term"),
                            n("Capitalized Leases - Current Portion")) if x) or 0.0
    ppe_gross = n("Property Plant & Equipment - Gross - Total")
    accdep = n("PPE - Accumulated Depreciation & Impairment - Total")
    ppe_net = (ppe_gross - accdep) if (ppe_gross is not None and accdep is not None) else ppe_gross
    return {
        "name": ws.cell(row=row, column=H["Company Name"]).value,
        "tick": ws.cell(row=row, column=H["Instrument"]).value,
        "country": ws.cell(row=row, column=H["Country of Headquarters"]).value,
        "ind": " ".join(str(ws.cell(row=row, column=H["GICS Industry Group Name"]).value or "").split()),
        "mktcap": mktcap, "ev": ev, "netdebt": netdebt, "gross": gross,
        "nopat": n("New Operating Income"), "nopat7sum": n("New Operating Income 7 years sum"),
        "roiic7": n("ROICm_total - 7 years"), "roiic21": n("ROICm_total - 21 years"),
        "rr": n("Reinvestment Rate"), "ic": n("Invested Capital"), "ebita": n("EBITA"),
        "sales": n("Sales"), "tax": n("Income Tax Rate - Instrument"),
        "equity": n("Shareholders' Equity - Attributable to Parent ShHold - Total"),
        "repro": n("Gross Reproduction Cost"), "ppe_net": ppe_net,
        "div": n("Dividends Paid - Cash - Total - Cash Flow"),
    }


def main():
    m = load_engine()
    import openpyxl
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("file")
    ap.add_argument("--re", type=float, default=0.09, help="required equity return (default 0.09)")
    ap.add_argument("--moat-assume", type=float, default=7.5, help="assumed Moat Score (no column in file)")
    ap.add_argument("--country-base", default=None)
    ap.add_argument("--only", default=None, help="value only one archetype: operating|cyclical|regulated|nav")
    ap.add_argument("--horizon", type=int, default=5)
    ap.add_argument("--sheet", default=None)
    ap.add_argument("--csv", default=None)
    args = ap.parse_args()
    ov = m.parse_kv_rates(args.country_base)
    wb = openpyxl.load_workbook(args.file, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb.worksheets[0]
    H = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

    out_rows = []
    counts = {}
    for row in latest_rows(ws, H):
        d = drivers(ws, H, row)
        if not d["mktcap"] or not d["nopat"] or d["netdebt"] is None:
            continue
        arch = archetype_for(d["ind"])
        counts[arch] = counts.get(arch, 0) + 1
        if arch == "financial":
            continue
        if args.only and arch != args.only:
            continue
        try:
            if arch == "cyclical":
                v = value_cyclical(m, d, args.re, args.moat_assume, ov, args.horizon)
            elif arch == "regulated":
                v = value_regulated(m, d, args.re, args.moat_assume, ov, args.horizon)
            elif arch == "nav":
                v = value_nav(m, d, args.re, ov, args.horizon)
            else:
                v = value_operating(m, d, args.re, args.moat_assume, ov, horizon=args.horizon)
        except Exception as e:
            continue
        ev = d["ev"]
        v.update({"arch": arch, "name": d["name"], "tick": d["tick"], "ind": d["ind"],
                  "co": d["country"], "mktcap": d["mktcap"], "ev": ev,
                  "mev": (v["total"] / ev) if (ev and v["total"]) else None})
        out_rows.append(v)

    _report(m, args, out_rows, counts)
    if args.csv:
        _csv(args, out_rows)


def _report(m, args, rows, counts):
    print(f"\nAIP SECTOR VALUATION — re {args.re*100:.0f}%, moat assumed {args.moat_assume}, "
          f"horizon {args.horizon}y")
    print(f"universe by archetype: {counts}  (financials skipped)\n")
    order = ["operating", "cyclical", "regulated", "nav"]
    for arch in (([args.only] if args.only else order)):
        grp = [r for r in rows if r["arch"] == arch]
        if not grp:
            continue
        grp.sort(key=lambda r: (r["er"] is not None, r["er"] if r["er"] is not None else -9), reverse=True)
        print(f"=== {arch.upper()} ({len(grp)}) — top/bottom by expected return ===")
        hdr = f"  {'Company':32}{'Ctry':5}{'MCap':>8}{'Model/EV':>9}{'ER':>7}  note"
        print(hdr)
        show = grp[:12] + ([("...",)] if len(grp) > 18 else []) + (grp[-6:] if len(grp) > 18 else [])
        for r in show:
            if isinstance(r, tuple):
                print("   ...")
                continue
            mev = f"{r['mev']:.2f}x" if r["mev"] else "  n/a"
            er = f"{r['er']*100:5.1f}%" if r["er"] is not None else "  n/a"
            note = r.get("extra", "")
            print(f"  {str(r['name'])[:31]:32}{_cc(r['co']):5}{r['mktcap']/1e9:6.1f}b{mev:>9}{er:>7}  {note[:40]}")
        print()
    print("Cyclical = cycle-normalised ROIC + reproduction-cost floor; Regulated = excess "
          "return faded to WACC (value~RAB); NAV = book equity + reproduction surplus.\n"
          "Financials (Banks/Insurance/Financial Services) need a residual-income model — not built.\n"
          "Analytical framework output, not investment advice.")


_CC = {'United States of America': 'US', 'Japan': 'JP', 'United Kingdom': 'UK', 'Sweden': 'SE',
       'Germany': 'DE', 'Canada': 'CA', 'France': 'FR', 'Switzerland': 'CH', 'China': 'CN',
       'India': 'IN', 'Saudi Arabia': 'SA', 'United Arab Emirates': 'AE', 'Indonesia': 'ID',
       'Brazil': 'BR', 'Philippines': 'PH', 'Australia': 'AU', 'Netherlands': 'NL'}


def _cc(c):
    return _CC.get(c, (str(c)[:4] if c else "?"))


def _csv(args, rows):
    with open(args.csv, "w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["Archetype", "Company", "Ticker", "Country", "Industry", "MktCap_bn",
                    "EV_bn", "Model_value_bn", "Model/EV", "ExpReturn", "Method", "Note"])
        for r in sorted(rows, key=lambda x: (x["arch"], -(x["er"] or -9))):
            w.writerow([r["arch"], r["name"], r["tick"], r["co"], r["ind"],
                        f"{r['mktcap']/1e9:.2f}", f"{r['ev']/1e9:.2f}" if r["ev"] else "",
                        f"{r['total']/1e9:.2f}" if r["total"] else "",
                        f"{r['mev']:.2f}" if r["mev"] else "",
                        f"{r['er']*100:.1f}%" if r["er"] is not None else "n/a",
                        r["method"], r.get("extra", "")])
    print(f"saved -> {args.csv}")


if __name__ == "__main__":
    main()
