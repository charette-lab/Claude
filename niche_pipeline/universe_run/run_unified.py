#!/usr/bin/env python3
"""Fully research-aware, current-price run over the whole universe.

Values every company at current prices from the LTM screener, runs the empirical
history cross-check and over-earning normalization off the 30-year .xlsb panel
(via panel30), and selects the Core Index / Satellite book.

Research records: real analyst records are used wherever they exist in the cache.
For names whose RAW records are not in the cache but whose DERIVED moat is known
(from a prior scored book), each name's moat scalar is injected by solving for the
chapter level that reproduces it exactly (ncc_score is monotonic) — so the
selection runs on the true researched moats without re-paying for research and
without fabricating any value that reaches the output (the book emits the moat
scalar, never the chapter vector).

    ANTHROPIC_API_KEY unset is fine (research=False; nothing is fetched).
    AIP_VALUE_ENGINE=/path/to/roiic_dcf.py \
    python3 run_unified.py \
        --ltm   LTM_current.xlsx \
        --panel 30_file_1.xlsb 30_file_2.xlsb 30_file_3.xlsb \
        --prior Universe_scored_with_research.xlsx \
        --out   Universe_final.xlsx

If you have the ORIGINAL .cache for every name, you don't need --prior at all —
just run niche_pipeline/complete_universe.py; this script exists for the case
where only the derived moats (in a prior book) are available for some names.
"""
import argparse, json, os, sys, hashlib, shutil, tempfile
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))


def solve_chapter(target, ncc_score):
    """Chapter level c such that ncc_score([c]*10) == target (monotonic bisection)."""
    if not isinstance(target, (int, float)):
        return None
    lo, hi = 0.0, 10.0
    for _ in range(60):
        mid = (lo + hi) / 2
        if ncc_score([mid] * 10) < target:
            lo = mid
        else:
            hi = mid
    return (lo + hi) / 2


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--ltm", required=True, help="current LTM screener (prices)")
    ap.add_argument("--panel", nargs="+", required=True, help="30-yr .xlsb history parts")
    ap.add_argument("--prior", help="prior scored book with derived moats (CompanyMoat/CoreMoat/RiskTags)")
    ap.add_argument("--out", required=True)
    ap.add_argument("--real-cache", default=os.path.join(os.path.dirname(__file__), "..", ".cache"))
    args = ap.parse_args()

    tmp = tempfile.mkdtemp(prefix="unified_cache_")
    os.environ["NCP_CACHE"] = tmp
    import openpyxl, frameworks as F, analyst, panel30, history, pipeline

    # inject derived moats from the prior book for names lacking a raw cache record
    real = synth = 0
    if args.prior:
        wb = openpyxl.load_workbook(args.prior, data_only=True, read_only=True); ws = wb["Scored"]
        rows = list(ws.iter_rows(values_only=True)); ci = {h: i for i, h in enumerate(rows[0])}
        def g(r, c): return r[ci[c]] if c in ci else None
        wl = openpyxl.load_workbook(args.ltm, data_only=True, read_only=True); lw = wl.worksheets[0]
        lci = {h: i for i, h in enumerate([c.value for c in next(lw.iter_rows(min_row=1, max_row=1))])}
        universe = {r[lci["Instrument"]] for r in lw.iter_rows(min_row=2, values_only=True) if r[lci["Instrument"]]}
        def tags_vec(s):
            present = {x.strip() for x in (s or "").split(",") if x.strip()}
            return [1 if n in present else 0 for n in F.RISK_TAGS_SHORT]
        for r in rows[1:]:
            t = r[ci["Ticker"]]
            if t not in universe:
                continue
            realp = os.path.join(args.real_cache, hashlib.md5((t or "").encode()).hexdigest()[:10] + ".json")
            dst = analyst._cache_path(t)
            if os.path.exists(realp):
                shutil.copyfile(realp, dst); real += 1; continue
            cm, km = g(r, "CompanyMoat(v3.2)"), g(r, "CoreMoat(v3.2)")
            if not isinstance(cm, (int, float)):
                continue
            json.dump({
                "ticker": t, "company": g(r, "Company"), "core_business": g(r, "CoreBusiness"),
                "chapters_company": [solve_chapter(cm, F.ncc_score)] * 10,
                "chapters_core": [solve_chapter(km if isinstance(km, (int, float)) else cm, F.ncc_score)] * 10,
                "core_detach": {}, "largest_bloc_pct": g(r, "OwnerBloc%"),
                "ownership_notes": g(r, "OwnerNotes") or "", "risk_tags": tags_vec(g(r, "RiskTags")),
                "risk_notes": "", "sources": [],
            }, open(dst, "w")); synth += 1
        print(f"cache prepared: real={real} synthetic={synth}")

    by, idx = panel30.load(args.panel)
    history.load_panel = lambda *a, **k: (by, idx)
    pipeline.run(args.ltm, args.out, re=0.07, re2=0.12, research=False,
                 history_path="merged", apply_overearning=True)
    print("DONE")


if __name__ == "__main__":
    main()
