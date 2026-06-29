"""
history.py — empirical moat / quality validation from the long-run panel.

The qualitative moat scores (from analyst.py) are a narrative read; this module
tests them against a company's actual multi-year economics, in the spirit of the
v3.2 Domain-B rule ("third-party standardized data — actual ROIC — takes
precedence over the corporate narrative").

Input: the wide time-series panel (one row per company per year). Column meanings
(per the screener's author):
  * "Ratio"                      = ROIC (level); "Ratio_rolling_avg_7yr" = 7y avg
  * "ROICm_total - 7 years"      = marginal / incremental ROIC over 7y
  * "average_C - 7 year"         = average reinvestment rate over 7y
  * "EBITA_Margin"               = latest EBITA margin
  * "EBITA_Average_Margin 10 yr" = 10y average EBITA margin
  * "current_multiple"           = cycle-normalised EV/EBITA (10y-avg margin x current sales)
  * "d_Direktavkastning"         = (dividend + buybacks) / price  (total shareholder yield)
  * impairments / restructuring / M&A  = capital-allocation quality

For each company it returns an empirical summary and a verdict that cross-checks
the qualitative core moat:
  CONFIRMED      - returns on capital comfortably support a moat
  CONTRADICTED   - a Watchlist+ moat was claimed but the economics don't back it
                   (weak/again-declining ROIC or eroding margins)  <- the Fagerhult case
  SOFT           - in between / inconclusive
"""
from __future__ import annotations
import openpyxl
import capital

# thresholds (returns-on-capital are judged vs a ~8% cost of capital, not the 12% hurdle)
COST_OF_CAPITAL = 0.08
ROIC_STRONG = 0.12
ROIC_WEAK = 0.08
MARGIN_ERODING = 0.75      # latest / 10y-avg below this = material erosion
MARGIN_STABLE = 0.90


def _num(v):
    return v if isinstance(v, (int, float)) else None


def load_panel(path, sheet=None):
    """Return {ticker: [rows sorted by date]} and the column index.

    Reads .xlsx via openpyxl and .xlsb (binary) via pyxlsb — the large
    whole-universe panels are often exported as .xlsb to stay under size limits."""
    if str(path).lower().endswith(".xlsb"):
        hdr, rowiter = _load_xlsb(path, sheet)
    else:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[sheet] if sheet else wb.worksheets[0]
        hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        rowiter = ws.iter_rows(min_row=2, values_only=True)
    idx = {h: i for i, h in enumerate(hdr) if h is not None}
    by = {}
    ti, di = idx.get("Instrument"), idx.get("Date")
    for r in rowiter:
        t = r[ti] if (ti is not None and ti < len(r)) else None
        if t:
            by.setdefault(t, []).append(r)
    keyd = (lambda r: (di is not None and di < len(r) and r[di] is not None,
                       r[di] if (di is not None and di < len(r)) else None))
    for t in by:
        by[t].sort(key=keyd)
    return by, idx


def _load_xlsb(path, sheet=None):
    """(header_list, row_value_iterator) for an .xlsb panel via pyxlsb.
    Rows are padded to the header width; pyxlsb yields sparse cells."""
    import pyxlsb
    wb = pyxlsb.open_workbook(path)
    ws = wb.get_sheet(sheet or wb.sheets[0])
    rows = ws.rows()
    hdr_cells = next(rows)
    ncol = max((c.c for c in hdr_cells), default=-1) + 1
    hdr = [None] * ncol
    for c in hdr_cells:
        if c.c < ncol:
            hdr[c.c] = c.v

    def gen():
        for rcells in rows:
            row = [None] * ncol
            for c in rcells:
                if c.c < ncol:
                    row[c.c] = c.v
            yield row
    return hdr, gen()


def summarize(rows, idx):
    """Empirical economics for one company from its yearly rows."""
    def col(name):
        i = idx.get(name)
        return [_num(r[i]) for r in rows] if i is not None else []
    last = lambda name: next((v for v in reversed(col(name)) if v is not None), None)

    margins = [v for v in col("EBITA_Margin") if v is not None]
    m_last = margins[-1] if margins else None
    m_avg10 = last("EBITA_Average_Margin 10 yr")
    margin_ratio = (m_last / m_avg10) if (m_last and m_avg10 and m_avg10 > 0) else None

    s = {
        # The through-cycle, base-consistent quality anchor (CAPITAL_SPEC §5):
        # NOI / invested-capital-with-intangibles-and-10%-cash, 7yr-rolling. This
        # is the binding veto signal — cycle-stable, unlike the trough-distorted
        # marginal ROIC. Recomputed here, not read from the panel's full-cash Ratio.
        "roic_star": capital.roic_star(rows, idx),
        "roic_level": last("Ratio_rolling_avg_7yr") or last("Ratio"),
        "roic_marginal_7y": last("ROICm_total - 7 years"),
        "roic_marginal_14y": last("ROICm_total - 14 years"),
        "reinvest_7y": last("average_C - 7 year"),
        "margin_last": m_last, "margin_avg10": m_avg10, "margin_ratio": margin_ratio,
        "norm_ev_ebita": last("current_multiple"),
        "shareholder_yield": last("d_Direktavkastning"),
        "sales_growth_last": last("Sales Growth"),
        "impairments": sum(v for v in col("Impairment - Goodwill (ASR)") if v) +
                       sum(v for v in col("Impairment - Intangibles excluding Goodwill (ASR)") if v),
        "restructuring": sum(v for v in col("Restructuring Charges") if v),
        "years": len(rows),
    }
    return s


def verdict(summary, core_moat):
    """Cross-check the qualitative core moat against the economics.

    The binding evidence is the THROUGH-CYCLE quality anchor ROIC* (the base-
    consistent level ROIC, 7yr-rolling — CAPITAL_SPEC §5), NOT a single trough-
    distorted marginal ROIC. Averaging levels avoids the endpoint collapse that
    knifes a quality cyclical caught at its trough (Somero: 7y marginal 2% but
    ROIC* ~29%). A long-window marginal that is also negative only CONFIRMS
    genuine capital destruction; it does not by itself veto."""
    rs = summary.get("roic_star")            # the binding through-cycle anchor
    rm14 = summary.get("roic_marginal_14y")  # cycle-robust destruction confirmation
    rm = summary.get("roic_marginal_7y")     # trough-sensitive; reported, not binding
    mr = summary.get("margin_ratio")
    reasons = []
    # Weak = the franchise does not earn its cost of capital THROUGH THE CYCLE.
    weak = rs is not None and rs < COST_OF_CAPITAL
    destroying = weak and (rm14 is not None and rm14 < 0)
    eroding = mr is not None and mr < MARGIN_ERODING
    strong = (rs is not None and rs >= ROIC_STRONG) and not eroding
    if rs is not None:
        reasons.append(f"through-cycle ROIC* {rs*100:.0f}%"
                       + (f" below ~{COST_OF_CAPITAL*100:.0f}% cost of capital" if weak else ""))
    if destroying:
        reasons.append(f"long-window (14y) marginal ROIC NEGATIVE ({rm14*100:.0f}%) — destroying capital")
    if eroding:
        reasons.append(f"caution: EBITA margin at {mr*100:.0f}% of its 10y average "
                       + ("(but through-cycle ROIC* healthy — growth/cyclical, not decline)"
                          if not weak else ""))
    claimed = core_moat is not None and core_moat >= 6.5
    if claimed and weak:
        return "CONTRADICTED", reasons
    if strong and not weak:
        return "CONFIRMED", reasons
    return "SOFT", reasons
