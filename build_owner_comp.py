"""Total compensation for the equity owners (CIO + two PMs):
 base salary + performance take (weight x unit value) + equity distribution (ownership% x EBIT).
Adds an 'Owner Total Comp' sheet (formulas link to Static & Profit Simulation) + PNG."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
import matplotlib; matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np

WB="/home/user/Claude/Performance_Fee_Framework.xlsx"
PNG="/home/user/Claude/Owner_total_comp.png"
scn=[20,300,500,1000,2000,5000]
THR=300
# from prior reconciled runs
unit=[0.006605,0.080604,0.102707,0.167442,0.252632,0.631579]   # value per performance unit (with new roles)
ebit=[-1.4174,3.0911,6.5389,15.2659,36.2974,110.0974]
# owners: (name, salary<300, salary>300, weight, equity%)
owners=[("CIO",0.20,0.30,16,0.636),
        ("Portfolio Manager 1",0.13,0.20,8,0.10),
        ("Portfolio Manager 2",0.13,0.20,8,0.10)]

def comp(o):
    nm,lo,hi,wt,eq=o
    sal=[lo if a<THR else hi for a in scn]
    perf=[wt*unit[j] for j in range(6)]
    equity=[eq*ebit[j] for j in range(6)]
    tot=[sal[j]+perf[j]+equity[j] for j in range(6)]
    return sal,perf,equity,tot

print("AUM:",scn)
for o in owners:
    sal,perf,equity,tot=comp(o)
    print(f"\n{o[0]} (wt {o[3]}, equity {o[4]:.1%})")
    print("  salary  ",["%7.3f"%x for x in sal])
    print("  perform ",["%7.3f"%x for x in perf])
    print("  equity  ",["%7.3f"%x for x in equity])
    print("  TOTAL   ",["%7.3f"%x for x in tot])

# ---------------- workbook sheet ----------------
wb=openpyxl.load_workbook(WB)
if "Owner Total Comp" in wb.sheetnames: del wb["Owner Total Comp"]
ws=wb.create_sheet("Owner Total Comp")
NAVY="1F3864";BLUE="2E5496";LIGHTER="EAF0FA";GOLD="FFF2CC";GREEN="E2EFDA";GREY="F2F2F2"
fnavy=PatternFill("solid",fgColor=NAVY);fhdr=PatternFill("solid",fgColor=BLUE);flighter=PatternFill("solid",fgColor=LIGHTER)
fgold=PatternFill("solid",fgColor=GOLD);fgreen=PatternFill("solid",fgColor=GREEN);fgrey=PatternFill("solid",fgColor=GREY)
title_f=Font(size=14,bold=True,color="FFFFFF");sub_f=Font(size=10,italic=True,color="FFFFFF")
hdr_f=Font(size=10,bold=True,color="FFFFFF");bold=Font(size=10,bold=True);norm=Font(size=10);ital=Font(size=9,italic=True,color="595959")
wfw=Font(size=11,bold=True,color="FFFFFF")
ctr=Alignment("center","center");lft=Alignment("left","center",wrap_text=True)
thin=Side("thin",color="BFBFBF");box=Border(thin,thin,thin,thin)
SC=["H","I","J","K","L","M"];MM='#,##0.000'
def s(coord,v,f=norm,fill=None,al=None,fmt=None,bd=True):
    c=ws[coord];c.value=v;c.font=f
    if fill:c.fill=fill
    if al:c.alignment=al
    if fmt:c.number_format=fmt
    if bd:c.border=box
    return c
ws.merge_cells("A1:M1");s("A1","Total compensation for equity owners — CIO and the two PMs",title_f,fnavy,ctr,bd=False)
ws.merge_cells("A2:M2");s("A2","Total comp = base salary + performance take (weight × unit value) + equity distribution (ownership × EBIT). Millions.",sub_f,fnavy,ctr,bd=False)
ws.row_dimensions[1].height=22
# Static role rows: CIO=21, PM=22 ; unit value row=35 ; EBIT in 'Profit Simulation'!14 ; threshold Static!F9
rolerow={"CIO":21,"Portfolio Manager 1":22,"Portfolio Manager 2":22}
r=4
s(f"A{r}","Component  (millions)",hdr_f,fhdr,lft)
for i,col in enumerate(SC):s(f"{col}{r}",f"AUM ${scn[i]:,}m",hdr_f,fhdr,ctr)
r=5
for (nm,lo,hi,wt,eq) in owners:
    rr=rolerow[nm]
    s(f"A{r}",f"{nm}  (wt {wt}, equity {eq:.1%})",wfw,fhdr,lft)
    for col in SC:s(f"{col}{r}","",norm,fhdr)
    r+=1
    s(f"A{r}","  Base salary",norm,flighter,lft)
    for col in SC:s(f"{col}{r}",f"=IF(Static!{col}15<Static!$F$9,Static!$B${rr},Static!$C${rr})",norm,flighter,ctr,MM)
    r+=1
    s(f"A{r}","  Performance take  (weight × unit value)",norm,flighter,lft)
    for col in SC:s(f"{col}{r}",f"=Static!$D${rr}*Static!{col}$35",norm,flighter,ctr,MM)
    r+=1
    s(f"A{r}",f"  Equity distribution  ({eq:.1%} × EBIT)",norm,fgold,lft)
    s(f"G{r}",eq,Font(size=10,bold=True,color="C00000"),fgold,ctr,'0.0%')
    for col in SC:s(f"{col}{r}",f"='Profit Simulation'!{col}14*$G{r}",norm,fgold,ctr,MM)
    r+=1
    s(f"A{r}",f"  TOTAL COMP — {nm}",bold,fgreen,lft)
    for col in SC:s(f"{col}{r}",f"=SUM({col}{r-3}:{col}{r-1})",bold,fgreen,ctr,MM)
    r+=2
# combined: two PMs together
s(f"A{r}","Both PMs combined (equity 20%)",bold,fgrey,lft)
for col in SC:s(f"{col}{r}",f"={col}13*2",bold,fgrey,ctr,MM)  # PM1 total is at row 13
ws.column_dimensions["A"].width=42
for c in "BCDEF":ws.column_dimensions[c].width=6
ws.column_dimensions["G"].width=8
for col in SC:ws.column_dimensions[col].width=12
ws.sheet_view.showGridLines=False;ws.freeze_panes="B5"
# chart: CIO total + PM total
bc=BarChart();bc.type="col";bc.title="Total comp: CIO vs PM (each), by AUM (millions)";bc.height=9;bc.width=20
cats=Reference(ws,min_col=8,min_row=4,max_col=13,max_row=4)
cio_tot=Reference(ws,min_col=1,min_row=9,max_col=7,max_row=9)   # CIO TOTAL row
pm_tot=Reference(ws,min_col=1,min_row=14,max_col=7,max_row=14)  # PM1 TOTAL row
bc.add_data(cio_tot,titles_from_data=True,from_rows=True)
bc.add_data(pm_tot,titles_from_data=True,from_rows=True)
bc.set_categories(cats);ws.add_chart(bc,f"A{r+2}")
wb.save(WB)
print("\nSheets:",wb.sheetnames)

# ---------------- PNG: stacked components ----------------
x=np.arange(6);labels=[f"${a:,}m" for a in scn]
fig,axes=plt.subplots(1,3,figsize=(17,6),sharey=True)
comp_colors={"Base salary":"#BFBFBF","Performance take":"#2E75B6","Equity distribution":"#C8A23C"}
for ax,o in zip(axes,owners):
    sal,perf,equity,tot=comp(o)
    bottoms_pos=np.zeros(6);bottoms_neg=np.zeros(6)
    for label,vals in [("Base salary",sal),("Performance take",perf),("Equity distribution",equity)]:
        vals=np.array(vals);bottoms=np.where(vals>=0,bottoms_pos,bottoms_neg)
        ax.bar(x,vals,bottom=bottoms,color=comp_colors[label],edgecolor="black",linewidth=0.3,label=label)
        bottoms_pos=bottoms_pos+np.where(vals>=0,vals,0);bottoms_neg=bottoms_neg+np.where(vals<0,vals,0)
    ax.plot(x,tot,"D",color="black",ms=6,label="Total comp")
    for xi,t in zip(x,tot):ax.text(xi,t+(1.0 if t>=0 else -1.8),f"{t:.1f}",ha="center",fontsize=8,fontweight="bold")
    ax.axhline(0,color="black",lw=0.8);ax.set_xticks(x);ax.set_xticklabels(labels,rotation=30)
    ax.set_title(f"{o[0]}  (equity {o[4]:.1%})",fontweight="bold");ax.grid(True,axis="y",alpha=0.3)
axes[0].set_ylabel("Total compensation (millions)");axes[0].legend(fontsize=9,frameon=False,loc="upper left")
fig.suptitle("Overall compensation of equity owners = salary + performance take + equity distribution",fontsize=14,fontweight="bold")
plt.tight_layout();plt.savefig(PNG,dpi=130,bbox_inches="tight");print("PNG done")
