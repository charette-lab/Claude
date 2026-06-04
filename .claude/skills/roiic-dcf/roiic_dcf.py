#!/usr/bin/env python3
"""Three-stage ROIIC-driven DCF valuation, run against an attached Excel file.

Economic premise: growth requires capital (g = ROIIC x RR), and a company's
competitive advantage fades over time so returns regress toward the cost of
capital. Value is built in three stages:

  Stage 1  Competitive Advantage Period (CAP)  -> high ROIIC, n1 years
  Stage 2  Fade Period                         -> ROIIC midpoints to r, n2 years
  Stage 3  Terminal                            -> ROIIC = r (growth value-neutral)

The script reads a screener-style spreadsheet (one company per row, headers in
the first row), locates the requested company, pulls the three driver inputs,
and prints the full valuation with a market comparison.

DEFAULT EXCEL COLUMN MAPPING (override with --col-* flags):
  NOPAT_0   <- "New Operating Income"
  ROIIC_1   <- "ROICm 7"
  RR_1      <- "RR 7"
  (comparison) "EV", "Market Cap", "Net debt",
               "Shares used to calculate Diluted EPS - Total", "Close Price"

DEFAULT PARAMETERS (override with flags):
  r       = 0.12     cost of capital / discount rate (required return)
  n1, n2  = from the Moat Score: total competitive life split 1/3 (CAP) : 2/3 (Fade)
  g_term  = 0.025    terminal growth (must be < r)

STAGE LOGIC (fixed by the framework + the user's mapping):
  Stage 1:  ROIIC_1 = ROICm7;  RR_1 = RR7;  g1 = ROIIC_1 * RR_1
  Stage 2:  ROIIC_2 = (ROIIC_1 + r) / 2;  RR_2 = RR_1 * 1.5;  g2 = ROIIC_2 * RR_2
  Terminal: ROIIC_term = r;  RR_term = g_term / r  (growth adds no value)

Usage:
  python3 roiic_dcf.py <file.xlsx> "<company name substring>" [options]
  python3 roiic_dcf.py <file.xlsx> --list        # list companies in the sheet

Options:
  --r FLOAT --n1 INT --n2 INT --gterm FLOAT --sheet NAME
  --col-nopat STR --col-roiic STR --col-rr STR --col-name STR
"""

import argparse
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")


def growing_annuity_pv(cf1, r, g, n):
    """PV at time 0 of an n-year annuity whose first payment cf1 occurs at t=1
    and grows at g. Handles the g == r limit safely."""
    if abs(r - g) < 1e-12:
        return n * cf1 / (1 + r)
    return cf1 / (r - g) * (1 - ((1 + g) / (1 + r)) ** n)


def moat_to_life(score):
    """Total competitive period (years) implied by a Moat Score.

    Thresholds (Moat Life Estimation by Score):
      score < 6.0      -> short-term moat, < 10y   (linear 0->10 over 0->6)
      6.0 <= s <= 7.5  -> medium-term,    10-20y   (linear 10->20 over 6->7.5)
      score > 7.5      -> permanent moat,  50y     ("50+")
    Returns a whole number of years (>= 1), or None if score is missing. The
    total life is split 1/3 into Stage 1 (CAP) and 2/3 into Stage 2 (Fade).
    """
    if score is None:
        return None
    if score < 6.0:
        years = score / 6.0 * 10.0
    elif score <= 7.5:
        years = 10.0 + (score - 6.0) / (7.5 - 6.0) * 10.0
    else:
        years = 50.0
    return max(1, int(round(years)))


def split_life(life):
    """Split a total competitive period into (n1, n2) = (1/3, 2/3), >= 1y each."""
    n1 = max(1, int(round(life / 3.0)))
    n2 = max(1, int(round(life * 2.0 / 3.0)))
    return n1, n2


def moat_band(score):
    if score is None:
        return "n/a"
    if score < 6.0:
        return "Pass / short-term"
    if score <= 7.5:
        return "Watchlist / medium-term"
    return "Compounder / permanent"


def find_columns(ws, wanted):
    """Map each wanted header string to its 1-based column index (exact, then
    case-insensitive, then prefix match on the first row)."""
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is not None:
            headers[str(v).strip()] = c
    lower = {k.lower(): idx for k, idx in headers.items()}
    out = {}
    for key, label in wanted.items():
        if label in headers:
            out[key] = headers[label]
        elif label.lower() in lower:
            out[key] = lower[label.lower()]
        else:
            match = next((idx for h, idx in lower.items()
                          if h.startswith(label.lower())), None)
            out[key] = match
    return out


def find_company_row(ws, name_col, query):
    q = query.strip().lower()
    matches = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=name_col).value
        if v and q in str(v).lower():
            matches.append((r, str(v)))
    return matches


def num(ws, row, col):
    if col is None:
        return None
    v = ws.cell(row=row, column=col).value
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def money(x):
    if x is None:
        return "n/a"
    a = abs(x)
    if a >= 1e9:
        return f"{x/1e9:,.2f} bn"
    if a >= 1e6:
        return f"{x/1e6:,.1f} mn"
    return f"{x:,.0f}"


def pct(x):
    return "n/a" if x is None else f"{x*100:.2f}%"


def main():
    ap = argparse.ArgumentParser(add_help=True, description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("file")
    ap.add_argument("company", nargs="?", default=None)
    ap.add_argument("--list", action="store_true", help="list companies and exit")
    ap.add_argument("--r", type=float, default=0.12)
    ap.add_argument("--n1", type=int, default=None,
                    help="Stage 1 (CAP) length in years; if omitted = 1/3 of the "
                         "Moat-Score competitive life")
    ap.add_argument("--n2", type=int, default=None,
                    help="Stage 2 (Fade) length in years; if omitted = 2/3 of the "
                         "Moat-Score competitive life")
    ap.add_argument("--gterm", type=float, default=0.025)
    ap.add_argument("--horizon", type=int, default=5,
                    help="holding period in years for the expected-return / IRR (default 5)")
    ap.add_argument("--payout-total", type=float, default=0.0,
                    help="cumulative dividends + net buybacks over the horizon, "
                         "subtracted from the cash sweep (default 0 = all FCF de-levers)")
    ap.add_argument("--sheet", default=None)
    ap.add_argument("--col-nopat", default="New Operating Income")
    ap.add_argument("--col-roiic", default="ROICm 7")
    ap.add_argument("--col-rr", default="RR 7")
    ap.add_argument("--col-name", default="Company Name")
    ap.add_argument("--col-moat", default="Moat Score")
    args = ap.parse_args()

    if args.gterm >= args.r:
        sys.exit(f"g_term ({args.gterm}) must be < r ({args.r}) for a finite "
                 "terminal value.")

    wb = openpyxl.load_workbook(args.file, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb.worksheets[0]

    cols = find_columns(ws, {
        "name": args.col_name, "nopat": args.col_nopat,
        "roiic": args.col_roiic, "rr": args.col_rr, "moat": args.col_moat,
        "ev": "EV", "mktcap": "Market Cap", "netdebt": "Net debt",
        "shares": "Shares used to calculate Diluted EPS - Total",
        "price": "Close Price", "ticker": "Instrument",
    })
    if cols["name"] is None:
        sys.exit(f'Could not find a "{args.col_name}" column in row 1.')

    if args.list or not args.company:
        print(f"Companies in '{ws.title}':")
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=cols["name"]).value
            if v:
                print(f"  {v}")
        if not args.company and not args.list:
            print("\nPass a company name substring to value one.")
        return

    matches = find_company_row(ws, cols["name"], args.company)
    if not matches:
        sys.exit(f'No company matching "{args.company}". Use --list to see options.')
    if len(matches) > 1:
        names = ", ".join(n for _, n in matches)
        sys.exit(f'Ambiguous "{args.company}" matches: {names}. Be more specific.')
    row, company = matches[0]

    nopat0 = num(ws, row, cols["nopat"])
    roiic1 = num(ws, row, cols["roiic"])
    rr1 = num(ws, row, cols["rr"])
    if None in (nopat0, roiic1, rr1):
        sys.exit(f"Missing driver inputs for {company}: "
                 f"NOPAT={nopat0}, ROIIC={roiic1}, RR={rr1}")

    r, g_term = args.r, args.gterm

    # Total competitive period from the Moat Score, split 1/3 CAP : 2/3 Fade.
    moat = num(ws, row, cols["moat"])
    if moat is not None:
        life = moat_to_life(moat)
        n1_src = f"Moat Score {moat:.2f} [{moat_band(moat)}], life {life}y, split 1/3:2/3"
    else:
        life = 15
        n1_src = "no Moat Score; default life 15y, split 1/3:2/3"
    dn1, dn2 = split_life(life)
    n1 = args.n1 if args.n1 is not None else dn1
    n2 = args.n2 if args.n2 is not None else dn2

    # --- Stage parameters (Step 1: g = ROIIC x RR) ---
    g1 = roiic1 * rr1
    roiic2 = (roiic1 + r) / 2
    g2 = (g1 + g_term) / 2            # growth fades to the midpoint of g1 and terminal
    rr2 = g2 / roiic2 if abs(roiic2) > 1e-9 else 0.0   # reinvestment derived (Step 3)
    g2 = roiic2 * rr2                 # restate via the identity (no-op unless guarded)
    roiic_term = r
    rr_term = g_term / r

    # --- Stage 1 ---
    cf1 = nopat0 * (1 + g1) * (1 - rr1)
    pv1 = growing_annuity_pv(cf1, r, g1, n1)

    # --- Stage 2 ---
    nopat_end_s1 = nopat0 * (1 + g1) ** n1
    cf_s2 = nopat_end_s1 * (1 + g2) * (1 - rr2)
    pv2 = growing_annuity_pv(cf_s2, r, g2, n2) / (1 + r) ** n1

    # --- Terminal ---
    nopat_end_s2 = nopat_end_s1 * (1 + g2) ** n2
    cf_term = nopat_end_s2 * (1 + g_term) * (1 - rr_term)
    tv = cf_term / (r - g_term)
    pv_tv = tv / (1 + r) ** (n1 + n2)

    total = pv1 + pv2 + pv_tv

    # --- Output ---
    ticker = ws.cell(row=row, column=cols["ticker"]).value if cols["ticker"] else ""
    print(f"\nTHREE-STAGE ROIIC DCF — {company} {f'({ticker})' if ticker else ''}")
    print(f"r = {pct(r)}   n1 = {n1}y  n2 = {n2}y  (total {n1+n2}y; {n1_src})   g_term = {pct(g_term)}")
    print("=" * 66)
    print(f"  NOPAT_0 (New Operating Income) .... {money(nopat0)}")
    print(f"  ROIIC_1 (ROICm 7) ................. {pct(roiic1)}")
    print(f"  RR_1 (RR 7) ....................... {pct(rr1)}")

    print("\n  STAGE 1 — Competitive Advantage Period")
    print(f"    g1 = ROIIC_1 x RR_1 ............. {pct(g1)}")
    print(f"    CF_1 ........................... {money(cf1)}")
    print(f"    PV_1 ........................... {money(pv1)}")

    print("\n  STAGE 2 — Fade Period")
    print(f"    ROIIC_2 = (ROIIC_1 + r)/2 ...... {pct(roiic2)}")
    print(f"    g2 = (g1 + g_term)/2 .......... {pct(g2)}")
    print(f"    RR_2 = g2 / ROIIC_2 ........... {pct(rr2)}")
    print(f"    CF_Stage2_Yr1 ................. {money(cf_s2)}")
    print(f"    PV_2 .......................... {money(pv2)}")

    print("\n  STAGE 3 — Terminal (ROIIC_term = r)")
    print(f"    RR_term = g_term/r ............ {pct(rr_term)}")
    print(f"    CF_term_Yr1 ................... {money(cf_term)}")
    print(f"    TV (at end of Stage 2) ........ {money(tv)}")
    print(f"    PV_TV ......................... {money(pv_tv)}")

    print("\n" + "=" * 66)
    print(f"  TOTAL OPERATING VALUE ........... {money(total)}")
    print(f"    PV_1 {pv1/total*100:5.1f}%   PV_2 {pv2/total*100:5.1f}%   "
          f"PV_TV {pv_tv/total*100:5.1f}%")

    # --- Market comparison ---
    ev = num(ws, row, cols["ev"])
    netdebt = num(ws, row, cols["netdebt"])
    shares = num(ws, row, cols["shares"])
    price = num(ws, row, cols["price"])
    mktcap = num(ws, row, cols["mktcap"])
    print("\n  MARKET COMPARISON")
    if ev is not None:
        upside = total / ev - 1
        print(f"    Enterprise Value (market) ..... {money(ev)}")
        print(f"    Model / EV .................... {total/ev:.2f}x  "
              f"({'+' if upside>=0 else ''}{upside*100:.1f}% vs EV)")
    if netdebt is not None:
        equity = total - netdebt
        print(f"    Net debt (neg = net cash) ..... {money(netdebt)}")
        print(f"    Implied equity value .......... {money(equity)}")
        if shares:
            iv_ps = equity / shares
            line = f"    Implied value / share ......... {iv_ps:,.2f}"
            if price:
                line += f"   (market {price:,.2f}, {'+' if iv_ps>=price else ''}{(iv_ps/price-1)*100:.1f}%)"
            print(line)
    if mktcap is not None:
        print(f"    Market cap .................... {money(mktcap)}")

    # --- Expected return (IRR) over the holding horizon ---
    # EV_target = today's modelled operating value, assumed realised at exit.
    # Cash sweep: free cash flow over the horizon de-levers net debt
    # dollar-for-dollar (less any dividends/buybacks paid out).
    def cf_for_year(t):
        """Stage-aware free cash flow in year t (t >= 1)."""
        if t <= n1:
            nopat_t = nopat0 * (1 + g1) ** t
            return nopat_t * (1 - rr1)
        if t <= n1 + n2:
            nopat_t = nopat0 * (1 + g1) ** n1 * (1 + g2) ** (t - n1)
            return nopat_t * (1 - rr2)
        nopat_t = nopat_end_s2 * (1 + g_term) ** (t - n1 - n2)
        return nopat_t * (1 - rr_term)

    n = args.horizon
    print(f"\n  EXPECTED RETURN  (horizon n = {n}y)")
    cf_sum = sum(cf_for_year(t) for t in range(1, n + 1))
    print(f"    EV_target (operating value) ... {money(total)}")
    print(f"    Sum FCF yrs 1..{n} ............. {money(cf_sum)}")
    if args.payout_total:
        print(f"    - Dividends/buybacks (horizon)  {money(args.payout_total)}")
    if netdebt is not None:
        nd5 = netdebt - (cf_sum - args.payout_total)
        eqv_target = total - nd5
        print(f"    ND_0 (current net debt) ....... {money(netdebt)}")
        print(f"    ND_{n} (after cash sweep) ....... {money(nd5)}")
        print(f"    EqV_target = EV_target - ND_{n} . {money(eqv_target)}")
        if mktcap and mktcap > 0 and eqv_target > 0:
            eq_irr = (eqv_target / mktcap) ** (1 / n) - 1
            print(f"    EqV_0 (current market cap) .... {money(mktcap)}")
            print(f"    >> Expected equity return (IRR) {pct(eq_irr)} / yr")
        elif eqv_target <= 0:
            print("    >> Projected equity value <= 0; equity IRR undefined.")
    if ev is not None and ev > 0 and total > 0:
        unlev = (total / ev) ** (1 / n) - 1
        print(f"    >> Unlevered return (EV_target/EV_0) {pct(unlev)} / yr")
        print("    (If equity IRR >> unlevered, the thesis leans on leverage.)")

    print("\n  Analytical framework output, not investment advice.")


if __name__ == "__main__":
    main()
