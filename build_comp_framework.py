"""
Build the 'Static' performance-fee distribution framework into the workbook.

Model logic
-----------
Performance fee  = Gross profit (AUM x assumed gross return) x performance-fee rate (20%)
Comp pool        = Performance fee x firm take rate (60%)        <- pool shared by staff
Performance unit = one share of the pool. Each role earns a fixed number of units
                   per person ("performance weight").
Total units      = SUMPRODUCT(weights, headcount)  -> grows as the firm hires
USD / unit       = Comp pool / Total units
Per-person bonus = weight x USD-per-unit
Role share %     = (headcount x weight) / Total units

As the firm grows, total units rise so each role's PERCENTAGE share falls, but because
the pool grows with AUM the ABSOLUTE USD per person increases. Fixed salaries step up
once AUM crosses the $300m threshold.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = "/root/.claude/uploads/cc747475-841f-5d65-b8c7-9a970e1dee23/f8b4b64a-Comparison_returns.xlsx"
OUT = "/home/user/Claude/Performance_Fee_Framework.xlsx"

wb = openpyxl.load_workbook(SRC)
if "Static" in wb.sheetnames:
    del wb["Static"]
ws = wb.create_sheet("Static", 0)  # place first

# ----- styles -----
NAVY = "1F3864"; BLUE = "2E5496"; LIGHT = "D6E0F0"; LIGHTER = "EAF0FA"
GREY = "F2F2F2"; GOLD = "FFF2CC"; GREEN = "E2EFDA"
title_f = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
sub_f = Font(name="Calibri", size=10, italic=True, color="FFFFFF")
hdr_f = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
sec_f = Font(name="Calibri", size=12, bold=True, color=NAVY)
bold = Font(name="Calibri", size=10, bold=True)
norm = Font(name="Calibri", size=10)
ital = Font(name="Calibri", size=9, italic=True, color="595959")
inp_f = Font(name="Calibri", size=10, bold=True, color="C00000")

fill_navy = PatternFill("solid", fgColor=NAVY)
fill_blue = PatternFill("solid", fgColor=BLUE)
fill_hdr = PatternFill("solid", fgColor=BLUE)
fill_light = PatternFill("solid", fgColor=LIGHT)
fill_lighter = PatternFill("solid", fgColor=LIGHTER)
fill_grey = PatternFill("solid", fgColor=GREY)
fill_gold = PatternFill("solid", fgColor=GOLD)
fill_green = PatternFill("solid", fgColor=GREEN)

ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
right = Alignment(horizontal="right", vertical="center")
thin = Side(style="thin", color="BFBFBF")
box = Border(left=thin, right=thin, top=thin, bottom=thin)

USD = '$#,##0'
USDm = '#,##0.0'
PCT = '0.0%'
NUM = '0'

SCN_COLS = ["H", "I", "J", "K", "L", "M"]  # 6 AUM scenarios

def setc(coord, value, font=norm, fill=None, align=None, fmt=None, border=False):
    c = ws[coord]
    c.value = value
    c.font = font
    if fill: c.fill = fill
    if align: c.alignment = align
    if fmt: c.number_format = fmt
    if border: c.border = box
    return c

# ============================================================ TITLE
ws.merge_cells("A1:M1")
setc("A1", "Performance Fee Distribution Framework", title_f, fill_navy, ctr)
ws.merge_cells("A2:M2")
setc("A2", "Athanase Industrial Partner  —  how the performance-fee pool is split by role", sub_f, fill_navy, ctr)
ws.row_dimensions[1].height = 26
ws.row_dimensions[2].height = 16

# ============================================================ GLOBAL INPUTS (red = editable)
setc("A4", "1.  Global assumptions  (edit the red cells)", sec_f)
inputs = [
    (5, "Performance-fee rate (of profits)", 0.20, PCT),
    (6, "Firm take rate — share of perf-fee funding the comp pool", 0.60, PCT),
    (7, "AUM tier threshold (USD millions)", 300, '#,##0'),
    (8, "Assumed gross return on AUM (drives the demo profit)", 0.15, PCT),
]
for r, label, val, fmt in inputs:
    ws.merge_cells(f"A{r}:E{r}")
    setc(f"A{r}", label, norm, fill_grey, left, border=True)
    setc(f"F{r}", val, inp_f, fill_gold, ctr, fmt, border=True)

setc("A9", "Comp pool = Performance fee × take rate. The other 40% is retained by the firm.", ital)

# ============================================================ SCENARIO ECONOMICS
setc("A11", "2.  AUM growth scenarios & performance-fee economics", sec_f)
# header row 12
setc("A12", "Metric  (USD millions unless noted)", hdr_f, fill_hdr, left, border=True)
for i, col in enumerate(SCN_COLS, 1):
    setc(f"{col}12", f"Scenario {i}", hdr_f, fill_hdr, ctr, border=True)

aum = [100, 300, 500, 1000, 2000, 5000]
# row 13 AUM (editable)
setc("A13", "AUM  (USD m)", bold, fill_light, left, border=True)
for col, a in zip(SCN_COLS, aum):
    setc(f"{col}13", a, inp_f, fill_gold, ctr, '#,##0', border=True)
# row 14 gross return (links to global, editable per scenario)
setc("A14", "Assumed gross return", norm, fill_lighter, left, border=True)
for col in SCN_COLS:
    setc(f"{col}14", "=$F$8", norm, None, ctr, PCT, border=True)
# row 15 gross profit
setc("A15", "Gross profit  (= AUM × return)", norm, fill_lighter, left, border=True)
for col in SCN_COLS:
    setc(f"{col}15", f"={col}13*{col}14", norm, None, ctr, USDm, border=True)
# row 16 performance fee
setc("A16", "Performance fee  (= profit × fee rate)", norm, fill_lighter, left, border=True)
for col in SCN_COLS:
    setc(f"{col}16", f"={col}15*$F$5", norm, None, ctr, USDm, border=True)
# row 17 comp pool
setc("A17", "Comp pool  (= perf fee × take rate)", bold, fill_green, left, border=True)
for col in SCN_COLS:
    setc(f"{col}17", f"={col}16*$F$6", bold, fill_green, ctr, USDm, border=True)
# row 18 firm retained
setc("A18", "Firm retained  (= perf fee × 40%)", norm, fill_lighter, left, border=True)
for col in SCN_COLS:
    setc(f"{col}18", f"={col}16*(1-$F$6)", norm, None, ctr, USDm, border=True)

# ============================================================ ORG / WEIGHTS / HEADCOUNT  (rows 20-37)
setc("A20", "3.  Organisation, weights & headcount per scenario", sec_f)
# header row 20 (B..M)  -- core table headers
setc("A20", "3.  Organisation, weights & headcount", sec_f)
# Use row 20 also as header band on B..M? Keep section title on A20, headers on row 20 cols B-M
setc("B20", "Fixed salary\nAUM < $300m", hdr_f, fill_hdr, ctr, border=True)
setc("C20", "Fixed salary\nAUM ≥ $300m", hdr_f, fill_hdr, ctr, border=True)
setc("D20", "Performance\nweight (per person)", hdr_f, fill_hdr, ctr, border=True)
setc("E20", "", hdr_f, fill_hdr, ctr, border=True)
setc("F20", "", hdr_f, fill_hdr, ctr, border=True)
setc("G20", "Headcount →", hdr_f, fill_hdr, ctr, border=True)
for i, col in enumerate(SCN_COLS):
    setc(f"{col}20", f"@ ${aum[i]:,}m", hdr_f, fill_hdr, ctr, border=True)
ws.row_dimensions[20].height = 30

# ----- role table data: (role, sal_low, sal_high, weight, headcount x6) -----
roles = [
    ("Managing Partner / CIO",            300000, 500000, 100, [1,1,1,1,1,1]),
    ("Partner / Co-Portfolio Manager",    250000, 400000,  60, [1,1,2,2,3,4]),
    ("Portfolio Manager",                 200000, 300000,  40, [0,1,1,2,3,5]),
    ("Head of Research",                  180000, 250000,  30, [0,1,1,1,1,1]),
    ("Senior Analyst",                    150000, 200000,  20, [1,1,2,3,4,6]),
    ("Analyst",                           110000, 140000,  12, [1,2,3,4,6,10]),
    ("Junior Analyst",                     85000, 100000,   6, [0,1,2,3,4,8]),
    ("Head of Trading",                   160000, 220000,  18, [0,0,1,1,1,1]),
    ("Trader / Execution",                120000, 150000,  10, [0,1,1,2,2,3]),
    ("COO / Head of Operations",          170000, 230000,  15, [1,1,1,1,1,1]),
    ("CFO / Finance",                     160000, 210000,  12, [0,1,1,1,1,1]),
    ("Compliance / Legal",                140000, 180000,   8, [0,0,1,1,1,2]),
    ("Investor Relations / Ops Assoc.",   100000, 130000,   5, [1,1,1,2,3,5]),
]
# rows 21..33
for idx, (name, lo, hi, w, hc) in enumerate(roles):
    r = 21 + idx
    fill_row = fill_lighter if idx % 2 else None
    setc(f"A{r}", name, norm, fill_row, left, border=True)
    setc(f"B{r}", lo, inp_f if not fill_row else inp_f, fill_gold, ctr, USD, border=True)
    setc(f"C{r}", hi, inp_f, fill_gold, ctr, USD, border=True)
    setc(f"D{r}", w, inp_f, fill_gold, ctr, NUM, border=True)
    setc(f"E{r}", "", norm, fill_row, None, border=True)
    setc(f"F{r}", "", norm, fill_row, None, border=True)
    setc(f"G{r}", "", norm, fill_row, None, border=True)
    for col, n in zip(SCN_COLS, hc):
        setc(f"{col}{r}", n, inp_f, fill_gold, ctr, NUM, border=True)

# ----- row 34: total performance units -----
setc("A34", "Total performance units", bold, fill_blue, left, border=True)
for c in ["B","C","D","E","F","G"]:
    setc(f"{c}34", "", bold, fill_blue, border=True)
setc("G34", "Σ weight×HC →", bold, fill_blue, right, border=True)
for col in SCN_COLS:
    setc(f"{col}34", f"=SUMPRODUCT($D$21:$D$33,{col}21:{col}33)", Font(size=10,bold=True,color="FFFFFF"), fill_blue, ctr, '#,##0', border=True)
# ----- row 35: USD per performance unit -----
setc("A35", "USD value per performance unit", bold, fill_green, left, border=True)
for c in ["B","C","D","E","F","G"]:
    setc(f"{c}35", "", bold, fill_green, border=True)
setc("G35", "pool ÷ units →", bold, fill_green, right, border=True)
for col in SCN_COLS:
    setc(f"{col}35", f"=IF({col}34=0,0,{col}17*1000000/{col}34)", bold, fill_green, ctr, USD, border=True)

# ----- row 36/37 helper totals -----
setc("A36", "Total fixed payroll (USD)", norm, fill_grey, left, border=True)
for c in ["B","C","D","E","F","G"]:
    setc(f"{c}36", "", norm, fill_grey, border=True)
for col in SCN_COLS:
    setc(f"{col}36", f"=IF({col}13<$F$7,SUMPRODUCT($B$21:$B$33,{col}21:{col}33),SUMPRODUCT($C$21:$C$33,{col}21:{col}33))",
         norm, fill_grey, ctr, USD, border=True)
setc("A37", "Total headcount", norm, fill_grey, left, border=True)
for c in ["B","C","D","E","F","G"]:
    setc(f"{c}37", "", norm, fill_grey, border=True)
for col in SCN_COLS:
    setc(f"{col}37", f"=SUM({col}21:{col}33)", norm, fill_grey, ctr, NUM, border=True)

# ============================================================ OUTPUT A: per-person bonus
def output_block(start_row, title, formula_fn, fmt, head_fill=fill_hdr):
    setc(f"A{start_row}", title, sec_f)
    hr = start_row + 1
    setc(f"A{hr}", "Role", hdr_f, head_fill, left, border=True)
    setc(f"D{hr}", "Weight", hdr_f, head_fill, ctr, border=True)
    for i, col in enumerate(SCN_COLS):
        setc(f"{col}{hr}", f"@ ${aum[i]:,}m", hdr_f, head_fill, ctr, border=True)
    for idx in range(len(roles)):
        r = hr + 1 + idx
        src = 21 + idx
        fr = fill_lighter if idx % 2 else None
        setc(f"A{r}", f"=A{src}", norm, fr, left, border=True)
        setc(f"D{r}", f"=D{src}", norm, fr, ctr, NUM, border=True)
        for col in SCN_COLS:
            setc(f"{col}{r}", formula_fn(col, src, r), norm, fr, ctr, fmt, border=True)
    return hr + 1 + len(roles)

# A) per-person performance payout = weight x USD/unit
endA = output_block(40, "4.  Performance payout PER PERSON, by scenario  (absolute USD — grows with AUM)",
                     lambda col, src, r: f"=$D{src}*{col}$35", USD)

# B) role share of pool % = headcount x weight / total units
startB = endA + 2
endB = output_block(startB, "5.  Each role's SHARE of the performance pool  (% — falls as the firm grows)",
                     lambda col, src, r: f"=IF({col}$34=0,0,{col}{src}*$D{src}/{col}$34)", PCT)

# C) total comp per person = tiered salary + per-person bonus
startC = endB + 2
endC = output_block(startC, "6.  Total annual comp PER PERSON  (fixed salary + performance payout)",
                    lambda col, src, r: f"=IF({col}$13<$F$7,$B{src},$C{src})+$D{src}*{col}$35", USD)

# ============================================================ METHODOLOGY NOTES
nr = endC + 2
setc(f"A{nr}", "How the framework works", sec_f)
notes = [
    "•  Performance fee = AUM × gross return × fee rate (20%). The firm keeps 40%; the remaining 60% (take rate) becomes the comp pool.",
    "•  Every role is assigned a 'performance weight' per person. One weight = one performance unit — one share of the pool.",
    "•  Total performance units = Σ (weight × headcount). USD per unit = comp pool ÷ total units.",
    "•  A person's performance payout = their role weight × USD-per-unit. A role's pool share % = (headcount × weight) ÷ total units.",
    "•  As the firm hires, total units rise, so each role's % share DROPS — but the pool grows with AUM, so the absolute USD per person RISES (see tables 4 vs 5).",
    "•  Fixed salary steps from the '< $300m' column to the '≥ $300m' column once scenario AUM crosses the threshold in F7.",
    "•  Red/gold cells are inputs — edit roles, salaries, weights and headcounts freely; every total recalculates automatically.",
]
for i, t in enumerate(notes):
    rr = nr + 1 + i
    ws.merge_cells(f"A{rr}:M{rr}")
    setc(f"A{rr}", t, norm, fill_grey if i % 2 == 0 else None, left)
    ws.row_dimensions[rr].height = 26

# ============================================================ column widths
ws.column_dimensions["A"].width = 34
for c in ["B", "C", "D"]:
    ws.column_dimensions[c].width = 13
ws.column_dimensions["E"].width = 3
ws.column_dimensions["F"].width = 11
ws.column_dimensions["G"].width = 12
for col in SCN_COLS:
    ws.column_dimensions[col].width = 12
ws.sheet_view.showGridLines = False
ws.freeze_panes = "B21"

wb.save(OUT)
print("Saved:", OUT)
print("Sheets:", wb.sheetnames)
